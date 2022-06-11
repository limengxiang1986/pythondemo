# coding:utf-8
# !/usr/bin/env python
import sys
import threading

from ldap.controls import SimplePagedResultsControl

from rcatrackingconfig import FC_addr_dict,addr_dict1,getjira,JiraRequest,gAdmin, gAuth,teams,jiraRequest1

reload(sys)
sys.setdefaultencoding("utf-8")
from rcatrackingconfig import FC_addr_dict, addr_dict1, getjira, teams, get_login_info
import traceback
import random
import datetime
from flask import Flask, session, request, flash, url_for, redirect, render_template, abort, g, send_from_directory, \
    jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager
from flask_login import login_user, logout_user, current_user, login_required
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, and_, or_
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from flask_login import UserMixin
import xlrd, xlwt
import mysql.connector
import numpy as np
import pandas as pd
from flask import send_file, make_response
from io import BytesIO
import re
import time
import os
from os.path import dirname, abspath
from xlrd import xldate_as_tuple
from jira import JIRA
from jira.client import JIRA
import urllib2
import urllib
import webbrowser

from fuzzywuzzy import fuzz
import requests
from office365.runtime.client_request import ClientRequest
from office365.runtime.utilities.request_options import RequestOptions
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.file import File
from os.path import basename
import shutil
import ldap
import json

from xlutils import copy
from openpyxl import load_workbook
import ast
from collections import OrderedDict
import logging
from logging.handlers import RotatingFileHandler
logger = logging.getLogger('mylogger')
logger.setLevel(logging.DEBUG)

# 创建一个handler，用于写入日志文件
# fh = logging.FileHandler('test.log')
# fh.setLevel(logging.DEBUG)
rh = RotatingFileHandler('myRotatingFileHandler.log',maxBytes=1000,backupCount=5)
rh.setLevel(logging.DEBUG)
rh.setLevel(logging.WARN)

# 再创建一个handler，用于输出到控制台
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)

# 定义handler的输出格式
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# fh.setFormatter(formatter)
rh.setFormatter(formatter)
ch.setFormatter(formatter)

# 给logger添加handler
logger.addHandler(rh)
logger.addHandler(ch)


app = Flask(__name__, template_folder='templates', static_folder='static')
#app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:root@mysql-rca-shark:3306/fddrca?charset=utf8'
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:root@10.157.5.174:31402/fddrca?charset=utf8'
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://lmx:12345678@localhost:3306/blog?charset=utf8'
app.config['SQLALCHEMY_COMMIT_ON_TEARDOWN'] = True
app.config['SQLALCHEMY_ECHO'] = False
app.config['SECRET_KEY'] = 'secret_key'
# app.config['SECRET_KEY'] = 'lslslslsl'
app.config['DEBUG'] = True
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['EXPLAIN_TEMPLATE_LOADING'] = True
app.config['SQLALCHEMY_POOL_RECYCLE'] = 299
app.config['SQLALCHEMY_POOL_TIMEOUT'] = 20

db = SQLAlchemy(app)
#
# jirains = None
#
# connins = None

@app.route('/', methods=['GET', 'POST'])
def home1():
    arr = RcaXtDm.query.order_by(RcaXtDm.dm.asc()).filter().all()
    return render_template('/mainframe.html', todos=arr)


class RcaXtDm(db.Model):
    __tablename__ = "rca_xt_dm"
    type_id = db.Column('type_id', db.String(10), primary_key=True)
    type_name = db.Column('type_name', db.String(100))
    dm = db.Column('dm', db.String(10), primary_key=True)
    csz = db.Column('csz', db.String(200))
    p_type_id = db.Column('p_type_id', db.String(10))
    p_dm = db.Column('p_dm', db.String(10))
    yxbz = db.Column('yxbz', db.String(1))
    kz1 = db.Column('kz1', db.String(45))
    kz2 = db.Column('kz2', db.String(45))
    kz3 = db.Column('kz3', db.String(45))
    kz4 = db.Column('kz4', db.String(45))
    kz5 = db.Column('kz5', db.String(45))
    kz6 = db.Column('kz6', db.String(45))
    kz7 = db.Column('kz7', db.String(45))
    kz8 = db.Column('kz8', db.String(45))
    kz9 = db.Column('kz9', db.String(45))
    kz10 = db.Column('kz10', db.String(45))

    def toJson(self):
        return {"type_id": self.type_id, "type_name": self.type_name,
                "dm": self.dm, "csz": self.csz, "p_type_id": self.p_type_id,
                "p_dm": self.p_dm}



def get_ldap_connectionnew():
    conn = ldap.initialize('ldap://ed-p-gl.emea.nsn-net.net')
    conn.simple_bind_s('cn=BOOTMAN_Acc,ou=SystemUsers,ou=Accounts,o=NSN', 'Eq4ZVLXqMbKbD4th')
    return conn

def get_ldap_connection():
    # conn = ldap.initialize(app.config['LDAP_PROVIDER_URL'])
    try:
        conn = ldap.initialize('ldap://ed-p-gl.emea.nsn-net.net')
        conn.simple_bind_s('cn=BOOTMAN_Acc,ou=SystemUsers,ou=Accounts,o=NSN', 'Eq4ZVLXqMbKbD4th')
    except:
        return redirect(url_for('login'))

    return conn

def getDisplayNameByAccountId(JiraIssueAssignee):
    flg = False
    try:
        conn = get_ldap_connection()
    except:
        flg = True
        return redirect(url_for('login'))
    if flg:
        return ''
    filter = '(mail=%s)' % JiraIssueAssignee
    attrs = ['mail', 'displayName', 'uid']
    base_dn = 'o=NSN'
    flg = False
    try:
        result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    except:
        flg = True
        print 'No such account!!!'
        AssignTo = ''
        return AssignTo
    if flg:
        return ''
    AssignTo = result[0][1]['displayName'][0]
    return AssignTo

def loadUsers():
    cursor = db.session.execute("select username, email from local_user")
    rstjson = []
    result = cursor.fetchall()
    for a in result:
        rstjson.append({"username": a[0], "email": a[1]})
    return rstjson

class SystemAdmin(db.Model):
    __tablename__ = 'system_admin'
    adminid = db.Column('adminid', db.Integer, primary_key=True)
    adminname = db.Column(db.String(100))
    email = db.Column(db.String(100))
    yxbz = db.Column(db.String(1))
    def __init__(self, adminname,email,yxbz = 'Y'):
        self.adminname = adminname
        self.email = email
        self.yxbz = yxbz

# JiraUserName = 'Ca_AutomationAccount'
# JiraPassWord= 'jira123'
JiraUserName = 'rcashark'
# JiraPassWord= 'Jira1234'
JiraPassWord= 'Jira4321'
def getjira():
    options = {'server': 'https://jiradc.int.net.nokia.com/'}
    options = {'server': 'https://jiradc.ext.net.nokia.com/'}
    _conn_status = True
    _conn_retry_count = 0
    while _conn_status:
        try:
            _conn_status = False
            print 'Connecting JIRA...'
            jira = JIRA(options, basic_auth=(JiraUserName, JiraPassWord))
            # return jira
        except:
            _conn_retry_count += 1
            print ("_conn_retry_count=%d" % _conn_retry_count)
            _conn_status = True
            print 'ProntoDb connecting Error'
            time.sleep(10)
        if _conn_status:
            continue
        else:
            print 'Connecting JIRA...OK'
            return jira

def leap_year(y):
    if (y % 4 == 0 and y % 100 != 0) or y % 400 == 0:
        return True
    else:
        return False

def days_in_month(y, m):
    if m in [1, 3, 5, 7, 8, 10, 12]:
        return 31
    elif m in [4, 6, 9, 11]:
        return 30
    else:
        if leap_year(y):
            return 29
        else:
            return 28

def days_this_year(year):
    if leap_year(year):
        return 366
    else:
        return 365

def days_passed(year, month, day):
    m = 1
    days = 0
    while m < month:
        days += days_in_month(year, m)
        m += 1
    return days + day

def daysBetweenDates(year1, month1, day1, year2, month2, day2):
    if year1 == year2:
        return days_passed(year2, month2, day2) - days_passed(year1, month1, day1)
    if year1 < year2:
        sum1 = 0
        y1 = year1
        while y1 < year2:
            sum1 += days_this_year(y1)
            y1 += 1
        daysYear1 = days_passed(year1,month1,day1)
        daysYear2 = days_passed(year2,month2,day2)
        daysPassed = sum1 - daysYear1 + daysYear2
        return daysPassed
    if year1 > year2:
        sum1 = 0
        y1 = year1
        while y1 > year2:
            sum1 -= days_this_year(y1)
            y1 -= 1
        daysYear1 = days_passed(year1,month1,day1)
        daysYear2 = days_passed(year2,month2,day2)
        daysPassed = sum1 - daysYear1 + daysYear2
        # return sum1-days_passed(year1,month1,day1)+days_passed(year2,month2,day2)
        return daysPassed

def daysBetweenDate(start,end):
    year1=int(start.split('-',2)[0])
    month1=int(start.split('-',2)[1])
    day1=int(start.split('-',2)[2])

    year2=int(end.split('-',2)[0])
    month2=int(end.split('-',2)[1])
    day2=int(end.split('-',2)[2])
    # print ("daysBetweenDates(year1, month1, day1, year2, month2, day2)=%d"%daysBetweenDates(year1, month1, day1, year2, month2, day2))
    return daysBetweenDates(year1, month1, day1, year2, month2, day2)


class SystemLog(db.Model):
    __tablename__ = 'syslog'
    id = db.Column('Index', db.Integer, primary_key=True)
    loginAccount = db.Column(db.String(64), default='')
    loginLocation = db.Column(db.String(128), default='')
    ipAddress = db.Column(db.String(128), default='')
    browserType = db.Column(db.String(128), default='')
    deviceType = db.Column(db.String(128), default='')
    osType = db.Column(db.String(128), default='')
    operationType = db.Column(db.String(128), default='')
    logTime = db.Column(db.DateTime)
    prIdorApId = db.Column(db.String(128), default='')
    log1 = db.Column(db.String(1024), default='')
    log2 = db.Column(db.String(1024), default='')
    log3 = db.Column(db.String(1024), default='')
    log4 = db.Column(db.String(1024), default='')
    log5 = db.Column(db.String(1024), default='')
    log6 = db.Column(db.String(1024), default='')
    log7 = db.Column(db.String(1024), default='')
    log8 = db.Column(db.String(1024), default='')
    log9 = db.Column(db.String(1024), default='')
    log10 = db.Column(db.String(1024), default='')
    log11 = db.Column(db.String(1024), default='')
    log12 = db.Column(db.String(1024), default='')

    def __init__(self, loginAccount, loginLocation, ipAddress, browserType, deviceType, osType, operationType,
                 prIdorApId, \
                 log1, log2, log3, log4, log5, log6, log7, log8, log9, log10, log11, log12):
        self.loginAccount = loginAccount
        self.loginLocation = loginLocation
        self.ipAddress = ipAddress
        self.browserType = browserType
        self.deviceType = deviceType
        self.osType = osType
        self.operationType = operationType
        self.logTime = datetime.datetime.now()
        self.prIdorApId = prIdorApId
        self.log1 = log1
        self.log2 = log2
        self.log3 = log3
        self.log4 = log4
        self.log5 = log5
        self.log6 = log6
        self.log7 = log7
        self.log8 = log8
        self.log9 = log9
        self.log10 = log10
        self.log11 = log11
        self.log12 = log12




def syslog_error(module, errormes,qkms='',systemname='rcashark-container',loginAccount='SYSTEMLOG', loginLocation='NOKIA', operationType='error',
                 user='', username=''):
    try:
        user = ''
        username = ''
        ipAddress=''
        browserType=''
        deviceType=''
        osType=''
        try:
            log_dict = get_login_info()
            ipAddress = log_dict['IP']
            browserType = log_dict['BR']
            deviceType = log_dict['DEV']
            osType = log_dict['OS']
        except BaseException as eu :
            print(eu.message + str(traceback.format_exc()))
        loginfo = SystemLog(loginAccount=loginAccount, loginLocation=loginLocation, ipAddress=ipAddress, browserType=browserType, deviceType=deviceType, osType=osType, operationType=operationType,
                            prIdorApId='', \
                            log1=systemname, log2=module, log3=errormes[0:2999], log4=qkms[0:999], log5=user, log6=username, log7='', log8='', log9='', log10='', log11='', log12='')
        db.session.add(loginfo)
        db.session.commit()
    except BaseException as e1 :
        print(e1)
        pass


class MnTodo(db.Model):
    __tablename__ = 'mnrcatable'
    # JiraIssueId = db.Column('JiraIssueId', db.String(64), primary_key=True)
    PRID = db.Column('PRID', db.String(64), primary_key=True)
    PRTitle = db.Column(db.String(1024))
    PRReportedDate = db.Column(db.String(64))
    PRClosedDate = db.Column(db.String(64))
    PRRelease = db.Column(db.String(128))
    PRAttached = db.Column(db.String(512))
    PRSeverity = db.Column(db.String(32))
    PRGroupInCharge = db.Column(db.String(64))
    PRRcaEdaAssessor = db.Column(db.String(128))
    PRProduct = db.Column(db.String(64))
    ReportedBy = db.Column(db.String(64))
    FaultCoordinator = db.Column(db.String(64))
    CustomerName = db.Column(db.String(128))
    BusinessUnit = db.Column(db.String(128))  # Parent task has, Child task does not have
    BusinessLine = db.Column(db.String(128))  # Parent task has, Child task does not have
    ProductLine = db.Column(db.String(128))  # Parent task has, Child task does not have
    Feature = db.Column(db.String(128))
    RCAEDACategory = db.Column(db.String(128))
    JiraCustomerName = db.Column(db.String(128))
    JIRAProject = db.Column(db.String(128))
    FaRcaEdaDecisionFlag = db.Column(db.String(32))
    EDA_AssessorEmail = db.Column(db.String(128))
    EDACreatingTime = db.Column(db.String(128))
    EdaCaseType = db.Column(db.String(128))
    JiraTaskIsCreated = db.Column(db.String(32))

    user_id = db.Column(db.Integer, db.ForeignKey('jirausers.user_id'))

    def __init__(self, PRID, PRTitle, PRReportedDate, PRClosedDate, PRRelease, PRSeverity, PRGroupInCharge, PRAttached,
                 PRRcaEdaAssessor, \
                 PRProduct, ReportedBy, FaultCoordinator, CustomerName, BusinessUnit, BusinessLine, ProductLine,
                 Feature, RCAEDACategory, \
                 JiraCustomerName, JIRAProject, FaRcaEdaDecisionFlag, EDA_AssessorEmail, EDACreatingTime, EdaCaseType,
                 JiraTaskIsCreated):
        self.PRID = PRID
        self.PRTitle = PRTitle
        self.PRReportedDate = PRReportedDate
        self.PRClosedDate = PRClosedDate
        self.PRRcaEdaAssessor = PRRcaEdaAssessor
        self.PRRelease = PRRelease
        self.PRAttached = PRAttached
        self.PRSeverity = PRSeverity
        self.PRGroupInCharge = PRGroupInCharge
        self.PRProduct = PRProduct
        self.ReportedBy = ReportedBy
        self.FaultCoordinator = FaultCoordinator
        self.CustomerName = CustomerName
        self.BusinessUnit = BusinessUnit
        self.BusinessLine = BusinessLine
        self.ProductLine = ProductLine
        self.Feature = Feature
        self.RCAEDACategory = RCAEDACategory
        self.JiraCustomerName = JiraCustomerName
        self.JIRAProject = JIRAProject
        self.FaRcaEdaDecisionFlag = FaRcaEdaDecisionFlag
        self.EDA_AssessorEmail = EDA_AssessorEmail
        self.EDACreatingTime = EDACreatingTime
        self.EdaCaseType = EdaCaseType
        self.JiraTaskIsCreated = JiraTaskIsCreated


class MnRcaMetrics(db.Model):
    __tablename__ = 'mnrcametricstable'
    JiraIssueId = db.Column('JiraIssueId', db.String(64), primary_key=True)
    JiraIssueParentTaskId = db.Column(db.String(64))
    JiraIssueSummary = db.Column(db.String(512))
    JiraIssueType = db.Column(db.String(64))
    JiraIssueCaseType = db.Column(db.String(64))
    JiraIssueStatus = db.Column(db.String(64))
    JiraIssueParentStatus = db.Column(db.String(64))
    JiraIssueAssignee = db.Column(db.String(128))
    JiraIssueReporter = db.Column(db.String(128))
    PRID = db.Column(db.String(64))
    FAID = db.Column(db.String(64))
    PRTitle = db.Column(db.String(1024))
    PRRcaEdaAssessor = db.Column(db.String(128))
    PRAttached = db.Column(db.String(512))
    PRSeverity = db.Column(db.String(32))
    PRGroupInCharge = db.Column(db.String(64))
    JiraIssueCreatedDate = db.Column(db.String(64))
    JiraIssueParentCreatedDate = db.Column(db.String(64))
    JiraIssueResolutionDate = db.Column(db.String(64))
    JiraIssueParentResolutionDate = db.Column(db.String(64))
    JiraIssueDueDate = db.Column(db.String(64))
    JiraIssueOpenDays = db.Column(db.Integer)
    JiraIssueParentOpenDays = db.Column(db.Integer)
    JiraIssueAssigneeSquadGroup = db.Column(db.String(64))
    JiraIssueAssigneeSquadGroupLead = db.Column(db.String(64))
    JiraIssueAssigneeTribe = db.Column(db.String(64))
    JiraIssueAssigneeTribeLead = db.Column(db.String(64))
    JiraIssueOverDue = db.Column(db.String(32))
    JiraIssueParentOverDue = db.Column(db.String(32))
    JiraIssueOverDueReason = db.Column(db.String(1024))
    JiraIssueParentOverDueReason = db.Column(db.String(1024))
    ReportedBy = db.Column(db.String(64))

    def __init__(self, JiraIssueId,JiraIssueParentTaskId,JiraIssueSummary,JiraIssueType,JiraIssueCaseType,\
                 JiraIssueStatus,JiraIssueParentStatus,JiraIssueAssignee,JiraIssueReporter,PRID,FAID,PRTitle,\
                 PRRcaEdaAssessor,PRAttached,PRSeverity,PRGroupInCharge,JiraIssueCreatedDate,JiraIssueParentCreatedDate,\
                 JiraIssueResolutionDate,JiraIssueParentResolutionDate,JiraIssueDueDate,JiraIssueOpenDays,\
                 JiraIssueParentOpenDays,JiraIssueAssigneeSquadGroup,JiraIssueAssigneeSquadGroupLead,\
                 JiraIssueAssigneeTribe,JiraIssueAssigneeTribeLead,JiraIssueOverDue,JiraIssueParentOverDue,\
                 JiraIssueOverDueReason,JiraIssueParentOverDueReason,ReportedBy):

        self.JiraIssueId = JiraIssueId
        self.JiraIssueParentTaskId = JiraIssueParentTaskId
        self.JiraIssueSummary = JiraIssueSummary
        self.JiraIssueType = JiraIssueType
        self.JiraIssueCaseType = JiraIssueCaseType
        self.JiraIssueStatus = JiraIssueStatus
        self.JiraIssueParentStatus = JiraIssueParentStatus
        self.JiraIssueAssignee = JiraIssueAssignee
        self.JiraIssueReporter = JiraIssueReporter
        self.PRID = PRID
        self.FAID = FAID
        self.PRTitle = PRTitle
        self.PRRcaEdaAssessor = PRRcaEdaAssessor
        self.PRAttached = PRAttached
        self.PRSeverity = PRSeverity
        self.PRGroupInCharge = PRGroupInCharge
        self.JiraIssueCreatedDate = JiraIssueCreatedDate
        self.JiraIssueParentCreatedDate = JiraIssueParentCreatedDate
        self.JiraIssueResolutionDate = JiraIssueResolutionDate
        self.JiraIssueParentResolutionDate = JiraIssueParentResolutionDate
        self.JiraIssueDueDate = JiraIssueDueDate
        self.JiraIssueOpenDays = JiraIssueOpenDays
        self.JiraIssueParentOpenDays = JiraIssueParentOpenDays
        self.JiraIssueAssigneeSquadGroup = JiraIssueAssigneeSquadGroup
        self.JiraIssueAssigneeSquadGroupLead = JiraIssueAssigneeSquadGroupLead
        self.JiraIssueAssigneeTribe = JiraIssueAssigneeTribe
        self.JiraIssueAssigneeTribeLead = JiraIssueAssigneeTribeLead
        self.JiraIssueOverDue = JiraIssueOverDue
        self.JiraIssueParentOverDue = JiraIssueParentOverDue
        self.JiraIssueOverDueReason = JiraIssueOverDueReason
        self.JiraIssueParentOverDueReason = JiraIssueParentOverDueReason
        self.ReportedBy = ReportedBy


class Rca5Why(db.Model):
    __tablename__ = 'rca5why'
    id = db.Column('why_id', db.Integer, primary_key=True)
    PRID=db.Column(db.String(64))
    Why1 = db.Column(db.String(1024))
    Why2 = db.Column(db.String(1024))
    Why3=db.Column(db.String(1024))
    Why4 = db.Column(db.String(1024))
    Why5 = db.Column(db.String(1024))
    pr_id = db.Column(db.String(64), db.ForeignKey('rcastatus.PRID'))

    def __init__(self, PRID,Why1,Why2,Why3,Why4,Why5):
        self.PRID = PRID
        self.Why1 = Why1
        self.Why2 = Why2
        self.Why3 = Why3
        self.Why4 = Why4
        self.Why5 = Why5
        #pr_id = db.Column(db.String(64), db.ForeignKey('rcastatus.PRID'))

def testSyslog():
    jira = getjira()
    issues = jira.search_issues(jql_str='project = MNPRCA AND type in \
        ("Analysis subtask","Action for RCA","Action for EDA")\
         and created >= -30d', maxResults=False)
    # issues = jira.search_issues(jql_str='project = MNPRCA  and key ="MNRCA-48598"  AND type in \
    # ("Analysis subtask","Action for RCA","Action for EDA","RCA Subtask","EDA subtask")\
    #  and created >= "-30d"',maxResults=False)
    counting = 0
    # conn = get_ldap_connection()
    sizenum = 0
    errornum = 0
    successrnum = 0
    for issue in issues:  # 3861
        try:
            JiraIssueId = str(issue.key)
            print 'JiraIssueId = %s' % JiraIssueId
            issue = jira.issue(JiraIssueId)
            parentExistFlag = True
            try:
                Rca5WhyParentIssue = issue.fields.parent
            except:
                parentExistFlag = False
                pass
            if parentExistFlag == False:
                JiraIssueParentTaskId = 'None'
                JiraIssueSummary = 'Exception: No Parent Task exist!!!'
                PRAttached = ''
                continue
            else:
                JiraIssueParentTaskId = str(Rca5WhyParentIssue.key)
                Rca5WhyParentIssue = jira.issue(JiraIssueParentTaskId)
                if Rca5WhyParentIssue.fields.issuetype.name == '5WhyRCA':
                    PRAttached = Rca5WhyParentIssue.fields.customfield_37060
                else:
                    continue
            JiraIssueType = issue.fields.issuetype.name
            JiraIssueCaseType = ''
            if JiraIssueType in ('Analysis subtask',):
                JiraIssueCaseType = str(issue.fields.customfield_10464)
            JiraIssueStatus = issue.fields.status.name
            JiraIssueParentStatus = Rca5WhyParentIssue.fields.status.name
            JiraIssueAssignee = str(issue.fields.assignee)
            JiraIssueReporter = issue.fields.reporter.displayName
            # issue = jira.issue(JiraIssueId)
            # Rca5WhyParentIssue = jira.issue(JiraIssueParentTaskId)
            # PRAttached = Rca5WhyParentIssue.fields.customfield_37060
            createDate = issue.fields.created
            JiraIssueCreatedDate = createDate.split('T')[0]
            createDate = Rca5WhyParentIssue.fields.created
            JiraIssueParentCreatedDate = createDate.split('T')[0]
            resolutiondate = str(Rca5WhyParentIssue.fields.resolutiondate)
            JiraIssueParentResolutionDate = resolutiondate.split('T')[0]
            JiraIssueParentOpenDays = daysBetweenDate(JiraIssueParentCreatedDate, JiraIssueParentResolutionDate)
            successrnum += 1
        except Exception as e1:
            errornum += 1
            errmes = 'error: jiraIssueId=' + JiraIssueParentTaskId + ', ' + str(e1) + ''
            print(errmes)
            syslog_error('self_test_case',errormes=errmes)
        sizenum += 1
        print(' sucess:' + str(successrnum) + ', error:' + str(errornum) +', total:'+str(sizenum))

def testDiffFromJiraAndDB():
    jira = getjira()
    issues = jira.search_issues(jql_str='project = MNPRCA AND type in \
    ("Analysis subtask","Action for RCA","Action for EDA")\
     and created >= -10d',maxResults=False)
    counting =0
    totalnum = len(issues)
    existsnum =0
    lacknum = 0
    for issue in issues: #3861
        try:
            JiraIssueId = str(issue.key)
            issue = jira.issue(JiraIssueId)
            parentExistFlag = True
            try:
                Rca5WhyParentIssue = issue.fields.parent
            except:
                parentExistFlag = False
                pass
            if parentExistFlag == False:
                continue
            else:
                JiraIssueParentTaskId = str(Rca5WhyParentIssue.key)
                Rca5WhyParentIssue = jira.issue(JiraIssueParentTaskId)
                if Rca5WhyParentIssue.fields.issuetype.name == '5WhyRCA':
                    PRAttached = Rca5WhyParentIssue.fields.customfield_37060
                else:
                    continue
            createDate = issue.fields.created
            JiraIssueParentCreatedDate = createDate.split('T')[0]

            if JiraIssueParentCreatedDate < '2019-01-01':
                continue
            todo_item = MnRcaMetrics.query.get(JiraIssueId)
            if todo_item:
                existsnum +=1
                pass
            else:
                lacknum += 1
            print('totalnum:' + str(totalnum) + ',existsnum:'+str(existsnum) + ',lack num :' + str(lacknum))
        except Exception as e1:
            pass
            print('totalnum:' + str(totalnum) + ',existsnum:'+str(existsnum) + ',lack num :' + str(lacknum))

    s2 = datetime.datetime.now().strftime('%H:%M')
    print('totalnum:' + str(totalnum) + ',existsnum:'+str(existsnum) + ',lack num :' + str(lacknum))

PriorityDict = {
    "1": 'Blocker',
    "2": 'Critical',
    "3": 'Major',
    "4": 'Minor',
    "5": 'Trivial',
    "6": 'Documentary',
    "10000": 'Unknown',
    "10001": 'Medium',
    "10002": 'Normal',
    "10100": 'High',
    "10101": 'Low',
}
SourceListDict = {
    "": 'None',
    "206960": 'NCDR',
    "206961": 'DQR deep dive.CES (VoC)',
    "206962": 'Internal R & D improvement',
    "206963": 'Lessons learned',
}
BusinessUnitDict = {
    "-1": 'None',
    "189744": 'Mobile Networks Products',
    "189745": 'Converged Core',
    "189746": 'Advanced Mobile Networks Solutions (AMS)',
    "189747": 'MN Service',
}

BusinessLineDict = {
    "-1": 'None',
    "205171": 'LTE',
    "205172": 'HetRAN',
    "205173": '5G',
    "205174": 'RF&AA',
    "205175": 'BBP',
    "205176": 'RMA',
    "205177": 'Customer Support',
    "205178": 'Architecture',
    "205179": 'Telco Cloud',
    "205180": '3G Core',
    "205181": 'Cloud SDM',
    "205182": 'Cloud IMS',
    "205183": 'Customer and Solution',
    "205184": 'Small Cell & WiFi Solutions',
    "205185": 'Public Sector, Complementary Solutions and Partner',
    "205187": 'X-Haul',
    "205188": 'NA',
    "205189": 'WVS',
    "219023": 'SRAN',
    "280313": 'Airphone',
    "276424": 'CDS',
    "280311": 'ECP',
    "280310": 'SoC',
    "280312": 'TRS',
}

ProductLineDict = {
    "-1":'None',
    "179991":'AED',
    "179992":'BBP HW',
    "179993":'BBP OPSW',
    "179995":'BTS Site Manager',
    "179996":'CFX',
    "179997":'Cloud BTS',
    "179998":'CuDo',
    "179999":'DCM',
    "180000":'FDD LTE',
    "180001":'Femto',
    "180002":'Flexi Platform SW',
    "180003":'Flexi Zone',
    "180004":'GSM BTS',
    "180005":'HetRAN BTS',
    "180006":'HSS/vHSS',
    "180007":'HW',
    "180008":'MEC',
    "180009":'Metro',
    "180010":'MGW/OSP',
    "180011":'MSS',
    "180012":'N/A',
    "180013":'NCIO',
    "180014":'NDCS',
    "180015":'NLS',
    "180016":'NT-HLR/vHLR',
    "180017":'One NDS',
    "180018":'One-EIR ',
    "180019":'One-MNP',
    "180020":'oTAS',
    "180021":'PMOD - VGP',
    "180022":'RFHW',
    "180023":'RFSW',
    "180024":'RNC, BSC, RCP, vRNC',
    "180025":'SDL',
    "180026":'SoC',
    "180027":'SRAN/SBTS',
    "180028":'TD LTE',
    "180029":'WBTS',
    "180030":'WiFi/ZC',
    "180031":'GSM_BTS',
    "180032":'5G Application',
    "180033":'5G L1',
    "180034":'nTAS',
    "280316":'Airphone',
    "280314":'ECP',
    "280315":'TRS',
}


CustomerNameDict = {
    "-1": 'None',
    "206299": 'ADVANCED WIRELESS NETWORKS CO., LTD',
    "206300": 'AIS, Thailand',
    "206301": 'America Movil Head Office, Mexico, incl.AMX Colombia',
    "206302": 'AMX',
    "206303": 'AMX(Claro), Argentina',
    "206304": 'AMX(Claro), Chile',
    "206305": 'AT&T, USA',
    "206306": 'Avea',
    "206307": 'Bell Canada',
    "206308": 'BELL MOBILITY INC.',
    "206309": 'BHARTI',
    "206310": 'Bharti, India',
    "206311": 'BSNLMTNL',
    "206312": 'BT, EE & MBNL, UK',
    "206313": 'CA',
    "206314": 'CEWA',
    "206315": 'China Mobile, China',
    "206316": 'China Telecom, China',
    "206317": 'China Unicom, China',
    "206318": 'Chunghwa Telecom, Taiwan',
    "206319": 'DT Croatia',
    "206320": 'DT, Germany',
    "206321": 'Elisa',
    "206322": 'ERICSSON TELECOMUNICAZIONI S.P.A.WIND IT',
    "206323": 'EVERYTHING EVERYWHERE LIMITED(ORANGE PCS)',
    "206324": 'France Telecom',
    "206325": 'Free, France',
    "206326": 'HBTSCIF',
    "206327": 'HBTSSB',
    "206328": 'Hutchison 3G UK',
    "206329": 'Hutchison IRL',
    "206330": 'ICE Communication Norge AS',
    "206331": 'IDEA',
    "206332": 'IDEA CELLULAR LTD - KERELA',
    "206333": 'Idea, India',
    "206334": 'Indosat, Indonesia',
    "206335": 'IRANCELL',
    "206336": 'KDDI, Japan',
    "206337": 'Korea Telecom, Korea',
    "206338": 'KT CORPORATION',
    "206339": 'KT DJ Korea',
    "206340": 'KT Korea',
    "206341": 'LG U +, Korea',
    "206342": 'LOCOP',
    "206343": 'MBNL',
    "206344": 'Megafon, Russia',
    "206345": 'MOBILE TELECOMMUNICATIONS, SAUDI',
    "206346": 'MobileOne Ltd.',
    "206347": 'Mobily, Saudi Arabia',
    "206348": 'MTS, Russia',
    "206349": 'MUV',
    "206350": 'NA',
    "206351": 'NTT DoCoMo, Japan',
    "206352": 'O2 CZECH REPUBLIC',
    "206353": 'O2 GBR',
    "206354": 'O2, HU',
    "206355": 'OI',
    "206356": 'Oi, Brazil',
    "206357": 'Ooredoo Tunisie S.A.',
    "206358": 'Ooredoo, Myanmar',
    "206359": 'Ooredoo, Qatar',
    "206360": 'Optus, Australia',
    "206361": 'ORANGE',
    "206362": 'Orange, France',
    "206363": 'PLDT Smart, Philipines',
    "206364": 'PROXIMUS, BE',
    "206365": 'PT Telkomsel',
    "206366": 'Q-TEL',
    "206367": 'SAFARICOM',
    "206368": 'SAV',
    "206369": 'SFR, France',
    "206370": 'SK telecom, Inc',
    "206371": 'SK Telecom, Korea',
    "206372": 'SKIL',
    "206373": 'Softbank (incl. WCP, Y-Mobile), Japan',
    "206374": 'Sprint, USA',
    "206375": 'STC, SaudiArabia',
    "206376": 'T COM DEU',
    "206377": 'T-Mobile, Deutschland',
    "206378": 'T-Mobile Deutschland',
    "206379": 'T-MOBILE POLSKA S.A.',
    "206380": 'T-Mobile, USA',
    "206381": 'Taiwan Mobile, Taiwan',
    "206382": 'TATA',
    "206383": 'TELE2 NLD',
    "206384": 'Tele2, Russia',
    "206385": 'Telecom(TIM), Italia',
    "206386": 'TELECOM ITALIA S.P.A.(MOBILE)',
    "206387": 'Telefonica O2, UK',
    "206388": 'Telefonica, Germany',
    "206389": 'Telefonica, Spain',
    "206390": 'TELIA MOBILE AB, SE',
    "206391": 'Telkomsel, Indonesia',
    "206392": 'TF',
    "206393": 'THM',
    "206394": 'TIM',
    "206395": 'TIM, Brazil',
    "206397": 'UNITED STATES CELLULAR CORP.',
    "206398": 'US MA',
    "206399": 'VERIZON',
    "206400": 'Verizon, USA',
    "206401": 'VF',
    "206402": 'Viettel, Vietnam',
    "206403": 'Vimpelcom, Russia',
    "206404": 'Vodacom(Pty.) Ltd.',
    "206405": 'Vodacome, South Africa(RSA)',
    "206406": 'Vodafone - Chennai',
    "206407": 'VODAFONE EGYPT TEL',
    "206408": 'Vodafone, India',
    "206409": 'Vodafone, Italy',
    "206410": 'VTG AFR TZ(VNM)',
    "206411": 'WE',
    "206412": 'Wind, Canada',
    "206413": 'Zain Saudi Arabia',
    "206414": 'Vodafone',
    "206415": 'CUC',
    "206416": 'Bharti Airtel Limited',
    "206417": 'MTS UA',
    "206418": 'T-MOBILE USA',
    "206419": 'SFR Siege',
    "206420": 'Optus',
    "206421": 'SOFTBANK MOBILE Corp.',
    "206422": 'Telefonica Moviles de Espana',
    "206423": 'Vodafone UK Limited',
    "206424": 'SoftBank Corp.',
    "206425": 'Optus Networks Pty Ltd',
    "206426": 'ORANGE SKILL CENTER',
    "206427": 'Tt-Netvaerket P/S',
    "206428": 'IDEA CELLULAR LTD - UPEAST',
    "206429": 'Free Mobile',
    "213306": 'CMCC Zhejiang',
    "213307": 'CMCC Fujian',
    "213308": 'CMCC Hunan',
    "213309": 'CUC Shanghai',
}

CuDoBusinessLineDict = {
"-1": 'None',
"205152": 'LTE',
"205153": 'HetRAN',
"205154": '5G',
"205155": 'RF&AA',
"205156": 'BBP',
"205157": 'RMA',
"205158": 'Customer Support',
"205159": 'Architecture',
"205160": 'Telco Cloud ',
"205161": '3G Core',
"205162": 'Cloud SDM',
"205163": 'Cloud IMS ',
"205164": 'Customer and Solution',
"205165": 'Small Cell & WiFi Solutions',
"205166": 'IoT',
"205167": 'Public Sector, Complementary Solutions and Partner',
"205168": 'X-Haul',
"205169": 'NA',
}
CuDoProductLineDict ={
"-1": 'None',
"180035": 'FDD LTE',
"180036": 'TD LTE',
"180037": 'RNC, BSC, RCP, vRNC',
"180038": 'WBTS',
"180039": 'GSM BTS',
"180040": 'SRAN/SBTS',
"180041": 'N/A',
"180042": 'RFHW',
"180043": 'RFSW',
"180044": 'BBP HW',
"180045": 'BBP Site Manager',
"180046": 'BBP OPSW',
"180047": 'CuDo',
"180048": 'SoC',
"180049": 'MGW/OSP',
"180050": 'MSS',
"180051": 'Flexi Platform SW',
"180052": 'HW',
"180053": 'One NDS',
"180054": 'NT-HLR/vHLR',
"180055": 'HSS/vHSS',
"180056": 'CFX',
"180057": 'oTAS',
"180058": 'NLS',
"180059": 'PMOD - VGP',
"180060": 'Flexi Zone',
"180061": 'WiFi/ZC',
"180062": 'Femto',
"180063": 'Metro',
"180064": 'SDL',
"180065": 'NDCS',
"180066": 'DCM',
"180067": 'NCIO',
"180068": 'One-EIR',
"180069": 'One-MNP',

}

CuDoCustomerNameDict = {
"-1": 'None',
"206430": 'America Movil Head Office, Mexico, incl. AMX Colombia',
"206431": 'AMX',
"206432": 'AT&T Mobility',
"206433": 'AT&T, USA',
"206434": 'Bell Canada',
"206435": 'Bell Mobility Inc.',
"206436": 'BHARTI',
"206437": 'Bharti, India',
"206438": 'BSNLMTNL',
"206439": 'BT, EE & MBNL, UK',
"206440": 'CA',
"206441": 'CEWA',
"206442": 'China Mobile, China',
"206443": 'China Telecom, China',
"206444": 'China Unicom, China',
"206445": 'Chunghwa Telecom, Taiwan',
"206446": 'Claro Argentina',
"206447": 'CT Jiangsu',
"206448": 'DT, Germany',
"206449": 'Egyptian Company for Mobile Service',
"206450": 'Everything Everywhere Limited (Orange PCS)',
"206451": 'France Telecom',
"206452": 'IDEA',
"206453": 'Idea, India',
"206454": 'Indosat, Indonesia',
"206455": 'KDDI, Japan',
"206456": 'LG U+, Korea',
"206457": 'LOCOP',
"206458": 'Megafon, Russia',
"206459": 'Mobily, Saudi Arabia',
"206460": 'MTS, Russia',
"206461": 'MUV',
"206462": 'NA',
"206463": 'NTT DoCoMo, Japan',
"206464": 'OI',
"206465": 'Oi, Brazil',
"206466": 'Optus Networks Pty Ltd',
"206467": 'Optus, Australia',
"206468": 'ORANGE',
"206469": 'Orange, France',
"206470": 'SAV',
"206471": 'SFR, France',
"206472": 'SKIL',
"206473": 'Softbank (incl. WCP, Y-Mobile), Japan',
"206474": 'SoftBank Corp.',
"206475": 'Sprint, USA',
"206476": 'STC, Saudi Arabia',
"206477": 'T-Mobile Deutschland',
"206478": 'T-Mobile, USA',
"206479": 'Taiwan Mobile Co Ltd',
"206480": 'Taiwan Mobile, Taiwan',
"206481": 'Taiwan Star Telecom Co., Ltd',
"206482": 'TATA',
"206483": 'Telefonica, Germany',
"206484": 'Telefonica, Spain',
"206485": 'Telkomsel, Indonesia',
"206486": 'TF',
"206487": 'THM',
"206488": 'TIM',
"206489": 'TMO US',
"206490": 'US MA',
"206491": 'VERIZON',
"206492": 'Verizon, USA',
"206493": 'VF',
"206494": 'Vodafone - UP East',
"206495": 'Vodafone Omnitel N.V. (Vodafone Italy)',
"206496": 'Vodafone, India',
"206497": 'Vodafone, Italy',
"206498": 'WE',

}


def tools_strip(str):
    if str == None:
        return ""
    return str.strip(",").strip(" ")

def createParent5WhyRcaTaskFirst(jira, reporter, request, r):
    a = r
    PRAttached = a['problemReportIds']
    if PRAttached:
        LabelItems = PRAttached
        b = ''
        for s in LabelItems:
            b = b + s + ' '
            PRAttached = tools_strip(b)
    else:
        PRAttached = ''
    # PRAttached = ''
    # for s in a['problemReportIds']:
    #     PRAttached = PRAttached + s
    JiraIssuePriority = request.form['Priority'].strip()  # Parent task has, Child task does not have
    if JiraIssuePriority in PriorityDict.values():
        new_dict = {v: k for k, v in PriorityDict.items()}
        JiraIssuePriority = new_dict[JiraIssuePriority]
    JiraIssueSourceList = request.form['SourceList'].strip()  # Parent task has, Child task does not have
    if JiraIssueSourceList in SourceListDict.values():
        new_dict = {v: k for k, v in SourceListDict.items()}
        JiraIssueSourceList = new_dict[JiraIssueSourceList]
    JiraIssueBusinessUnit = request.form['BusinessUnit'].strip()  # Parent task has, Child task does not have
    if JiraIssueBusinessUnit in BusinessUnitDict.values():
        new_dict = {v: k for k, v in BusinessUnitDict.items()}
        JiraIssueBusinessUnit = new_dict[JiraIssueBusinessUnit]
    JiraIssueBusinessLine = request.form['BusinessLine'].strip()  # Parent task has, Child task does not have
    if JiraIssueBusinessLine in BusinessLineDict.values():
        new_dict = {v: k for k, v in BusinessLineDict.items()}
        JiraIssueBusinessLine = new_dict[JiraIssueBusinessLine]
    JiraIssueProductLine = request.form['ProductLine'].strip()  # Parent task has, Child task does not have
    if JiraIssueProductLine in ProductLineDict.values():
        new_dict = {v: k for k, v in ProductLineDict.items()}
        JiraIssueProductLine = new_dict[JiraIssueProductLine]
    else:
        PRProduct = a['product']
        JiraIssueProductLine = '123'
        new_dict = {v: k for k, v in ProductLineDict.items()}
        JiraIssueProductLine = new_dict[JiraIssueProductLine]
    # JiraIssueSourceList = request.form['SourceList'].strip()  # Parent task has, Child task does not have
    # JiraIssueBusinessUnit = request.form['BusinessUnit'].strip()  # Parent task has, Child task does not have
    # JiraIssueBusinessLine = request.form['BusinessLine'].strip()  # Parent task has, Child task does not have
    # JiraIssueProductLine = request.form['ProductLine'].strip()  # Parent task has, Child task does not have
    JiraIssueCustomerName = request.form['Customer'].strip()  # Parent task has, Child task does not have
    if JiraIssueCustomerName in CustomerNameDict.values():
        new_dict = {v: k for k, v in CustomerNameDict.items()}
        JiraIssueCustomerName = new_dict[JiraIssueCustomerName]
    JiraIssueFeature = a['feature']  # request.form['Feature'].strip()
    JiraIssueFeatureComponent = request.form['FeatureComponent'].strip()
    JiraIssueOther = request.form['Other'].strip()  # Parent task has, Child task does not have
    JiraIssueType = '5WhyRCA'
    JiraIssueCaseType = ''
    JiraIssueLabels = request.form['Labels'].strip().split()

    # JiraIssueAssignee = request.form['AssignTo'].strip()
    # JiraIssueAssignee = getAccountIdByEmailName(JiraIssueAssignee)

    # PRID = request.form['PRID'].strip()
    PRID = a['id']
    PRTitle = a['title']  # request.form['PRTitle'].strip()
    # PRRcaEdaAssessor = request.form['AssignTo'].strip()
    PRRelease = a['softwareRelease']  # request.form['PRRelease'].strip()
    # #PRAttached = request.form['PRAttached'].strip()
    # PRSeverity = a['severity'] #request.form['Severity'].strip()
    # PRGroupInCharge = a['groupIncharge'] #request.form['GroupInCharge'].strip()
    # PRProduct = a['product'] #request.form['product'].strip()
    # ReportedBy = a['reportedBy']
    # FaultCoordinator = a['devFaultCoord']
    # if 'customerName' in a.keys():
    #     CustomerName = a['customerName']
    # else:
    #     CustomerName = 'Nokia'
    upload_status, link = (True,'2')
    if upload_status == False:
        return False, ''
    # prontolink = "https://pronto.inside.nsn.com/pronto/problemReportSearch.html?freeTextdropDownID=prId&searchTopText={0}".format(
    prontolink = "https://pronto.int.net.nokia.com/pronto/problemReport.html?prid=={0}".format(PRID)
    analysislink = str(link)
    Prontonumber = PRID + ' ' + PRAttached
    if len(Prontonumber) > 255:
        Prontonumber = Prontonumber[:255]
    Prontonumber = tools_strip(Prontonumber)
    summaryinfo = PRID + ':' + PRTitle  # '5WhyRca Parent Task for ' + PRID
    JiraIssueAssignee = ''
    # fourtyFiveDaysLater = (datetime.datetime.now() + datetime.timedelta(days=45)).strftime("%Y-%m-%d")
    # dict1 = {'customfield_38090': fourtyFiveDaysLater}
    issueaddforRCAsubtask = {
        'project': {'id': u'41675'},
        'issuetype': {'id': u'17678'},
        'summary': summaryinfo,
        'customfield_37060': Prontonumber,
        'customfield_37069': {'id': JiraIssueSourceList},  # Source List u'206960'
        'customfield_37460': {'id': JiraIssueBusinessUnit},  # Business Unit{}
        'customfield_37061': [{'id': JiraIssueBusinessLine}],  # 'Business Linef{}',
        'customfield_37062': {'id': JiraIssueProductLine},  # {'Product Line'},
        'customfield_37063': JiraIssueFeatureComponent,  # 'Feature Component'
        'customfield_37015': prontolink,
        'priority': {'id': JiraIssuePriority},
        'labels': JiraIssueLabels,  # ['AAA', 'BBB'],
        'customfield_37059': PRRelease,  # 'Release Fault Introduced', #softwareRelease
        'customfield_32006': [{'id': JiraIssueCustomerName}],  # {'Customer':}, # less[
        'customfield_38068': JiraIssueOther,  # {Other} #More filled CMCC
        'customfield_37064': analysislink,
        'reporter': {'name': reporter},
        # 'customfield_38090': fourtyFiveDaysLater,
        'assignee': {'name': JiraIssueAssignee}  # Optional
    }
    # issues = jira.search_issues(jql_str='project = MNPRCA AND summary~' + PRID + ' AND type = "5WhyRCA" ', maxResults=5)
    # existFlag = True  # existFlag not reset ,represent the issue not exist.
    # for newissue in issues:
    #     newissue.update(issueaddforRCAsubtask)
    #     existFlag = False
    # if existFlag:  # Flag be reset to False, no need create new issue at all
    newissue = jira.create_issue(issueaddforRCAsubtask)
    return True, newissue


def getAccountIdByEmailName(JiraIssueAssignee):
    # AssignTo = JiraIssueAssignee
    # user = JiraUser.query.filter_by(email=JiraIssueAssignee).first()
    # if user:
    #     AssignTo = user.username
    # else:
    if JiraIssueAssignee == '':
        return JiraIssueAssignee
    flg = True
    try:
        conn = get_ldap_connection()
        # flg = False
    except:
        # flg=False
        return redirect(url_for('login'))
    # if flg:
    #     return ''
    filter = '(mail=%s)' % JiraIssueAssignee
    attrs = ['sn', 'mail', 'cn', 'displayName', 'uid']
    base_dn = 'o=NSN'
    flg = True
    try:
        result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
        # flg = False
    except:
        # flg=False
        print 'No such account!!!'
        AssignTo = ''
        # return AssignTo
        # pass
    if flg:
        AssignTo = result[0][1]['uid'][0]
    return AssignTo


def APIPr(PRID):
    url1 = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/problemReport/%s" % PRID
    r = requests.get(url1, verify=False, auth=('krma', 'Nokia123'))
    r_status = r.status_code
    if r_status != 200:
        return
    b = json.loads(r.text)
    return b

def APIFA(faultAnalysisId):
    faultAnalysisId = faultAnalysisId.strip()
    urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
    r = requests.get(urlfa, verify=False, auth=('krma', 'Nokia123'))
    r_status = r.status_code
    if r_status != 200:
        raise Exception('APIFA error content='+str(r.content))
    b = json.loads(r.text) # FA json
    return b

def testxls():
    filename = 'C:\\Work\\6.Work\\2.RCA\\4.src\\1.DockerProductSrc\\product-mengli\\fddrcatracking\\fdd_rca_tracking\\rca-data\\apupload\\pronto01309318fdr\\01309318.xlsx'
    workbook = xlrd.open_workbook(filename)
    print(workbook)
#
# def getjirains():
#     if jirains != None:
#         return jirains
#     else :
#         global jirains
#         jirains = getjira()
#         return jirains
#
# def initjira():
#     global jirains
#     jirains = getjira()

def myjob(par):
    time.sleep(5)
    print(par)


class SystemUser(db.Model):
    __tablename__ = 'system_user'
    employeeNumber = db.Column('employeeNumber', db.String(20), primary_key=True)
    displayName = db.Column(db.String(200), default='')
    mail = db.Column(db.String(200), default='')
    uid = db.Column(db.String(200), default='')
    o = db.Column(db.String(200), default='')
    ou = db.Column(db.String(200), default='')
    cn = db.Column(db.String(200), default='')

    def __init__(self, employeeNumber, displayName, mail, uid, o,ou,cn):
        self.employeeNumber = employeeNumber
        self.displayName = displayName
        self.mail = mail
        self.uid = uid
        self.o = o
        self.ou = ou
        self.cn = cn

def pullEmailByName(uname):
    conn = get_ldap_connection()
    filter = 'displayName=%s' % uname
    attrs = ['mail', 'displayName', 'uid']
    base_dn = 'o=NSN'
    flg = False
    result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    for i in result:
        try:
            employeeNumber = ""
            o = ''
            ou = ''
            cn = ''
            for j in i[0].split(','):
                k = j.split('=')[0]
                v = j.split('=')[1]
                if 'employeeNumber' == k:
                    employeeNumber = v
                elif 'ou' == k:
                    ou = v
                elif 'cn' == k:
                    cn = v
                elif 'o' == k:
                    o = v
            displayName = i[1]['displayName'][0]
            mail = i[1]['mail'][0]
            uid = i[1]['uid'][0]
            print('employeeNumber:' + employeeNumber + ',displayName:' + displayName + ',mail:' + mail + ',uid:' + uid + ',o=' + o)
        except BaseException as e1:
            print('error:' + str(i) + ',e1=' + str(e1))
            pass
def pullEmail(*args):
    conn = get_ldap_connection()
    filter = '(mail=*%s*%s*)' % (args[0], args[1])
    attrs = ['mail', 'displayName', 'uid']
    base_dn = 'o=NSN'
    flg = False
    req_ctrl = SimplePagedResultsControl(size=30, cookie='')
    result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs, serverctrls=[req_ctrl])
    for i in result:
        try:
            employeeNumber = ""
            o = ''
            ou = ''
            cn = ''
            for j in i[0].split(','):
                k = j.split('=')[0]
                v = j.split('=')[1]
                if 'employeeNumber' == k:
                    employeeNumber = v
                elif 'ou' == k:
                    ou = v
                elif 'cn' == k:
                    cn = v
                elif 'o' == k:
                    o = v
            displayName = i[1]['displayName'][0]
            mail = i[1]['mail'][0]
            uid = i[1]['uid'][0]
            #print('employeeNumber:' + employeeNumber + ',displayName:' + displayName + ',mail:' + mail + ',uid:' + uid + ',o=' + o)
            rstc = SystemUser.query.filter(SystemUser.employeeNumber == employeeNumber).count()
            if rstc == 0:
                su = SystemUser(employeeNumber, displayName, mail, uid, o, ou, cn)
                db.session.add(su)
                db.session.commit()
        except BaseException as e1:
            print('error:' + str(i) + ',e1=' + str(e1))
            pass
    print("pull : "+ str(len(result)))

def batchPullEmail():
    strr = ""
    for i in range(65,91):
        for j in range(65,91):
            for k in range(65,91):
                strr = chr(i) + chr(j) + chr(k)
                if strr in ['AGA','ACH','ADA']:
                    continue
                print("=====================" + strr + "===============")
                pullEmail(strr,'nokia.com')
                pullEmail(strr,'nokia-sbell.com')

def pullTribe(*args):
    conn = get_ldap_connection()
    # filter = '(mail=*%s*%s*)' % (args[0], args[1])
    # filter = '(&(mail=*%s*%s*)(title=*Squad*))' % (args[0], args[1])
    filter = '(&(mail=*%s*%s*)(|(title=*Tribe Leader*)(title=*Head of*)))' % (args[0], args[1])
    # filter = '&(mail=*%s*%s*)(title=*Title*)' % (args[0], args[1])
    # filter = '(title=*%s*)' % ('Tribe')
    attrs = ['mail', 'displayName', 'uid','nsnManagerAccountName', 'nsnTeamName', 'title']
    base_dn = 'o=NSN'
    flg = False
    req_ctrl = SimplePagedResultsControl(size=30, cookie='')
    try:
        result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    except:
        if len(args[0]) > 4:
            print('too long')
            return
        # print(traceback.format_exc())
        for ai in range(65, 91):
            pullTribe(args[0] + chr(ai), args[1])
        pullTribe(args[0] + "@", args[1])
        pullTribe(args[0] + ".", args[1])
        pullTribe("." + args[0], args[1])
        pullTribe(" " + args[0], args[1])
        pullTribe(args[0] + " ", args[1])
        return

    for i in result:
        try:
            employeeNumber = ""
            o = ''
            ou = ''
            cn = ''
            for j in i[0].split(','):
                k = j.split('=')[0]
                v = j.split('=')[1]
                if 'employeeNumber' == k:
                    employeeNumber = v
                elif 'ou' == k:
                    ou = v
                elif 'cn' == k:
                    cn = v
                elif 'o' == k:
                    o = v
            displayName = i[1]['displayName'][0]
            mail = i[1]['mail'][0]
            uid = i[1]['uid'][0]
            tribename = i[1]['nsnTeamName'][0]
            try:
                if 'title' in i[1].keys() :
                    title = i[1]['title'][0]
                    title = title.split(',')[0]
                if 'Tribe' in title or 'Head of' in title:
                    print('%s;%s,%s' % (tribename,displayName,title))
            except:
                print(tribename)

            # if 'EXT' not in displayName \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'Deploy' not in displayName  \
            #         and 'EXT' not in displayName  \
            #         and 'Services' not in displayName  \
            #         and 'People' not in displayName  \
            #         and 'service' not in displayName  \
            #         and 'Access' not in displayName  \
            #         and 'Process' not in displayName  \
            #         and 'INFO' not in displayName  \
            #         and 'GOVERNANCE' not in displayName  \
            #         and 'Finance' not in displayName  \
            #         and 'MGMT' not in displayName  \
            #         and 'People' not in displayName  \
            #         and 'SSO' not in displayName  \
            #         and 'platform' not in displayName  \
            #         and 'TECHCloud' not in displayName  \
            #         and 'SUPPORT' not in displayName  \
            #         and 'APPLICATION' not in displayName  \
            #         and 'Applications' not in displayName  \
            #         and 'notification' not in displayName  \
            #         and 'application' not in displayName  \
            #         and 'communications' not in displayName  \
            #         and 'notification' not in displayName  \
            #         and 'COVID' not in displayName  \
            #         and 'notifications' not in displayName  \
            #         and 'COMMUNICATION' not in displayName  \
            #         and 'Hotline' not in displayName  \
            #         and 'Admin' not in displayName  \
            #         and 'admin' not in displayName  \
            #         and 'Beijing' not in displayName  \
            #         and 'Center' not in displayName  \
            #         and 'Innovation' not in displayName  \
            #         and 'Security' not in displayName  \
            #         and 'exceptions' not in displayName  \
            #         and 'PDF' not in displayName  \
            #         and 'DepREQUESTloy' not in displayName  \
            #         and 'EXCEPTION' not in displayName  \
            #         and 'COMPLIANCE' not in displayName  \
            #         and 'ACCEPTANCE' not in displayName  \
            #         and 'Lte' not in displayName  \
            #         and 'Work' not in displayName  \
            #         and 'Certification' not in displayName  \
            #         and 'EDU' not in displayName  \
            #         and 'Sales' not in displayName  \
            #         and 'assessment' not in displayName  \
            #         and 'Notification' not in displayName  \
            #         and 'Program' not in displayName  \
            #         and 'Technology' not in displayName  \
            #         and 'Communication' not in displayName  \
            #         and 'QINGDAO' not in displayName  \
            #         and 'LOCATION' not in displayName  \
            #         and 'Beijing' not in displayName  \
            #         and 'Chengdu' not in displayName  \
            #         and 'Communications' not in displayName  \
            #         and 'CAPS' not in displayName  \
            #         and 'tools' not in displayName  \
            #         and 'API' not in displayName  \
            #         and 'Management' not in displayName  \
            #         and 'Assistant' not in displayName  \
            #         and 'Team' not in displayName  \
            #         and 'Escalation' not in displayName  \
            #         and 'access' not in displayName  \
            #         and 'Technical' not in displayName  \
            #         and 'Office' not in displayName  \
            #         and 'Production' not in displayName  \
            #         and 'Mission' not in displayName  \
            #         and 'Capability' not in displayName  \
            #         and 'Capacity' not in displayName  \
            #         and 'PET' not in displayName  \
            #         and 'Test' not in displayName  \
            #         and 'EBOX' not in displayName  \
            #         and 'Scanning' not in displayName  \
            #         and 'Support' not in displayName  \
            #         and 'scanning' not in displayName  \
            #         and 'VS,' not in displayName  \
            #         and 'Event' not in displayName  \
            #         and 'support' not in displayName  \
            #         and 'Service' not in displayName  \
            #         and 'AI_cable' not in displayName  \
            #         and 'Credit' not in displayName  \
            #         and '_' not in displayName  \
            #         and 'Cable' not in displayName  \
            #         and 'WEBTOOLS' not in displayName  \
            #         and 'CERTIFICACION' not in displayName  \
            #         and 'Comunicacion' not in displayName  \
            #         and 'account' not in displayName  \
            #         and 'manager' not in displayName  \
            #         and 'Certificaciones' not in displayName  \
            #         and 'REGISTRATION' not in displayName  \
            #         and 'Nsn,' not in displayName  \
            #         and 'cacusthesecond' not in displayName  \
            #         and 'applicable' not in displayName  \
            #         and 'It,' not in displayName:
            #     if 'title' in i[1].keys():
            #         title = i[1]['title'][0]
            #         title = title.split(',')[0]
            #     else:
            #         title = ''
            #     if 'displayName' in i[1].keys():
            #         displayName = i[1]['displayName'][0]
            #     else:
            #         displayName = ''
            #     if 'nsnTeamName' in i[1].keys():
            #         squadGroupName = i[1]['nsnTeamName'][0]
            #         # nsnTeamName=i[1]['nsnOperativeOrgName'][0]
            #     else:
            #         squadGroupName = 'External People'
            #     if "Squad Group" in title: # in ['Tribe Leader', 'R&D Tribe Leader']
            #         # lineDisplayName = lineDisplayName1
            #         # tribeLeadDisplayName = displayName
            #         sgName = squadGroupName
            #         linem = i[1]['nsnManagerAccountName'][0]
            #         # lineManagerEmail = lineManagerEmail1
            #         #print('%s,%s,%s,' % (displayName,mail,sgName))
            #
            #         filter2 = '(uid=*%s*)' % (linem)
            #         # filter = '&(mail=*mingjie.fu*)(title=*Group*)'
            #         # filter = '&(mail=*%s*%s*)(title=*Title*)' % (args[0], args[1])
            #         # filter = '(title=*%s*)' % ('Tribe')
            #         attrs2 = ['mail', 'displayName', 'uid', 'nsnManagerAccountName', 'nsnTeamName', 'title']
            #         base_dn2 = 'o=NSN'
            #         result2 = conn.search_s(base_dn2, ldap.SCOPE_SUBTREE, filter2, attrs2)
            #
            #         for i2 in result2:
            #             try:
            #                 tribename2 = i2[1]['nsnTeamName'][0]
            #                 displayName2 = i2[1]['displayName'][0]
            #                 title2 = i2[1]['title'][0]
            #                 title2 = title2.split(',')[0]
            #                 if 'Tribe' in title2 :
            #                     print('%s;%s,%s' % (tribename2,displayName2,title2))
            #             except:
            #                 print(tribename2)
            #print('employeeNumber:' + employeeNumber + ',displayName:' + displayName + ',mail:' + mail + ',uid:' + uid + ',o=' + o)
            # rstc = SystemUser.query.filter(SystemUser.employeeNumber == employeeNumber).count()
            # if rstc == 0:
            #     su = SystemUser(employeeNumber, displayName, mail, uid, o, ou, cn)
            #     db.session.add(su)
            #     db.session.commit()
        except BaseException as e1:
            print('error:' + str(i) + ',e1=' + str(e1))

            pass


def batchPullTribe():
    threads = []
    strr = ""
    for i in range(65,91):
        for j in range(65,91):
            for k in range(65,66): #91
                # strr = chr(i) + chr(j) + chr(k)
                strr = chr(i) + chr(j) + ""
                # if chr(i) in ['A','B','C','D']:
                #     continue
                # if strr in ['AGA','ACH','ADA','ADI','CAR','CHA','CHE','CHI']:
                #     continue
                print("=====================" + strr + "===============")
                # try:
                #     pullTribe(strr,'nokia.com')
                # except:
                #     print(traceback.format_exc())
                #
                # try:
                #     pullTribe(strr, 'nokia-sbell.com')
                # except:
                #     print(traceback.format_exc())


                thr = threading.Thread(target=pullTribe, args=(strr,'nokia.com'))
                thr2 = threading.Thread(target=pullTribe, args=(strr,'nokia-sbell.com'))
                # thr = threading.Thread(target=getCustomerPR_ThreadTry1, args=(start,m,t,i))
                thr.start()
                thr2.start()
                threads.append(thr)
                threads.append(thr2)
                # getCustomerPR_Thread(start, total_num, t)
            for t in threads:
                t.join()

def raiseExp2():
    1/0


def raiseExe():
    raiseExp2()

# def APIGetconn():
#     if connins != None:
#         return connins
#     else :
#         global connins
#         connins = get_ldap_connection()
#         return connins

def FindProductLine(s1):
    Str=s1
    dict={}
    for key in ProductLineDict.keys():
        s2=ProductLineDict[key]
        # value = difflib.SequenceMatcher(None, Str, s2).quick_ratio()
        value = fuzz.partial_ratio(Str, s2)
        # value = get_num(Str,s2)
        dict.setdefault(key,value)
    # dict.sort(key=lambda x: x[1], reverse=True)
    # aaa= sorted(dict.iteritems(), key=lambda d: d[1], reverse=False)
    d=dict
    dd = sorted(d.items(), key=lambda x: x[1], reverse=True)
    k = dd[0]
    kk=k[0] # value
    aa= ProductLineDict[kk]
    # print aa
    return aa

def getBusinessLineAndProductLineWithJson(a):
    PRRelease = a['softwareRelease']
    PRProduct = a['product']
    BusinessLine = 'None'
    ProductLine = FindProductLine(PRProduct)
    if PRProduct == 'Flexi LTE Base Station' or PRProduct == 'Flexi LTE_FDD Base Station' or \
        PRProduct == 'Flexi LTE_TDD Base Station':
        BusinessLine = 'LTE'
    if PRProduct == 'Flexi Multiradio BTS SingleRAN' or PRProduct == 'Flexi WCDMA Base Station':
        BusinessLine = 'SRAN'
    if PRProduct == 'FlexiZone Micro':
        BusinessLine = 'Small Cell & WiFi Solutions'
    if PRProduct == 'Flexi LTE_FDD Base Station':
        BusinessLine = 'LTE'
    if PRRelease:
        if (len(PRRelease.split('TL')) != 1):
            ProductLine = 'TD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('FL')) != 1 and len(PRRelease.split('FLF')) == 1 and \
              len(PRRelease.split('TLF')) == 1): #FLF/TLF
            ProductLine = 'FDD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('CBTS')) != 1):
            ProductLine = 'Cloud BTS'
            BusinessLine = 'Telco Cloud'
        elif (len(PRRelease.split('SBTS')) != 1):
            ProductLine = 'SRAN/SBTS'
            BusinessLine = 'SRAN'
        elif (len(PRRelease.split('FLF')) != 1):
            ProductLine = 'FDD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('WBTS')) != 1):
            ProductLine = 'WBTS'
            BusinessLine = 'SRAN'
    return BusinessLine,ProductLine

class JiraUser(db.Model):
    __tablename__ = "jirausers"
    id = db.Column('user_id',db.Integer , primary_key=True)
    username = db.Column('username', db.String(64), unique=True , index=True)
    password = db.Column('password' , db.String(256))
    email = db.Column('email',db.String(128),unique=True , index=True)
    displayName = db.Column(db.String(128))
    lineManagerAccountId = db.Column(db.String(128)) #nsnManagerAccountName
    lineManagerDisplayName = db.Column(db.String(128)) #nsnManagerName
    lineManagerEmail = db.Column(db.String(128)) # Further search thru uid.
    squadGroupName = db.Column(db.String(128))  # Further search thru uid.
    registered_on = db.Column('registered_on' , db.DateTime)

    # todos = db.relationship('Todo', backref='jirauser', lazy='select')
    # jiratodos = db.relationship('JiraTodo', backref='jirauser', lazy='select')
    # todoaps = db.relationship('TodoAP', backref='jirauser', lazy='select')
    # todolongcycletimercas = db.relationship('TodoLongCycleTimeRCA', backref='jirauser', lazy='select')
    # inchargegroups = db.relationship('InChargeGroup', backref='jirauser', lazy='select')

    def __init__(self , username, email,displayName,lineManagerAccountId,lineManagerDisplayName,lineManagerEmail,squadGroupName):
        self.username = username
        self.email = email
        self.displayName = displayName
        self.lineManagerAccountId = lineManagerAccountId
        self.lineManagerDisplayName = lineManagerDisplayName
        self.lineManagerEmail = lineManagerEmail
        self.squadGroupName = squadGroupName
        self.registered_on = datetime.datetime.now() #datetime.utcnow()

    @staticmethod
    def try_login(username, password):
        global gSSOPWD
        try:
            conn = get_ldap_connection()
        except:
            flash('LDAP Connection Failed!!!')
            return redirect(url_for('login'))
        filter = '(uid=%s)'%username
        attrs = ['sn', 'mail', 'cn', 'displayName', 'nsnManagerAccountName','nsnOperativeOrgName']
        base_dn = 'o=NSN'
        try:
            result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
        except:
            flash('LDAP Searching Failed!!!')
            return redirect(url_for('login'))
        dn = result[0][0]
        try:
            a = conn.simple_bind_s(dn,password)
            gSSOPWD = password # For privacy policy, cannot save this for other purpose.
        except ldap.INVALID_CREDENTIALS:
            print "Your  password is incorrect!!!"
            flash('Password is incorrect!!')
            return redirect(url_for('login'))
            #sys.exit()
        except ldap.LDAPError, e:
            if type(e.message) == dict and e.message.has_key('desc'):
                print e.message['desc']
            else:
                print e
            #sys.exit()
            flash('LDAP Bind Failed!!!')
            return redirect(url_for('login'))
        except:
            flash('Other reason login Failed!!!')
            return redirect(url_for('login'))
        return result,conn
    def is_authenticated(self):
        return True
    def is_active(self):
        return True
    def is_anonymous(self):
        return False
    def get_id(self):
        return unicode(self.id)
    def __repr__(self):
        return '<JiraUser %r>' % (self.username)

class InChargeGroup(db.Model):
    __tablename__ = "inchargegroups"
    # id = db.Column('index',db.Integer)
    id = db.Column(db.Integer,primary_key=True)
    InChargeGroupName = db.Column(db.String(128))
    AssessorEmail = db.Column(db.String(64),index=True)
    Labels = db.Column(db.String(128))
    BusinessUnit = db.Column(db.String(128))
    BusinessLine = db.Column(db.String(128))
    ProductLine = db.Column(db.String(128),default='All')
    JIRAProject = db.Column(db.String(128))
    AddedBy = db.Column(db.String(64))
    registered_on = db.Column('registered_on', db.DateTime)
    BG = db.Column(db.String(64))
    DU = db.Column(db.String(64))
    Tribe = db.Column(db.String(64))
    RCAFilter = db.Column(db.String(128))
    EDA_AssessorEmail = db.Column(db.String(128))
    EDACreatingTime = db.Column(db.String(128))
    EdaCaseType = db.Column(db.String(128))
    RCAEDACategory = db.Column(db.String(128))

    # user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'))
    jirauser_id = db.Column(db.Integer, db.ForeignKey('jirausers.user_id'))

    def __init__(self , InChargeGroupName ,AssessorEmail,Labels,BusinessUnit,BusinessLine,ProductLine,JIRAProject,AddedBy, \
                 BG, DU, Tribe,RCAFilter,EDA_AssessorEmail,EDACreatingTime,EdaCaseType,RCAEDACategory):
        self.InChargeGroupName = InChargeGroupName
        self.AssessorEmail = AssessorEmail
        self.Labels = Labels
        self.BusinessUnit = BusinessUnit
        self.BusinessLine = BusinessLine
        self.ProductLine = ProductLine
        self.JIRAProject = JIRAProject
        self.AddedBy = AddedBy
        self.registered_on = datetime.datetime.now()
        self.BG = BG
        self.DU = DU
        self.Tribe = Tribe
        self.RCAFilter = RCAFilter
        self.EDA_AssessorEmail = EDA_AssessorEmail
        self.EDACreatingTime = EDACreatingTime
        self.EdaCaseType = EdaCaseType
        self.RCAEDACategory = RCAEDACategory

def getBusinessLineAndProductLine(PRID):
    flg = False
    url = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/problemReport/%s" % PRID
    try:
        r = requests.get(url, verify=False, auth=(gAdmin, gAuth))
        a = r.json()
        PRRelease = a['softwareRelease']
        PRProduct = a['product']
    except:
        flg = True
        print 'Get problemReports/%s ERROR'%PRID
        pass
    if flg:
        print 'Get problemReports/%s ERROR' % PRID
        return '',''
    # print PRID
    # BusinessLine, ProductLine = '',''
    # a = r.json()
    # PRRelease = a['softwareRelease']
    # PRProduct = a['product']
    BusinessLine = 'None'
    ProductLine = FindProductLine(PRProduct)
    if PRProduct == 'Flexi LTE Base Station' or PRProduct == 'Flexi LTE_FDD Base Station' or \
        PRProduct == 'Flexi LTE_TDD Base Station':
        BusinessLine = 'LTE'
    if PRProduct == 'Flexi Multiradio BTS SingleRAN' or PRProduct == 'Flexi WCDMA Base Station':
        BusinessLine = 'SRAN'
    if PRProduct == 'FlexiZone Micro':
        BusinessLine = 'Small Cell & WiFi Solutions'
    if PRProduct == 'Flexi LTE_FDD Base Station':
        BusinessLine = 'LTE'
    if PRRelease:
        if (len(PRRelease.split('TL')) != 1):
            ProductLine = 'TD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('FL')) != 1 and len(PRRelease.split('FLF')) == 1 and \
              len(PRRelease.split('TLF')) == 1): #FLF/TLF
            ProductLine = 'FDD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('CBTS')) != 1):
            ProductLine = 'Cloud BTS'
            BusinessLine = 'Telco Cloud'
        elif (len(PRRelease.split('SBTS')) != 1):
            ProductLine = 'SRAN/SBTS'
            BusinessLine = 'SRAN'
        elif (len(PRRelease.split('FLF')) != 1):
            ProductLine = 'FDD LTE'
            BusinessLine = 'LTE'
        elif (len(PRRelease.split('WBTS')) != 1):
            ProductLine = 'WBTS'
            BusinessLine = 'SRAN'
    return BusinessLine,ProductLine

def getGroupInChargeInfo(PRID,PRGroupInCharge):
    groupInChargeInfo = tryGetGroupInChargeInfo(PRID, PRGroupInCharge)
    return groupInChargeInfo

def tryGetGroupInChargeInfo(PRID,PRGroupInCharge):
    groupInChargeInfo = ''
    BusinessLine,ProductLine = getBusinessLineAndProductLine(PRID)
    groupInChargeInfo = InChargeGroup.query.get(1)
    groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
                                                     BusinessLine=BusinessLine,
                                                     ProductLine='All').first()
    # print 'get All'
    if not groupInChargeInfo:
        groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
                                                         BusinessLine=BusinessLine,
                                                         ProductLine=ProductLine).first()
    # print 'get Productline first'
    if not groupInChargeInfo:
        groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
                                                         AddedBy='sufang.huang@nokia-sbell.com').first()
    # print 'get AddedBy'
    if not groupInChargeInfo:
        num = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge).count()
        if num == 1:
            groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge).first()
        # elif num == 0:
        #     print 'No registerred yet'
        # else:
        #     print 'There are shared GIC....'
    # print 'get by GIC only'
    return groupInChargeInfo


def isFNRedNew(a):
    # url = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/problemReport/%s" % PRID
    # r = requests.get(url, verify=False, auth=(gAdmin, gAuth))
    # a = r.json()
    FNRed = True
    correctionIds = a['correctionIds']
    for correctionId in correctionIds:
        urlcf = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/correction/%s" % correctionId
        # try:
        # print 'isFNRedNew  correctionId= %s!!!!!!!!!!!!!!!!!!!!!' % correctionId
        r = requests.get(urlcf, verify=False, auth=('krma', 'Nokia123'))
        # except:
        #     FNRed = True
        #     pass
        # try:
        r_status = r.status_code
        if r_status != 200:
            return True
            # b = r.text
        b = json.loads(r.text)
        # b = r.json()
        targetBuild = b['targetBuild']
        state1 = b['state']
        if targetBuild:
            targetBuildFNR = targetBuild.split('FNR')
            targetBuildFNE = targetBuild.split('FNE')
            # if len(targetBuildFNR) > 1 or len(targetBuildFNE) > 1:
            #     FNRed = True
            if state1 == 'Tested':
                if len(targetBuildFNR) == 1 and len(targetBuildFNE) == 1:
                    FNRed = False
                    return FNRed
                else:
                    FNRed = True
            else:
                FNRed = True
        # except:
        #     FNRed = True
        #     pass
    return FNRed

def checkPRnotgeneratelist(Plist):
    for i in Plist:
        try:
            PRID = i
            po = APIPr(PRID)
            rst = MnTodo.query.get(PRID)
            if not rst :
                BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(po)
                state = po['state']
                if state == 'Closed' or state == 'First Correction Complete' or \
                        state == 'Final Response Sent':  # "Correction Plan Ready",
                    author = po['author']
                    PRGroupInCharge = po['groupIncharge']
                else:
                    continue
                todo_item = getGroupInChargeInfo(PRID, PRGroupInCharge)
                if todo_item:
                    JIRAProject = todo_item.JIRAProject
                    RCAFilter = todo_item.RCAFilter
                    # Below for Only Customer (A+B)
                    if author != 'Electra':
                        continue
                    # FNRed = False
                    FNRed = isFNRedNew(po)
                    if FNRed == True and JIRAProject != 'CUDOCIPBL':
                        continue
                    if RCAFilter == 'All Customer(A+B+C), All Internal':
                        print('All Customer(A+B+C), All Internal')
                        continue
                    if RCAFilter == 'Partial Customer(A+B) and Internal, both according to FA rcaEdaDecision Flag':
                        print('Partial Customer(A+B) and Internal, both according to FA rcaEdaDecision Flag')
                        continue
                    if RCAFilter == 'No Internal, Partial Customer(A+B) according to FA rcaEdaDecision Flag':
                        print('No Internal, Partial Customer(A+B) according to FA rcaEdaDecision Flag')
                        continue
                    if RCAFilter == 'All Customer(A+B), Partial Internal according to FA rcaEdaDecision Flag':
                        print('All Customer(A+B), Partial Internal according to FA rcaEdaDecision Flag')
                    if RCAFilter == 'Customer A + Key Customers B':
                        addToMnRcaCustomerABnew(PRID, po, author, JIRAProject)
                    print('lack pr:' + str(PRID))
                    if PRID in ['02408288','02399531','02357691']:
                        print('find!!!============================================' + str(PRID))
        except BaseException as ec1:
            print(ec1)

def APIsiglepr(prid):
    try:
        # print '********************* i=%d @@@@@@@@@@@@@@@@@@@@@@@@@@@@@' % i
        PRID = prid
        probj= APIPr(PRID)
        # if PRID != '02458610':
        #     continue
        BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(probj)
        state = probj['state']
        if state == 'Closed' or state == 'First Correction Complete' or \
                state == 'Final Response Sent':  # "Correction Plan Ready",
            author = probj['author']
            PRGroupInCharge = probj['groupIncharge']
        else:
            return
        todo_item = getGroupInChargeInfo(PRID, PRGroupInCharge)
        if todo_item:
            JIRAProject = todo_item.JIRAProject
            RCAFilter = todo_item.RCAFilter
            # if state == 'Closed' or state == 'First Correction Complete' or \
            #         state == 'Final Response Sent':  # "Correction Plan Ready",
            if RCAFilter == 'All Customer(A+B+C), All Internal':
                print('addToMnRcaAll(PRID)')
                return
            if RCAFilter == 'Partial Customer(A+B) and Internal, both according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                print('addToMnRca(PRID, urlfa, probj)')
                return
            if RCAFilter == 'No Internal, Partial Customer(A+B) according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                print('addToMnRca(PRID, urlfa, probj)')
                return
            if RCAFilter == 'All Customer(A+B), Partial Internal according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                print('addToMnRca(PRID, urlfa, probj)')
                return
                # continue
            if RCAFilter == 'Customer A + Key Customers B':
                addToMnRcaCustomerABnew(PRID, probj, author, JIRAProject)
                return
            # Below for Only Customer (A+B)
            if author != 'Electra':
                return
            # FNRed = False
            problemType = probj['problemType']  # Documentation
            FNRed = isFNRedNew(probj)
            if FNRed == True and JIRAProject != 'CUDOCIPBL':
                todo_item = MnTodo.query.get(PRID)
                if todo_item:
                    if todo_item.FaRcaEdaDecisionFlag == 'Yes':
                        return None
                    else:
                        db.session.delete(todo_item)
                        db.session.commit()
                        return None
            # PRID = probj['id']
            PRReportedDate = probj['reportedDate'].split('T')[0]
            PRClosedDate = probj['lastMod'].split('T')[0]
            # print PRID
            try:
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                # else:
                #     print 'No '
            except:
                pass

            PRTitle = probj['title']
            ReportedBy = probj['reportedBy']
            PRGroupInCharge = probj['groupIncharge']
            FaultCoordinator = probj['devFaultCoord']
            PRSeverity = probj['severity']
            if PRSeverity == 'A - Critical':
                rank = 1
            elif PRSeverity == 'B - Major':
                rank = 1
            elif PRSeverity == 'C - Minor':
                rank = 3
            PRRelease = probj['softwareRelease']
            Feature = probj['feature']
            PRProduct = probj['product']
            if 'customerName' in probj.keys():
                CustomerName = probj['customerName']
            else:
                CustomerName = 'Nokia'
            PRAttached = probj['problemReportIds']
            attachedlist = PRAttached
            if not PRAttached:
                PRAttached = ''
            else:
                LabelItems = PRAttached
                b = ''
                for s in LabelItems:
                    b = b + s + ','
                    PRAttached = b.strip(',')
            for item in attachedlist:
                c = APIPr(item)
                Attachedauthor = c['author']
                if Attachedauthor == 'Electra':
                    attachedSeverity = c['severity']
                    if attachedSeverity == 'A - Critical':
                        rank1 = 1
                    elif attachedSeverity == 'B - Major':
                        rank1 = 1
                    elif attachedSeverity == 'C - Minor':
                        rank1 = 3
                    if rank1 == rank:
                        AttachedPRReportedDate = c['reportedDate'].split('T')[0]
                        if AttachedPRReportedDate < PRReportedDate:
                            PRID = c['id']
                            BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                            PRReportedDate = c['reportedDate'].split('T')[0]
                            PRGroupInCharge = c['groupIncharge']
                            # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                            PRClosedDate = c['lastMod'].split('T')[0]
                            dateFlag, realDate = getPRCloseDate(PRID)
                            if dateFlag:
                                PRClosedDate = realDate
                                # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                            PRTitle = c['title']
                            ReportedBy = c['reportedBy']
                            FaultCoordinator = c['devFaultCoord']
                            PRSeverity = c['severity']
                            if PRSeverity == 'A - Critical':
                                rank = 1
                            elif PRSeverity == 'B - Major':
                                rank = 1
                            elif PRSeverity == 'C - Minor':
                                rank = 3
                            PRRelease = c['softwareRelease']
                            Feature = c['feature']
                            PRAttached = c['problemReportIds']
                            if not PRAttached:
                                PRAttached = ''
                            else:
                                LabelItems = PRAttached
                                b = ''
                                for s in LabelItems:
                                    b = b + s + ','
                                    PRAttached = b.strip(',')
                            PRProduct = c['product']
                            if 'customerName' in c.keys():
                                CustomerName = c['customerName']
                            else:
                                CustomerName = 'Nokia'
                    elif rank1 < rank:
                        PRID = c['id']
                        BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                        PRReportedDate = c['reportedDate'].split('T')[0]
                        PRGroupInCharge = c['groupIncharge']
                        # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                        PRClosedDate = c['lastMod'].split('T')[0]
                        # print 'before getPRCloseDate  PRID= %s!!!!!!!!!!!!!!!!!!!!!' % PRID
                        dateFlag, realDate = getPRCloseDate(PRID)
                        if dateFlag:
                            PRClosedDate = realDate
                            # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                        PRTitle = c['title']
                        ReportedBy = c['reportedBy']
                        FaultCoordinator = c['devFaultCoord']
                        PRSeverity = c['severity']
                        if PRSeverity == 'A - Critical':
                            rank = 1
                        elif PRSeverity == 'B - Major':
                            rank = 1
                        elif PRSeverity == 'C - Minor':
                            rank = 3
                        PRRelease = c['softwareRelease']
                        Feature = c['feature']
                        PRAttached = c['problemReportIds']
                        if not PRAttached:
                            PRAttached = ''
                        else:
                            LabelItems = PRAttached
                            b = ''
                            for s in LabelItems:
                                b = b + s + ','
                                PRAttached = b.strip(',')
                        PRProduct = c['product']
                        if 'customerName' in c.keys():
                            CustomerName = c['customerName']
                        else:
                            CustomerName = 'Nokia'
                # else:
                #     print 'Internal PR, ignore'
            # BusinessLine, ProductLine = getBusinessLineAndProductLine(PRID)
            groupInChargeInfo = getGroupInChargeInfo(PRID, PRGroupInCharge)
            if groupInChargeInfo:
                JIRAProject = groupInChargeInfo.JIRAProject
                JiraCustomerName = FindCustomerName(CustomerName.lower())
                # ProductLine = FindProductLine(PRProduct)
                BusinessLine = groupInChargeInfo.BusinessLine
                if groupInChargeInfo.ProductLine != 'All':
                    ProductLine = groupInChargeInfo.ProductLine

                if PRAttached:
                    attachedList = PRAttached.split(',')
                    for item in attachedList:
                        todo_item = MnTodo.query.get(item)
                        if todo_item:
                            db.session.delete(todo_item)
                            db.session.commit()
                todomn = MnTodo.query.get(PRID)
                if todomn:
                    todomn.ProductLine = ProductLine
                    todomn.JiraCustomerName = JiraCustomerName
                    todomn.RCAEDACategory = groupInChargeInfo.RCAEDACategory
                    todomn.PRClosedDate = PRClosedDate
                    todomn.PRAttached = PRAttached
                    todomn.JIRAProject = JIRAProject
                    todomn.PRSeverity = PRSeverity
                    todomn.PRGroupInCharge = PRGroupInCharge
                    db.session.commit()
                    todo_item = Todo.query.get(PRID)
                    if todo_item:
                        todo_item.PRSeverity = todomn.PRSeverity
                        todo_item.PRAttached = todomn.PRAttached
                        todo_item.PRGroupInCharge = todomn.PRGroupInCharge
                        todo_item.PRProduct = todomn.PRProduct
                        todo_item.ReportedBy = todomn.ReportedBy
                        todo_item.FaultCoordinator = todomn.FaultCoordinator
                        todo_item.CustomerName = todomn.CustomerName
                        todo_item.Feature = todomn.Feature
                        todo_item.JiraCustomerName = todomn.JiraCustomerName
                        todo_item.JIRAProject = todomn.JIRAProject
                        db.session.commit()
                else:
                    PRRcaEdaAssessor = groupInChargeInfo.AssessorEmail
                    EDA_AssessorEmail = groupInChargeInfo.EDA_AssessorEmail
                    EDACreatingTime = groupInChargeInfo.EDACreatingTime
                    EdaCaseType = groupInChargeInfo.EdaCaseType
                    BusinessUnit = groupInChargeInfo.BusinessUnit
                    # ProductLine = groupInChargeInfo.ProductLine
                    RCAEDACategory = groupInChargeInfo.RCAEDACategory
                    FaRcaEdaDecisionFlag = ''
                    JiraTaskIsCreated = 'No'
                    todo = MnTodo(PRID, PRTitle, PRReportedDate, PRClosedDate, PRRelease, PRSeverity,
                                  PRGroupInCharge, PRAttached, PRRcaEdaAssessor, \
                                  PRProduct, ReportedBy, FaultCoordinator, CustomerName, BusinessUnit,
                                  BusinessLine, ProductLine, Feature, RCAEDACategory, \
                                  JiraCustomerName, JIRAProject, FaRcaEdaDecisionFlag, EDA_AssessorEmail,
                                  EDACreatingTime, EdaCaseType, JiraTaskIsCreated)
                    # todo.jirauser = g.user
                    db.session.add(todo)
                    db.session.commit()
                    todo_item = Todo.query.get(PRID)
                    if todo_item:
                        todo_item.PRSeverity = PRSeverity
                        todo_item.PRAttached = PRAttached
                        todo_item.PRGroupInCharge = PRGroupInCharge
                        todo_item.PRProduct = PRProduct
                        todo_item.ReportedBy = ReportedBy
                        todo_item.FaultCoordinator = FaultCoordinator
                        todo_item.CustomerName = CustomerName
                        todo_item.Feature = Feature
                        todo_item.JiraCustomerName = JiraCustomerName
                        todo_item.JIRAProject = JIRAProject
                    db.session.commit()
            else:
                return None
        else:
            return None
    except BaseException as ec1:
        errmes = 'error: PRID=' + str(PRID) + ', errorMes=' + str(ec1)
        print(errmes)
        syslog_error('getCustomerPR_Thread', errormes=errmes)


class Todo(db.Model):
    __tablename__ = 'rcastatus'
    PRID = db.Column('PRID', db.String(64), primary_key=True)
    PRTitle = db.Column(db.String(1024))
    PRReportedDate = db.Column(db.String(64))
    PRClosedDate=db.Column(db.String(64))
    PROpenDays=db.Column(db.Integer)
    PRRcaCompleteDate = db.Column(db.String(64))
    PRRelease = db.Column(db.String(128))
    PRAttached = db.Column(db.String(128))
    IsLongCycleTime = db.Column(db.String(32))
    IsCatM = db.Column(db.String(32))
    IsRcaCompleted = db.Column(db.String(32))
    NoNeedDoRCAReason = db.Column(db.String(64))
    RootCauseCategory=db.Column(db.String(1024))
    FunctionArea = db.Column(db.String(1024))
    CodeDeficiencyDescription = db.Column(db.String(2048))
    CorrectionDescription=db.Column(db.String(1024))
    RootCause = db.Column(db.String(1024))
    IntroducedBy = db.Column(db.String(128))
    Handler = db.Column(db.String(64))

    # New added field for JIRA deployment
    LteCategory=db.Column(db.String(32))
    CustomerOrInternal = db.Column(db.String(32))
    JiraFunctionArea=db.Column(db.String(32))
    TriggerScenarioCategory = db.Column(db.String(128))
    FirstFaultEcapePhase=db.Column(db.String(32))
    FaultIntroducedRelease = db.Column(db.String(256))
    TechnicalRootCause = db.Column(db.String(1024))
    TeamAssessor = db.Column(db.String(64))
    EdaCause = db.Column(db.String(1024))
    RcaRootCause5WhyAnalysis = db.Column(db.String(2048))
    JiraRcaBeReqested = db.Column(db.String(32))
    JiraIssueStatus = db.Column(db.String(32))
    JiraIssueAssignee = db.Column(db.String(128))
    JiraRcaPreparedQualityRating = db.Column(db.Integer)
    JiraRcaDeliveryOnTimeRating = db.Column(db.Integer)
    RcaSubtaskJiraId = db.Column(db.String(32))
    ShouldHaveBeenDetected = db.Column(db.String(128))
    rcaDueDate = db.Column(db.String(64))
    # End of new added field for JIRA

    PRSeverity = db.Column(db.String(32))
    PRGroupInCharge = db.Column(db.String(64))
    PRProduct = db.Column(db.String(64))
    ReportedBy = db.Column(db.String(64))
    FaultCoordinator = db.Column(db.String(64))
    CustomerName = db.Column(db.String(128))
    Feature = db.Column(db.String(128))
    JiraCustomerName = db.Column(db.String(128))
    JIRAProject = db.Column(db.String(128))
    EdaJiraIssueAssignee = db.Column(db.String(128))
    EdaSubtaskJiraId = db.Column(db.String(32))
    edaDueDate = db.Column(db.String(64))
    EdaJiraIssueStatus = db.Column(db.String(32))


    rca5whys = db.relationship('Rca5Why' , backref='todo',lazy='select')
    # user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'))
    jirauser_id = db.Column(db.Integer, db.ForeignKey('jirausers.user_id'))

    def __init__(self, PRID,PRTitle,PRReportedDate,PRClosedDate,PROpenDays,PRRcaCompleteDate,PRRelease,PRAttached,IsLongCycleTime,\
                 IsCatM,IsRcaCompleted,NoNeedDoRCAReason,RootCauseCategory,FunctionArea,CodeDeficiencyDescription,\
		 CorrectionDescription,RootCause,IntroducedBy,Handler,LteCategory,CustomerOrInternal,JiraFunctionArea,TriggerScenarioCategory, \
                 FirstFaultEcapePhase,FaultIntroducedRelease,TechnicalRootCause,TeamAssessor,EdaCause,RcaRootCause5WhyAnalysis, \
                 JiraRcaBeReqested,JiraIssueStatus,JiraIssueAssignee,JiraRcaPreparedQualityRating,\
                 JiraRcaDeliveryOnTimeRating,RcaSubtaskJiraId,ShouldHaveBeenDetected,rcaDueDate,PRSeverity,PRGroupInCharge, \
                 PRProduct,ReportedBy,FaultCoordinator,CustomerName,Feature,JiraCustomerName,JIRAProject,EdaJiraIssueAssignee,\
                 EdaSubtaskJiraId,edaDueDate,EdaJiraIssueStatus):
        self.PRID = PRID
        self.PRTitle = PRTitle
        self.PRReportedDate = PRReportedDate
        self.PRClosedDate = PRClosedDate
        self.PROpenDays = PROpenDays
        self.PRRcaCompleteDate = PRRcaCompleteDate
        self.PRRelease = PRRelease
        self.PRAttached = PRAttached
        self.IsLongCycleTime = IsLongCycleTime
        self.IsCatM = IsCatM
        self.IsRcaCompleted = IsRcaCompleted
        self.NoNeedDoRCAReason = NoNeedDoRCAReason
        self.RootCauseCategory = RootCauseCategory
        self.FunctionArea = FunctionArea
        self.CodeDeficiencyDescription = CodeDeficiencyDescription
        self.CorrectionDescription = CorrectionDescription
        self.RootCause = RootCause
        self.IntroducedBy = IntroducedBy
        self.Handler = Handler

        # New added for JIRA
        self.LteCategory = LteCategory
        self.CustomerOrInternal = CustomerOrInternal
        self.JiraFunctionArea = JiraFunctionArea
        self.TriggerScenarioCategory = TriggerScenarioCategory
        self.FirstFaultEcapePhase = FirstFaultEcapePhase
        self.FaultIntroducedRelease = FaultIntroducedRelease
        self.TechnicalRootCause = TechnicalRootCause
        self.TeamAssessor = TeamAssessor
        self.EdaCause = EdaCause
        self.RcaRootCause5WhyAnalysis = RcaRootCause5WhyAnalysis
        self.JiraRcaBeReqested = JiraRcaBeReqested
        self.JiraIssueStatus = JiraIssueStatus
        self.JiraIssueAssignee = JiraIssueAssignee
        self.JiraRcaPreparedQualityRating = JiraRcaPreparedQualityRating
        self.JiraRcaDeliveryOnTimeRating = JiraRcaDeliveryOnTimeRating
        self.RcaSubtaskJiraId = RcaSubtaskJiraId
        self.ShouldHaveBeenDetected = ShouldHaveBeenDetected
        self.rcaDueDate = rcaDueDate
        # End of new added field for JIRA
        self.PRSeverity = PRSeverity
        self.PRGroupInCharge = PRGroupInCharge
        self.PRProduct = PRProduct
        self.ReportedBy = ReportedBy
        self.FaultCoordinator = FaultCoordinator
        self.CustomerName = CustomerName
        self.Feature = Feature
        self.JiraCustomerName = JiraCustomerName
        self.JIRAProject = JIRAProject
        self.EdaJiraIssueAssignee = EdaJiraIssueAssignee
        self.EdaSubtaskJiraId = EdaSubtaskJiraId
        self.edaDueDate = edaDueDate
        self.EdaJiraIssueStatus = EdaJiraIssueStatus

def get_num(s1, s2):
    num = 0
    len_s1 = len(s1)
    list_s1 = []
    for i in range(len_s1):
        two_s1 = s1[0:i+1]
        list_s1.append(two_s1)
    for i in list_s1:
        if s2.startswith(i) and len(i) > num:
            num = len(i)
    return num

def FindCustomerName(s1):
    Str=s1
    dict={}
    for key in CustomerNameDict.keys():
        s2=CustomerNameDict[key].lower()
        # value = difflib.SequenceMatcher(None, Str, s2).quick_ratio()
        value = get_num(Str,s2)
        dict.setdefault(key,value)
    d=dict
    dd = sorted(d.items(), key=lambda x: x[1], reverse=True)
    k = dd[0]
    kk=k[0] # value
    aa= CustomerNameDict[kk]
    # print aa
    return aa

def getPRCloseDate(PRID):
    existingFlag = False
    PRClosedDate = ''
    startAt = 0
    # link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=2019-01-01' % startAt
    link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport/%s/revisionHistory?startAt=%d&maxResults=50' % (PRID,startAt)
    # https: // pronto.int.net.nokia.com / prontoapi / rest / api / latest / problemReport / CAS - 133509 - F9C8 / revisionHistory
    # https: // pronto.int.net.nokia.com / prontoapi / rest / api / latest / problemReport / CAS - 133509 - F9C8 / revisionHistory
    url = link
    errorFlag = True
    try:
        r = requests.get(url, verify=False, auth=('krma', 'Nokia123'))
    except:
        errorFlag = False
        pass
    if errorFlag:
        sucessfulFlag = False
        try:
            a = r.json()
        except:
            sucessfulFlag = True
            pass
        if sucessfulFlag:
            return existingFlag, PRClosedDate
        total = a['total']
        prlist = a['values']
        num = len(prlist)
        while (total > 0):
            for i in range(num):
                changeComment = prlist[i]['changeComment'] # First Correction Ready For Testing to Closed
                modificationDate = prlist[i]['modificationDate']
                if changeComment == 'The state of the problem changed from First Correction Complete to Closed' or \
                        changeComment == 'The state of the problem changed from Final Response Sent to Closed' or \
                        changeComment == "The state of the problem changed from Correction Plan Ready to Closed" or \
                        changeComment == 'The state of the problem changed from First Correction Ready For Testing to Closed' \
                        or changeComment == 'The state of the problem changed from Correction Plan Ready to Final Response Sent':
                    PRClosedDate = modificationDate.split('T')[0]
                    existingFlag = True
                    # return True,PRClosedDate
                else:
                    continue
            startAt = startAt + num
            total = total - num
            # link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=2019-01-01'%startAt
            link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport/%s/revisionHistory?startAt=%d&maxResults=50' % (
            PRID, startAt)
            url = link # 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=2019-04-01'%startAt
            errorFlag = True
            try:
                r = requests.get(url, verify=False, auth=('krma', 'Nokia123'))
            except:
                errorFlag = False
                print 'errorFlag@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@'
                pass
            if errorFlag:
                sucessfulFlag = False
                try:
                    a = r.json()
                except:
                    sucessfulFlag = True
                    pass
                if sucessfulFlag:
                    return existingFlag, PRClosedDate
                prlist = a['values']
                num = len(prlist)
                # print total
                # print " PRClosedDate Searching $$$$$$$$$$$$$$$$%%%%%%%%%%%%%%&&&&&&&&&&&"
        return existingFlag,PRClosedDate
    else:
        return existingFlag, PRClosedDate

def APIjiraissue(jiraissueid):
    jira = getjira()
    ob = jira.issue(jiraissueid)
    conn = get_ldap_connection()
    if ob.fields.assignee:
        JiraIssueAssigneeEmail = ob.fields.assignee.emailAddress
    LineManagerEmail, squadGroup, lineDisplayName, tribeLeadDisplayName, tribeName = getSquadTribeInfoWithConn(JiraIssueAssigneeEmail, conn)
    print(tribeName)

def APIjiracreate():
    jira = getjira()
    issueaddforRCAsubtask = {
        'project': {'id': u'41675'},
        'parent': {"key": 'MNRCA-46699'},
        'issuetype': {'id': u'19800'},  # Analysis subtask
        'summary': '01309318 test ',
        'customfield_10464': {'id': '219576'},  # Case Type
        'customfield_27792': (datetime.datetime.now() + datetime.timedelta(days=15)).strftime(
                    "%Y-%m-%dT%H:%M:00.00+0200"),
        # 'assignee': {'name': JiraIssueAssignee},
        'reporter': {'name': 'qmxh38'},
        # 'watcher': {'name': watcher},
    }
    newissue = jira.create_issue(issueaddforRCAsubtask)
    try:
        jira.assign_issue(newissue, 'qmxh38')
    except:
        pass
    jira.remove_watcher(newissue, 'qmxh38')
    try:
        jira.add_watcher(newissue, 'qmxh38')
    except:
        pass
    print(newissue)
    APIjiraissue(newissue.key)

def addToMnRcaCustomerABnew(PRID,probj,author,JIRAProject):
    # Below for Only Customer (A+B)
    FNRed = False
    if author != 'Electra':
        return
    FNRed = isFNRedNew(probj)
    if FNRed == True and JIRAProject != 'CUDOCIPBL':
        todo_item = MnTodo.query.get(PRID)
        if todo_item:
            db.session.delete(todo_item)
            db.session.commit()
        return
    PRReportedDate = probj['reportedDate'].split('T')[0]
    PRClosedDate = probj['lastMod'].split('T')[0]
    # print PRID
    try:
        dateFlag, realDate = getPRCloseDate(PRID)
        if dateFlag:
            PRClosedDate = realDate
        #     print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
        # else:
        #     print 'No '
    except:
        pass

    PRTitle = probj['title']
    ReportedBy = probj['reportedBy']
    PRGroupInCharge = probj['groupIncharge']
    FaultCoordinator = probj['devFaultCoord']
    PRSeverity = probj['severity']
    HightestLevel = -1         #the hightest level in self pr and attach pr
    if PRSeverity == 'A - Critical':
        HightestLevel = 1
    elif PRSeverity == 'B - Major':
        HightestLevel = 2
    elif PRSeverity == 'C - Minor':
        HightestLevel = 3
    PRRelease = probj['softwareRelease']
    Feature = probj['feature']
    PRProduct = probj['product']
    JiraCustomerName = probj['customerName']
    if 'customerName' in probj.keys():
        CustomerName = probj['customerName']
    else:
        CustomerName = 'Nokia'
    PRAttached = probj['problemReportIds']
    attachedlist = PRAttached
    if not PRAttached:
        PRAttached = ''
    else:
        LabelItems = PRAttached
        b = ''
        for s in LabelItems:
            b = b + s + ','
            PRAttached = b.strip(',')
    for item in attachedlist:
        c = APIPr(item)
        Attachedauthor = c['author']
        if 'reportedBy' in c.keys():
            AttachedreportedBy = c['reportedBy']
        else:
            AttachedreportedBy = ''
        if Attachedauthor == 'Electra':
            attachedSeverity = c['severity']
            if attachedSeverity == 'A - Critical':
                rank1 = 1
            elif attachedSeverity == 'B - Major':
                rank1 = 2
            elif attachedSeverity == 'C - Minor':
                rank1 = 3
            if rank1 == HightestLevel:
                AttachedPRReportedDate = c['reportedDate'].split('T')[0]
                if AttachedPRReportedDate < PRReportedDate:
                    PRID = c['id']
                    PRReportedDate = c['reportedDate'].split('T')[0]
                    PRGroupInCharge = c['groupIncharge']
                    # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                    PRClosedDate = c['lastMod'].split('T')[0]
                    dateFlag, realDate = getPRCloseDate(PRID)
                    if dateFlag:
                        PRClosedDate = realDate
                        # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                    PRTitle = c['title']
                    ReportedBy = c['reportedBy']
                    FaultCoordinator = c['devFaultCoord']
                    PRSeverity = c['severity']
                    if PRSeverity == 'A - Critical':
                        HightestLevel = 1
                    elif PRSeverity == 'B - Major':
                        HightestLevel = 2
                    elif PRSeverity == 'C - Minor':
                        HightestLevel = 3
                    PRRelease = c['softwareRelease']
                    Feature = c['feature']
                    PRAttached = c['problemReportIds']
                    if not PRAttached:
                        PRAttached = ''
                    else:
                        LabelItems = PRAttached
                        b = ''
                        for s in LabelItems:
                            b = b + s + ','
                            PRAttached = b.strip(',')
                    PRProduct = c['product']
                    JiraCustomerName = c['customerName']
                    if 'customerName' in c.keys():
                        CustomerName = c['customerName']
                    else:
                        CustomerName = 'Nokia'
            elif rank1 < HightestLevel:
                PRID = c['id']
                PRReportedDate = c['reportedDate'].split('T')[0]
                PRGroupInCharge = c['groupIncharge']
                # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                PRClosedDate = c['lastMod'].split('T')[0]
                print 'before getPRCloseDate  PRID= %s!!!!!!!!!!!!!!!!!!!!!' % PRID
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                PRTitle = c['title']
                ReportedBy = c['reportedBy']
                FaultCoordinator = c['devFaultCoord']
                PRSeverity = c['severity']
                if PRSeverity == 'A - Critical':
                    HightestLevel = 1
                elif PRSeverity == 'B - Major':
                    HightestLevel = 2
                elif PRSeverity == 'C - Minor':
                    HightestLevel = 3
                PRRelease = c['softwareRelease']
                Feature = c['feature']
                PRAttached = c['problemReportIds']
                if not PRAttached:
                    PRAttached = ''
                else:
                    LabelItems = PRAttached
                    b = ''
                    for s in LabelItems:
                        b = b + s + ','
                        PRAttached = b.strip(',')
                PRProduct = c['product']
                JiraCustomerName = c['customerName']
                if 'customerName' in c.keys():
                    CustomerName = c['customerName']
                else:
                    CustomerName = 'Nokia'
        else:
            print 'Internal PR, ignore'
    BusinessLine, ProductLine = getBusinessLineAndProductLine(PRID)
    groupInChargeInfo = getGroupInChargeInfo(PRID, PRGroupInCharge)
    if groupInChargeInfo:
        JIRAProject = groupInChargeInfo.JIRAProject
        MatchFlag = False
        if HightestLevel == 1:
            MatchFlag = True
        elif HightestLevel == 2:
            customerDict = CustomerDict.query.filter_by(RCAFilterId='A_KEY_B').all()
            proik = APIPr(PRID)
            MatchFlag,cusname = FindCustomerNameAB(PRGetCustomerNameByDefault(proik), customerDict)
            if not MatchFlag:
                for ik in PRAttached.split(','):
                    proik = APIPr(ik)
                    MatchFlag,cusname = FindCustomerNameAB(PRGetCustomerNameByDefault(proik), customerDict)
                    if MatchFlag :
                        break
        elif HightestLevel == 3:
            MatchFlag = False
        if not MatchFlag:
            return
        BusinessLine = groupInChargeInfo.BusinessLine
        if groupInChargeInfo.ProductLine != 'All':
            ProductLine = groupInChargeInfo.ProductLine
        if PRAttached:
            attachedList = PRAttached.split(',')
        todomn = MnTodo.query.get(PRID)
        if todomn:
            print('update Todo')
        else:
            print('add Todo')

def PRGetCustomerNameByDefault(PRJson, defaultName = 'Nokia'):
    if 'customerName' in PRJson.keys():
        CustomerName = PRJson['customerName']
    else:
        CustomerName = defaultName
    return CustomerName

def FindCustomerNameAB(s1, customerdict):
    rst = ''
    for row in customerdict:
        if row.SearchRule == 'ALL_CHAR_MATCH' and row.CustomerName == s1:
            rst = row.CustomerName
            break
        elif row.SearchRule == 'PRE_CHAR_MATCH':
            prestr = row.CustomerName[:row.CustomerName.index('*') if '*' in row.CustomerName else None]
            if prestr in s1 and s1.index(prestr) == 0:
                rst = s1
                break
    if rst == '':
        return False, ''
    return True, rst

class CustomerDict(db.Model):
    __tablename__ = 'customerdict'
    RCAFilterId = db.Column('RCAFilterId', db.Integer, primary_key=True)
    RCAFilterName = db.Column(db.String(128))
    CustomerId = db.Column('CustomerId', db.Integer, primary_key=True)
    CustomerName = db.Column(db.String(128))
    SearchRule = db.Column(db.String(20))
    Comment = db.Column(db.String(500))
    RCAFilterIdXh = db.Column(db.Integer)
    def __init__(self, RCAFilterId,RCAFilterName,CustomerId,CustomerName,SearchRule,Comment,RCAFilterIdXh):
        self.RCAFilterId = RCAFilterId
        self.RCAFilterName = RCAFilterName
        self.CustomerId = CustomerId
        self.CustomerName = CustomerName
        self.SearchRule = SearchRule
        self.Comment = Comment
        self.RCAFilterIdXh = RCAFilterIdXh


def checkPRnotgeneratebatch():
    # find mnrcatable
    # find jira, subtask, and not exists in mnrcametricstable
    lastYearDate = (datetime.datetime.now() + datetime.timedelta(days=-30)).strftime("%Y-%m-%d")
    link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=%s' % (0, lastYearDate)
    r = requests.get(link, verify=False, auth=('krma', 'Nokia123'))
    r_status = r.status_code
    if r_status != 200:
        return
    a = json.loads(r.text)
    prlist = a['values']
    num = len(prlist)
    total = a['total']
    startAt = 0
    while (total > 0):
        for i in range(num):
            try:
                PRID = prlist[i]['id']
                rst = MnTodo.query.get(PRID)
                if not rst :
                    BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(prlist[i])
                    state = prlist[i]['state']
                    if state == 'Closed' or state == 'First Correction Complete' or \
                            state == 'Final Response Sent':  # "Correction Plan Ready",
                        author = prlist[i]['author']
                        PRGroupInCharge = prlist[i]['groupIncharge']
                    else:
                        continue
                    todo_item = getGroupInChargeInfo(PRID, PRGroupInCharge)
                    if todo_item:
                        JIRAProject = todo_item.JIRAProject
                        RCAFilter = todo_item.RCAFilter
                        # Below for Only Customer (A+B)
                        if author != 'Electra':
                            continue
                        # FNRed = False
                        FNRed = isFNRedNew(prlist[i])
                        if FNRed == True and JIRAProject != 'CUDOCIPBL':
                            continue
                        print('lack pr:' + str(PRID))
                        if PRID in ['02408288','02399531','02357691']:
                            print('find!!!============================================' + str(PRID))
            except BaseException as ec1:
                print(ec1)
            total -= 1
        try:
            startAt = startAt + num
            total = total - num
            url = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=%s' % (startAt, lastYearDate)
            r = requests.get(url, verify=False, auth=('krma', 'Nokia123'))
            r_status = r.status_code
            if r_status != 200:
                print "if r_status != 200 Means Leftover statrAt is larger than than the end: ThreadId= %d getCustomerPr Total = %d num = %d "
                return
            a = json.loads(r.text)
            prlist = a['values']
            num = len(prlist)
        except BaseException as ec1:
            pass

def checkPR():
    # find mnrcatable
    # find jira, subtask, and not exists in mnrcametricstable
    rst = MnTodo.query.all()
    jira = getjira()
    lacknum = 0
    prnum = 0
    lastprid = ''
    deta = str(-30)+'d'
    for i in rst :
        issues = jira.search_issues(jql_str='project = MNRCA AND summary~"' + i.PRID + '" AND type = "5WhyRCA" and created >= ' + deta + '',maxResults=1)
        for j in issues:
            childissues = j.fields.subtasks
            for k in childissues:
                key = k.key
                issue1 = jira.issue(key)
                if str(issue1.fields.issuetype) == 'Analysis subtask' and issue1.fields.assignee is not None:
                    jiraissue = MnRcaMetrics.query.filter(*[MnRcaMetrics.PRID==str(i.PRID),MnRcaMetrics.JiraIssueId==key]).first()
                    if not jiraissue:
                        prnum += 1
                        lacknum += 1
                        print(key + ', PRID= ' + str(i.PRID) + ', parent task id = ' + str(j.key) + ', total lacknum=' + str(lacknum) + ', prnum=' + str(prnum))

def checkAP():
    # find mnrcatable
    # find jira, subtask, and not exists in mnrcametricstable
    rst = MnTodo.query.all()
    jira = getjira()
    lacknum = 0
    prnum = 0
    lastprid = ''
    deta = str(-80) + 'd'
    for i in rst:
        issues = jira.search_issues(
            jql_str='project = MNRCA AND summary~"' + i.PRID + '" AND type = "5WhyRCA" and created >= ' + deta + '',
            maxResults=1)
        print(i.PRID)
        for j in issues:
            childissues = j.fields.subtasks
            for k in childissues:
                key = k.key
                issue1 = jira.issue(key)
                if str(issue1.fields.issuetype) in ['Action for EDA','Action for RCA'] and issue1.fields.assignee is not None:
                    jiraissue = MnRcaMetrics.query.filter(
                        *[MnRcaMetrics.PRID == str(i.PRID), MnRcaMetrics.JiraIssueId == key]).first()
                    if not jiraissue:
                        prnum += 1
                        lacknum += 1
                        print(key + ', PRID= ' + str(i.PRID) + ', parent task id = ' + str(
                            j.key) + ', total lacknum=' + str(lacknum) + ', prnum=' + str(prnum))

def checkJiraissue():
    jira = getjira()
    issue = jira.issue('MNRCA-48740')
    createDate = issue.fields.created
    JiraIssueParentCreatedDate = createDate.split('T')[0]
    print(JiraIssueParentCreatedDate)

    issue = jira.issue('MNRCA-48741')
    Rca5WhyParentIssue = issue.fields.parent
    createDate = Rca5WhyParentIssue.fields.created
    JiraIssueParentCreatedDate = createDate.split('T')[0]
    print(JiraIssueParentCreatedDate)


def getSquadTribeInfoWithConn(JiraIssueAssignee, conn):
    AssignTo = JiraIssueAssignee.strip()
    if JiraIssueAssignee == '':
        return '', '', '', '', ''
    # conn = get_ldap_connection()
    filter = '(mail=%s)' % AssignTo
    # Title:'Squad Group Lead', 'Tribe Leader', 'NSB MN SRAN RD L2 HZH'
    # nsnTeamName:'NSB MN SRAN RD L2 HZH Arch Ops',
    # conn = app.config['ldap']
    attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
    base_dn = 'o=NSN'
    # conn = app.config['ldap']
    result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    if not result:
        return '', '', '', '', ''
    if 'title' in result[0][1].keys():
        title = result[0][1]['title'][0]
        title = title.split(',')[0]
    else:
        title = ''
    if 'displayName' in result[0][1].keys():
        displayName = result[0][1]['displayName'][0]
    else:
        displayName = ''
    if 'nsnTeamName' in result[0][1].keys():
        squadGroupName = result[0][1]['nsnTeamName'][0]
        # nsnTeamName=result[0][1]['nsnOperativeOrgName'][0]
    else:
        squadGroupName = 'External People'
    if 'nsnManagerAccountName' in result[0][1].keys():
        lineManagerAccountId = result[0][1]['nsnManagerAccountName'][0]
    else:
        print 'JiraAssigneeEmail=%s' % JiraIssueAssignee
        return '', squadGroupName, '', '', ''
    if len(title.split('Tribe Lead')) > 1 or len(title.split('Head of')) > 1:
    # if title in ['Tribe Leader', 'R&D Tribe Leader']:
        # lineDisplayName = lineDisplayName1
        # tribeLeadDisplayName = displayName
        tribeName = squadGroupName
        # lineManagerEmail = lineManagerEmail1
        return '', squadGroupName, '', '', tribeName

    filter = '(uid=%s)' % lineManagerAccountId
    attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
    # 'nsnTeamName','title'
    base_dn = 'o=NSN'
    lineResult = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    if 'title' in lineResult[0][1].keys():
        title1 = lineResult[0][1]['title'][0]
        title1 = title1.split(',')[0]
    else:
        title1 = ''
    if 'nsnTeamName' in lineResult[0][1].keys():
        squadGroupName1 = lineResult[0][1]['nsnTeamName'][0]
    else:
        squadGroupName1 = 'External Squad'
    if 'displayName' in lineResult[0][1].keys():
        lineDisplayName1 = lineResult[0][1]['displayName'][0]
    else:
        lineDisplayName1 = ''
    if 'mail' in lineResult[0][1].keys():
        lineManagerEmail1 = lineResult[0][1]['mail'][0]
    else:
        lineManagerEmail1 = ''
    if 'nsnManagerAccountName' in lineResult[0][1].keys():
        lineManagerAccountId1 = lineResult[0][1]['nsnManagerAccountName'][0]
    else:
        squadGroupName = squadGroupName1
        return lineManagerEmail1, squadGroupName, lineDisplayName1, '', ''

    # if title1 in ['Tribe Leader', 'R&D Tribe Leader','Head of NM RD UTF']:
    if len(title1.split('Tribe Lead')) > 1 or len(title1.split('Head of')) > 1:
        # lineDisplayName = lineDisplayName1
        # tribeLeadDisplayName = displayName
        tribeName = squadGroupName1
        # lineManagerEmail = lineManagerEmail1
        return lineManagerEmail1, squadGroupName, lineDisplayName1, lineDisplayName1, tribeName
    if len(title.split('Squad Group Lead')) > 1:
    # if title == 'Squad Group Lead':
        tribeName = squadGroupName1
        tribeLeadDisplayName = lineDisplayName1
        lineManagerEmail = lineManagerEmail1
        lineDisplayName = lineDisplayName1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName
    if len(title.split('Tribe Lead')) > 1 or len(title.split('Head of')) > 1:
    # if title in ['Tribe Leader', 'R&D Tribe Leader']:
        lineDisplayName = lineDisplayName1
        tribeLeadDisplayName = displayName
        tribeName = squadGroupName
        lineManagerEmail = lineManagerEmail1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName
    if len(title.split('Tribe Lead')) == 1 and len(title.split('Head of')) == 1 and \
        len(title.split('Squad Group Lead')) == 1:
    # if title not in ['Squad Group Lead', 'Tribe Leader', 'R&D Tribe Leader']:
        # Next Level Org:
        squadGroupName = squadGroupName1
        TribeLeadAccountId = lineManagerAccountId1
        filter = '(uid=%s)' % TribeLeadAccountId
        attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
        base_dn = 'o=NSN'
        TribeLeadResult = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
        tribeLeadDisplayName = TribeLeadResult[0][1]['displayName'][0]
        tribeName = TribeLeadResult[0][1]['nsnTeamName'][0]
        lineManagerEmail = lineManagerEmail1
        lineDisplayName = lineDisplayName1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName

def strip_non_ascii(string):
    ''' Returns the string without non ASCII characters'''
    stripped = (c for c in string if 0 < ord(c) < 127)
    return ''.join(stripped)

class SquadGroupAssessor(db.Model):
    __tablename__ = 'squadGroupAssessorTable'
    # assessorId = db.Column('assessorId', db.String(64), primary_key=True)
    assessorId = db.Column('assessorId', db.String(64), primary_key=True)
    # id = db.Column(db.Integer, primary_key=True)
    squadGroupRcaEdaAssessor = db.Column(db.String(64))
    tribeRcaEdaLeadAssessor = db.Column(db.String(64))
    assessorSquadGroupName = db.Column(db.String(64))
    assessorSquadGroupLead = db.Column(db.String(64))
    assessorTribe = db.Column(db.String(64))
    assessorTribeLead = db.Column(db.String(64))
    squadGroupEmailGroup = db.Column(db.String(64))
    registeredBy = db.Column(db.String(64))

    def __init__(self, assessorId,squadGroupRcaEdaAssessor,tribeRcaEdaLeadAssessor,assessorSquadGroupName,assessorSquadGroupLead,\
                 assessorTribe,assessorTribeLead,squadGroupEmailGroup,registeredBy):

        self.assessorId = assessorId
        self.squadGroupRcaEdaAssessor = squadGroupRcaEdaAssessor
        self.tribeRcaEdaLeadAssessor = tribeRcaEdaLeadAssessor
        self.assessorSquadGroupName = assessorSquadGroupName
        self.assessorSquadGroupLead = assessorSquadGroupLead
        self.assessorTribe = assessorTribe
        self.assessorTribeLead = assessorTribeLead
        self.squadGroupEmailGroup = squadGroupEmailGroup
        self.registeredBy = registeredBy

# sigle jira sync to mnrcametrics table
def syncjira(PRID):
    jira = getjira()
    conn = get_ldap_connection()

    issues = jira.search_issues(jql_str='project = MNPRCA AND type in \
        ("Analysis subtask","Action for RCA","Action for EDA")\
         and created >= ' + str(-5)+'d \
         and summary~"' + PRID + '"', maxResults=False)
    # issues = jira.search_issues(jql_str='project = MNPRCA AND type in \
    #     ("5WhyRCA")\
    #      and  summary~"' + PRID + '"', maxResults=False) #PR635549
    for issue in issues: #3861
        # try:
        JiraIssueId = str(issue.key)
        issue = jira.issue(JiraIssueId)
        parentExistFlag = True
        try:
            Rca5WhyParentIssue = issue.fields.parent
        except:
            parentExistFlag = False
            pass
        if parentExistFlag == False:
            continue
        else:
            JiraIssueParentTaskId = str(Rca5WhyParentIssue.key)
            summary = strip_non_ascii(Rca5WhyParentIssue.fields.summary)
            JiraIssueSummary = strip_non_ascii(Rca5WhyParentIssue.fields.summary)
            Rca5WhyParentIssue = jira.issue(JiraIssueParentTaskId)
            if Rca5WhyParentIssue.fields.issuetype.name == '5WhyRCA':
                PRAttached = Rca5WhyParentIssue.fields.customfield_37060
            else:
                continue
        JiraIssueType = issue.fields.issuetype.name
        JiraIssueCaseType = ''
        if JiraIssueType in ('Analysis subtask',):
            JiraIssueCaseType = str(issue.fields.customfield_10464)
        JiraIssueStatus = issue.fields.status.name
        JiraIssueParentStatus = Rca5WhyParentIssue.fields.status.name
        JiraIssueAssignee = str(issue.fields.assignee)
        JiraIssueReporter = issue.fields.reporter.displayName

        createDate = issue.fields.created
        JiraIssueCreatedDate = createDate.split('T')[0]
        createDate = Rca5WhyParentIssue.fields.created
        JiraIssueParentCreatedDate = createDate.split('T')[0]

        if JiraIssueParentCreatedDate < '2019-01-01':
            todo_item = MnRcaMetrics.query.get(JiraIssueId)
            if todo_item:
                db.session.delete(todo_item)
                db.session.commit()
            continue
        # PRID = ''
        FAID = 'FAID'
        # PRTitle = ''
        PRAttached = PRAttached
        # PRSeverity = ''
        # PRGroupInCharge = ''
        # ReportedBy = ''
        ProntoNumber = Rca5WhyParentIssue.fields.customfield_37060
        # if parentExistFlag:
        try:
            # PRID =summary.split('for',2)[1].strip()
            PRID = summary
            PRID = PRID.split(':')
            PRID = PRID[0].strip()
            PRID = PRID.split(' ')
            PRID = PRID[0]
        except:
            print "Bad@@@@@@@@@@@@@@@@" + summary
            continue
        PRID = PRID
        if len(PRID) > 64:
            PRID = PRID[:64]
        PRTitle = summary
        PRGroupInCharge = ''
        PRSeverity = ''
        url = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/problemReport/%s" % PRID
        try1 = False
        try:
            r = requests.get(url, verify=False, auth=(gAdmin, gAuth))
            a = r.json()

            PRGroupInCharge = a['groupIncharge']
            PRSeverity = a['severity']
            PRAttached = a['problemReportIds']
            author = a['author']
            if not PRAttached:
                PRAttached = ''
            else:
                LabelItems = PRAttached
                b = ''
                for s in LabelItems:
                    b = b + s + ','
                    PRAttached = b.strip(',')
            PRTitle = a['title']
            FAID = a['faultAnalysisId']
            # ReportedBy = a['reportedBy']
        except:
            try1 = True
            pass
        if try1:
            try2 = False
        try:
            PRID1 = ProntoNumber.split(',')[0]
            PRID = PRID1.split(' ')[0]
            if len(PRID) > 64:
                PRID = PRID[:64]
            url = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/problemReport/%s" % PRID
            r = requests.get(url, verify=False, auth=(gAdmin, gAuth))
            a = r.json()

            PRGroupInCharge = a['groupIncharge']
            PRSeverity = a['severity']
            author = a['author']
            PRAttached = a['problemReportIds']
            if not PRAttached:
                PRAttached = ''
            else:
                LabelItems = PRAttached
                b = ''
                for s in LabelItems:
                    b = b + s + ','
                    PRAttached = b.strip(',')
            PRTitle = a['title']
            FAID = a['faultAnalysisId']
            # ReportedBy = a['reportedBy']
        except:
            try2 = True
            pass

        if try1 == True and try2 == True:
            continue
        if author == 'Electra':
            ReportedBy = 'Customer'
        else:
            ReportedBy = 'Nokia'
        if issue.fields.assignee is None:
            continue
        JiraIssueAssigneeEmail = ''
        if issue.fields.assignee:
            JiraIssueAssigneeEmail = issue.fields.assignee.emailAddress
        # print '!!!JiraIssueAssigneeEmail =%s!!!' % JiraIssueAssigneeEmail
        # LineManagerEmail,squadGroup,lineDisplayName,tribeLeadDisplayName,tribeName = getSquadTribeInfo(JiraIssueAssigneeEmail)
        LineManagerEmail, squadGroup, lineDisplayName, tribeLeadDisplayName, tribeName = getSquadTribeInfoWithConn(
            JiraIssueAssigneeEmail, conn)
        PRRcaEdaAssessor = ''
        squadGroupAssessorInfo = SquadGroupAssessor.query.filter_by(assessorSquadGroupLead=LineManagerEmail). \
            order_by(SquadGroupAssessor.assessorId.desc()).first()
        if squadGroupAssessorInfo:
            assessorEmailAddress = squadGroupAssessorInfo.squadGroupRcaEdaAssessor
            PRRcaEdaAssessor = assessorEmailAddress
            JiraIssueAssigneeSquadGroupLead = lineDisplayName
            JiraIssueAssigneeTribeLead = tribeLeadDisplayName  # "zhuofei.chen@nokia-sbell.com"
            JiraIssueAssigneeTribe = tribeName
            JiraIssueAssigneeSquadGroup = squadGroup
        else:
            PRRcaEdaAssessor = lineDisplayName
            JiraIssueAssigneeSquadGroupLead = lineDisplayName
            JiraIssueAssigneeTribeLead = tribeLeadDisplayName  # "zhuofei.chen@nokia-sbell.com"
            JiraIssueAssigneeTribe = tribeName
            JiraIssueAssigneeSquadGroup = squadGroup
        JiraIssueDueDate = ''
        dueDate = ''
        # print "Other type"
        if JiraIssueType == 'Analysis subtask':
            dueDate = issue.fields.customfield_27792
        if JiraIssueType == 'Action for EDA':
            dueDate = issue.fields.customfield_37058
        if JiraIssueType == 'Action for RCA':
            dueDate = issue.fields.customfield_37070
        if dueDate:
            JiraIssueDueDate = dueDate.split('T')[0]
        current_date = datetime.datetime.now()  # + datetime.timedelta(days=days)
        current_date = current_date.strftime('%Y-%m-%d')
        JiraIssueResolutionDate = current_date
        JiraIssueParentResolutionDate = current_date
        if JiraIssueStatus in ['Resolved', 'Closed', 'RCA Done', 'EDA Done', 'RCA Action Done']:
            if issue.fields.resolutiondate == None:
                resolutiondate = str(issue.fields.updated)
            else:
                resolutiondate = str(issue.fields.resolutiondate)
            JiraIssueResolutionDate = resolutiondate.split('T')[0]
        fiveWhyRcaStatus = 'Not Approved'
        if JiraIssueParentStatus in ['Resolved', 'Closed']:
            resolutiondate = str(Rca5WhyParentIssue.fields.resolutiondate)
            JiraIssueParentResolutionDate = resolutiondate.split('T')[0]
            fiveWhyRcaStatus = 'Approved'
        JiraIssueOpenDays = daysBetweenDate(JiraIssueCreatedDate, JiraIssueResolutionDate)
        JiraIssueOverDue = 'No'
        if JiraIssueDueDate == '':
            if JiraIssueOpenDays > 15:
                JiraIssueOverDue = 'Yes'
        else:
            if JiraIssueResolutionDate > JiraIssueDueDate:
                JiraIssueOverDue = 'Yes'
        JiraIssueParentOpenDays = 0
        try:
            JiraIssueParentOpenDays = daysBetweenDate(JiraIssueParentCreatedDate, JiraIssueParentResolutionDate)
        except:
            pass
        JiraIssueParentOverDue = 'No'
        if JiraIssueParentOpenDays > 30:
            JiraIssueParentOverDue = 'Yes'

        PRRcaEdaAssessor = PRRcaEdaAssessor

        PRTitle = strip_non_ascii(PRTitle)
        todo_item = MnRcaMetrics.query.get(JiraIssueId)
        if todo_item:
            # print 'Existing%s!!!!'%PRID
            todo_item.PRID = PRID
            todo_item.PRTitle = PRTitle
            todo_item.FAID = FAID
            todo_item.PRAttached = PRAttached
            todo_item.JiraIssueStatus = JiraIssueStatus
            todo_item.JiraIssueCaseType = JiraIssueCaseType
            todo_item.JiraIssueParentCreatedDate = JiraIssueParentCreatedDate
            todo_item.JiraIssueParentStatus = JiraIssueParentStatus
            todo_item.JiraIssueOpenDays = JiraIssueOpenDays
            todo_item.JiraIssueParentOpenDays = JiraIssueParentOpenDays
            todo_item.JiraIssueResolutionDate = JiraIssueResolutionDate
            todo_item.JiraIssueParentResolutionDate = JiraIssueParentResolutionDate
            todo_item.JiraIssueDueDate = JiraIssueDueDate
            todo_item.JiraIssueAssignee = JiraIssueAssignee
            todo_item.PRRcaEdaAssessor = PRRcaEdaAssessor
            todo_item.JiraIssueAssigneeSquadGroup = JiraIssueAssigneeSquadGroup
            todo_item.JiraIssueAssigneeSquadGroupLead = JiraIssueAssigneeSquadGroupLead
            todo_item.JiraIssueAssigneeTribe = JiraIssueAssigneeTribe
            todo_item.JiraIssueAssigneeTribeLead = JiraIssueAssigneeTribeLead
            todo_item.PRAttached = PRAttached
            todo_item.JiraIssueOverDue = JiraIssueOverDue
            todo_item.JiraIssueParentOverDue = JiraIssueParentOverDue
            todo_item.ReportedBy = ReportedBy
            todo_item.PRSeverity = PRSeverity
            todo_item.JiraIssueOverDueReason = fiveWhyRcaStatus
            db.session.commit()
        else:
            JiraIssueOverDueReason = ''
            JiraIssueParentOverDueReason = ''
            todo = MnRcaMetrics(JiraIssueId, JiraIssueParentTaskId, JiraIssueSummary, JiraIssueType,
                                JiraIssueCaseType, \
                                JiraIssueStatus, JiraIssueParentStatus, JiraIssueAssignee, JiraIssueReporter, PRID,
                                FAID, PRTitle, \
                                PRRcaEdaAssessor, PRAttached, PRSeverity, PRGroupInCharge, JiraIssueCreatedDate,
                                JiraIssueParentCreatedDate, \
                                JiraIssueResolutionDate, JiraIssueParentResolutionDate, JiraIssueDueDate,
                                JiraIssueOpenDays, \
                                JiraIssueParentOpenDays, JiraIssueAssigneeSquadGroup,
                                JiraIssueAssigneeSquadGroupLead, \
                                JiraIssueAssigneeTribe, JiraIssueAssigneeTribeLead, JiraIssueOverDue,
                                JiraIssueParentOverDue, \
                                JiraIssueOverDueReason, JiraIssueParentOverDueReason, ReportedBy)
            db.session.add(todo)
            db.session.commit()
            counting = counting + 1
            print 'Action now!'
            print counting

def testjiracomment():
    jira = getjira()
    # comment_to_edit = jira.add_comment('MNRCA-46700', 'Li, Mengxiang (NSB - CN/Hangzhou) : Change this content later')
    # comment_to_edit.update(body='New Content.')

    # comment_to_edit = jira.comment('MNRCA-46700', '11462801')
    # comment_to_edit.update(body='Edit the content to something new.')

    # comment_to_edit = jira.comment('MNRCA-46700', '11462830')
    # comment_to_edit.delete()

    for i in ['11479285','11479480','11479572','11479483']:
        coms = jira.comment('MNRCA-46699',i)
        coms.delete()

    # print(comment_to_edit)

def getSquadTribeInfo(JiraIssueAssignee):
    AssignTo = JiraIssueAssignee.strip()
    if JiraIssueAssignee == '':
        return '', '', '', '', ''
    conn = get_ldap_connection()
    filter = '(mail=%s)' % AssignTo
    # Title:'Squad Group Lead', 'Tribe Leader', 'NSB MN SRAN RD L2 HZH'
    # nsnTeamName:'NSB MN SRAN RD L2 HZH Arch Ops',
    # conn = app.config['ldap']
    attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
    base_dn = 'o=NSN'
    # conn = app.config['ldap']
    result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    if not result:
        return '', '', '', '', ''
    if 'title' in result[0][1].keys():
        title = result[0][1]['title'][0]
        title = title.split(',')[0]
    else:
        title = ''
    if 'displayName' in result[0][1].keys():
        displayName = result[0][1]['displayName'][0]
    else:
        displayName = ''
    if 'nsnTeamName' in result[0][1].keys():
        squadGroupName = result[0][1]['nsnTeamName'][0]
        # nsnTeamName=result[0][1]['nsnOperativeOrgName'][0]
    else:
        squadGroupName = 'External People'
    if 'nsnManagerAccountName' in result[0][1].keys():
        lineManagerAccountId = result[0][1]['nsnManagerAccountName'][0]
    else:
        print 'JiraAssigneeEmail=%s' % JiraIssueAssignee
        return '', squadGroupName, '', '', ''
    filter = '(uid=%s)' % lineManagerAccountId
    attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
    # 'nsnTeamName','title'
    base_dn = 'o=NSN'
    lineResult = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
    if 'nsnTeamName' in lineResult[0][1].keys():
        squadGroupName1 = lineResult[0][1]['nsnTeamName'][0]
    else:
        squadGroupName1 = 'External Squad'
    if 'displayName' in lineResult[0][1].keys():
        lineDisplayName1 = lineResult[0][1]['displayName'][0]
    else:
        lineDisplayName1 = ''
    if 'mail' in lineResult[0][1].keys():
        lineManagerEmail1 = lineResult[0][1]['mail'][0]
    else:
        lineManagerEmail = ''
    if 'nsnManagerAccountName' in lineResult[0][1].keys():
        lineManagerAccountId1 = lineResult[0][1]['nsnManagerAccountName'][0]
    else:
        squadGroupName = squadGroupName1
        return lineManagerEmail, squadGroupName, lineDisplayName1, '', ''
    if title == 'Squad Group Lead':
        tribeName = squadGroupName
        tribeLeadDisplayName = lineDisplayName1
        lineManagerEmail = lineManagerEmail1
        lineDisplayName = lineDisplayName1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName
    if title == 'Tribe Leader':
        lineDisplayName = lineDisplayName1
        tribeLeadDisplayName = displayName
        tribeName = squadGroupName
        lineManagerEmail = lineManagerEmail1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName
    if title not in ['Squad Group Lead', 'Tribe Leader']:
        # Next Level Org:
        squadGroupName = squadGroupName1
        TribeLeadAccountId = lineManagerAccountId1
        filter = '(uid=%s)' % TribeLeadAccountId
        attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
        base_dn = 'o=NSN'
        TribeLeadResult = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
        tribeLeadDisplayName = TribeLeadResult[0][1]['displayName'][0]
        tribeName = TribeLeadResult[0][1]['nsnTeamName'][0]
        lineManagerEmail = lineManagerEmail1
        lineDisplayName = lineDisplayName1
        return lineManagerEmail, squadGroupName, lineDisplayName, tribeLeadDisplayName, tribeName


def singleton(cls, *args, **kwargs):
    instances = {}

    def _singleton():
        if cls not in instances:
            instances[cls] = cls(*args, **kwargs)
        return instances[cls]

    return _singleton


@singleton
class RcaLdap(object):
    def __init__(self):
        self.conninstance = get_ldap_connectionnew()
    def getldap(self):
        if self.testconnection():
            return self.conninstance
        else :
            self.conninstance = get_ldap_connectionnew()
            print('reconnect ldap')
            return self.conninstance
    def testconnection(self):
        try:
            filter = '(mail=%s)' % 'mengxiang.li@nokia-sbell.com'
            attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
            base_dn = 'o=NSN'
            result = self.conninstance.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
            if 'displayName' in result[0][1].keys():
                displayName = result[0][1]['displayName'][0]
            if displayName:
                return True
            else :
                return False
        except:
            print(traceback.format_exc())
            return False


def testConnSingle():
    init_date = datetime.datetime.now()
    while(True):
        conn = RcaLdap().getldap()
        filter = '(mail=%s)' % 'mengxiang.li@nokia-sbell.com'
        attrs = ['mail', 'displayName', 'nsnManagerAccountName', 'nsnTeamName', 'title']
        base_dn = 'o=NSN'
        result = conn.search_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs)
        if 'displayName' in result[0][1].keys():
            displayName = result[0][1]['displayName'][0]
        current_date = datetime.datetime.now()
        print(displayName + ', ' + str((current_date-init_date).seconds) + 's')
        time.sleep(random.randint(1,60))

def testJiraGrade():
    jira = getjira()
    coms = jira.issue('MNRCA-46699')
    # dictgrade = {'customfield_37464': {"id":"261614","value":"A – Excellent"}}
    dictgrade = {'customfield_37464': {"id":"261618","value":"N/A – Grade not applicable"}}
    coms.update(dictgrade)

    # print(comment_to_edit)

def addToMnRca(PRID,urlfa,c):
    # flg=False
    # try:
    r = requests.get(urlfa, verify=False, auth=('krma', 'Nokia123'))
    # except:
    #     flg=True
    #     pass
    # if flg:
    #     return
    # urlcf = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/correction/%s" % correctionId
    # try:
    # print 'isFNRedNew  correctionId= %s!!!!!!!!!!!!!!!!!!!!!' % correctionId
    # r = requests.get(urlcf, verify=False, auth=('krma', 'Nokia123'))
    r_status=r.status_code
    if r_status != 200:
        return
    # b = r.text
    b = json.loads(r.text)
    # b = r.json() # FA json
    if 'rcaEdaDecision' in b.keys():
        rcaEdaDecision = b['rcaEdaDecision']
    else:
        return
    FaRcaEdaStatus = b['rcaEda']['status']
    if rcaEdaDecision == 'Yes' and FaRcaEdaStatus == None:
        # FaRcaEdaDecisionFlag=rcaEdaDecision
        # c = APIPr(PRID)
        # PRID = c['id']
        BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
        PRReportedDate = c['reportedDate'].split('T')[0]
        PRGroupInCharge = c['groupIncharge']
        PRClosedDate = c['lastMod'].split('T')[0]
        dateFlag, realDate = getPRCloseDate(PRID)
        if dateFlag:
            PRClosedDate = realDate
            # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
        PRTitle = c['title']
        ReportedBy = c['reportedBy']
        FaultCoordinator = c['devFaultCoord']
        PRSeverity = c['severity']
        if PRSeverity == 'A - Critical':
            rank = 1
        elif PRSeverity == 'B - Major':
            rank = 2
        elif PRSeverity == 'C - Minor':
            rank = 3
        PRRelease = c['softwareRelease']
        Feature = c['feature']
        PRProduct = c['product']
        if 'customerName' in c.keys():
            CustomerName = c['customerName']
        else:
            CustomerName = 'Nokia'
        PRAttached = c['problemReportIds']
        attachedlist = PRAttached
        if not PRAttached:
            PRAttached = ''
        else:
            LabelItems = PRAttached
            b = ''
            for s in LabelItems:
                b = b + s + ','
                PRAttached = b.strip(',')
        for item in attachedlist:
            c = APIPr(item)
            # if 'reportedBy' in c.keys():
            #     AttachedreportedBy = c['reportedBy']
            # else:
            #     AttachedreportedBy = ''
            # if AttachedreportedBy == 'Customer':
            attachedSeverity = c['severity']
            if attachedSeverity == 'A - Critical':
                rank1 = 1
            elif attachedSeverity == 'B - Major':
                rank1 = 2
            elif attachedSeverity == 'C - Minor':
                rank1 = 3
            if rank1 == rank:
                AttachedPRReportedDate = c['reportedDate'].split('T')[0]
                if AttachedPRReportedDate < PRReportedDate:
                    PRID = c['id']
                    BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                    PRReportedDate = c['reportedDate'].split('T')[0]
                    PRGroupInCharge = c['groupIncharge']
                    # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                    PRClosedDate = c['lastMod'].split('T')[0]
                    dateFlag, realDate = getPRCloseDate(PRID)
                    if dateFlag:
                        PRClosedDate = realDate
                        # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                    PRTitle = c['title']
                    ReportedBy = c['reportedBy']
                    FaultCoordinator = c['devFaultCoord']
                    PRSeverity = c['severity']
                    if PRSeverity == 'A - Critical':
                        rank = 1
                    elif PRSeverity == 'B - Major':
                        rank = 2
                    elif PRSeverity == 'C - Minor':
                        rank = 3
                    PRRelease = c['softwareRelease']
                    Feature = c['feature']
                    PRAttached = c['problemReportIds']
                    if not PRAttached:
                        PRAttached = ''
                    else:
                        LabelItems = PRAttached
                        b = ''
                        for s in LabelItems:
                            b = b + s + ','
                            PRAttached = b.strip(',')
                    PRProduct = c['product']
                    if 'customerName' in c.keys():
                        CustomerName = c['customerName']
                    else:
                        CustomerName = 'Nokia'
            elif rank1 < rank:
                PRID = c['id']
                BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                PRReportedDate = c['reportedDate'].split('T')[0]
                PRGroupInCharge = c['groupIncharge']
                # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                PRClosedDate = c['lastMod'].split('T')[0]
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                PRTitle = c['title']
                ReportedBy = c['reportedBy']
                FaultCoordinator = c['devFaultCoord']
                PRSeverity = c['severity']
                if PRSeverity == 'A - Critical':
                    rank = 1
                elif PRSeverity == 'B - Major':
                    rank = 2
                elif PRSeverity == 'C - Minor':
                    rank = 3
                PRRelease = c['softwareRelease']
                Feature = c['feature']
                PRAttached = c['problemReportIds']
                if not PRAttached:
                    PRAttached = ''
                else:
                    LabelItems = PRAttached
                    b = ''
                    for s in LabelItems:
                        b = b + s + ','
                        PRAttached = b.strip(',')
                PRProduct = c['product']
                if 'customerName' in c.keys():
                    CustomerName = c['customerName']
                else:
                    CustomerName = 'Nokia'
        # PRProduct = c['product']
        # CustomerName = 'Nokia'
        # groupInChargeInfo = InChargeGroup.query.get(PRGroupInCharge)
        # BusinessLine,ProductLine = getBusinessLineAndProductLine(PRID)
        # # groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName = PRGroupInCharge, \
        # #                         BusinessLine = BusinessLine,ProductLine = ProductLine).first()
        # groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
        #                                                  BusinessLine=BusinessLine,
        #                                                  ProductLine='All').first()
        # if not groupInChargeInfo:
        #     groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
        #                                                      BusinessLine=BusinessLine,
        #                                                      ProductLine=ProductLine).first()
        # if not groupInChargeInfo:
        #     groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
        #                                                      AddedBy='sufang.huang@nokia-sbell.com').first()
        groupInChargeInfo = getGroupInChargeInfo(PRID, PRGroupInCharge)
        JIRAProject = groupInChargeInfo.JIRAProject
        RCAFilter = groupInChargeInfo.RCAFilter
        JiraCustomerName = CustomerName
        # ProductLine = FindProductLine(PRProduct)
        if PRAttached:
            attachedList = PRAttached.split(',')
            for item in attachedList:
                todo_item = MnTodo.query.get(item)
                if todo_item:
                    db.session.delete(todo_item)
                    db.session.commit()
        # 'All Customer, No internal'
        # 'All Customer, Partial Internal according to FA rcaEdaDecision Flag'
        # 'Partial Customer and Internal, both according to FA rcaEdaDecision Flag'
        # 'No Internal, Partial Customer according to FA rcaEdaDecision Flag'
        if RCAFilter == 'Partial Customer(A+B) and Internal, both according to FA rcaEdaDecision Flag'\
            and CustomerName != 'Nokia':
            JiraCustomerName = FindCustomerName(CustomerName.lower())
        if RCAFilter == 'No Internal, Partial Customer(A+B) according to FA rcaEdaDecision Flag':
            JiraCustomerName = FindCustomerName(CustomerName.lower())
        # if RCAFilter == 'No Internal, Partial Customer according to FA rcaEdaDecision Flag':
        BusinessLine = groupInChargeInfo.BusinessLine
        if groupInChargeInfo.ProductLine != 'All':
            ProductLine = groupInChargeInfo.ProductLine
        # if groupInChargeInfo.BusinessLine == 'LTE' and PRRelease:
        #     if (len(PRRelease.split('TL')) != 1):
        #         ProductLine = 'TD LTE'
        #     elif (len(PRRelease.split('FL')) != 1):
        #         ProductLine = 'FDD LTE'
        #     elif (len(PRRelease.split('CBTS')) != 1):
        #         ProductLine = 'Cloud BTS'
        #     elif (len(PRRelease.split('SBTS')) != 1):
        #         ProductLine = 'SRAN/SBTS'
        #         BusinessLine = 'SRAN'
        # if groupInChargeInfo.BusinessLine == 'LTE' and PRProduct=='Flexi WCDMA Base Station':
        #     ProductLine = 'WBTS'
        #     BusinessLine = 'SRAN'
        todomn= MnTodo.query.get(PRID)
        if todomn:
            todomn.ProductLine = ProductLine
            todomn.JiraCustomerName = JiraCustomerName
            # todomn.RCAEDACategory = groupInChargeInfo.ProductLine
            todomn.RCAEDACategory = groupInChargeInfo.RCAEDACategory
            todomn.PRClosedDate = PRClosedDate
            todomn.PRAttached = PRAttached
            todomn.JIRAProject = JIRAProject
            todomn.FaRcaEdaDecisionFlag = rcaEdaDecision
            todomn.PRSeverity = PRSeverity
            db.session.commit()
        else:
            PRRcaEdaAssessor = groupInChargeInfo.AssessorEmail
            EDA_AssessorEmail = groupInChargeInfo.EDA_AssessorEmail
            EDACreatingTime = groupInChargeInfo.EDACreatingTime
            EdaCaseType = groupInChargeInfo.EdaCaseType
            BusinessUnit = groupInChargeInfo.BusinessUnit
            # BusinessLine = groupInChargeInfo.BusinessLine
            # ProductLine = groupInChargeInfo.ProductLine
            RCAEDACategory = groupInChargeInfo.RCAEDACategory
            FaRcaEdaDecisionFlag = rcaEdaDecision
            JiraTaskIsCreated = 'No'
            todo = MnTodo(PRID,PRTitle,PRReportedDate,PRClosedDate,PRRelease,PRSeverity,PRGroupInCharge,PRAttached,PRRcaEdaAssessor,\
                 PRProduct,ReportedBy,FaultCoordinator,CustomerName,BusinessUnit,BusinessLine,ProductLine,Feature,RCAEDACategory,\
                          JiraCustomerName,JIRAProject,FaRcaEdaDecisionFlag,EDA_AssessorEmail,EDACreatingTime,EdaCaseType, \
                          JiraTaskIsCreated)
            db.session.add(todo)
            db.session.commit()

def addToMnRcaCustomerABnew(PRID,probj,author,JIRAProject):
    # Below for Only Customer (A+B)
    FNRed = False
    if author != 'Electra':
        return
    FNRed = isFNRedNew(probj)
    if FNRed == True and JIRAProject != 'CUDOCIPBL':
        todo_item = MnTodo.query.get(PRID)
        if todo_item:
            db.session.delete(todo_item)
            db.session.commit()
        return
    PRReportedDate = probj['reportedDate'].split('T')[0]
    PRClosedDate = probj['lastMod'].split('T')[0]
    # print PRID
    try:
        dateFlag, realDate = getPRCloseDate(PRID)
        if dateFlag:
            PRClosedDate = realDate
        #     print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
        # else:
        #     print 'No '
    except:
        pass

    PRTitle = probj['title']
    ReportedBy = probj['reportedBy']
    PRGroupInCharge = probj['groupIncharge']
    FaultCoordinator = probj['devFaultCoord']
    PRSeverity = probj['severity']
    HightestLevel = -1         #the hightest level in self pr and attach pr
    if PRSeverity == 'A - Critical':
        HightestLevel = 1
    elif PRSeverity == 'B - Major':
        HightestLevel = 2
    elif PRSeverity == 'C - Minor':
        HightestLevel = 3
    PRRelease = probj['softwareRelease']
    Feature = probj['feature']
    PRProduct = probj['product']
    JiraCustomerName = probj['customerName']
    if 'customerName' in probj.keys():
        CustomerName = probj['customerName']
    else:
        CustomerName = 'Nokia'
    PRAttached = probj['problemReportIds']
    attachedlist = PRAttached
    if not PRAttached:
        PRAttached = ''
    else:
        LabelItems = PRAttached
        b = ''
        for s in LabelItems:
            b = b + s + ','
            PRAttached = b.strip(',')
    for item in attachedlist:
        c = APIPr(item)
        Attachedauthor = c['author']
        if 'reportedBy' in c.keys():
            AttachedreportedBy = c['reportedBy']
        else:
            AttachedreportedBy = ''
        if Attachedauthor == 'Electra':
            attachedSeverity = c['severity']
            if attachedSeverity == 'A - Critical':
                rank1 = 1
            elif attachedSeverity == 'B - Major':
                rank1 = 2
            elif attachedSeverity == 'C - Minor':
                rank1 = 3
            if rank1 == HightestLevel:
                AttachedPRReportedDate = c['reportedDate'].split('T')[0]
                if AttachedPRReportedDate < PRReportedDate:
                    PRID = c['id']
                    PRReportedDate = c['reportedDate'].split('T')[0]
                    PRGroupInCharge = c['groupIncharge']
                    # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                    PRClosedDate = c['lastMod'].split('T')[0]
                    dateFlag, realDate = getPRCloseDate(PRID)
                    if dateFlag:
                        PRClosedDate = realDate
                        # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                    PRTitle = c['title']
                    ReportedBy = c['reportedBy']
                    FaultCoordinator = c['devFaultCoord']
                    PRSeverity = c['severity']
                    if PRSeverity == 'A - Critical':
                        HightestLevel = 1
                    elif PRSeverity == 'B - Major':
                        HightestLevel = 2
                    elif PRSeverity == 'C - Minor':
                        HightestLevel = 3
                    PRRelease = c['softwareRelease']
                    Feature = c['feature']
                    PRAttached = c['problemReportIds']
                    if not PRAttached:
                        PRAttached = ''
                    else:
                        LabelItems = PRAttached
                        b = ''
                        for s in LabelItems:
                            b = b + s + ','
                            PRAttached = b.strip(',')
                    PRProduct = c['product']
                    JiraCustomerName = c['customerName']
                    if 'customerName' in c.keys():
                        CustomerName = c['customerName']
                    else:
                        CustomerName = 'Nokia'
            elif rank1 < HightestLevel:
                PRID = c['id']
                PRReportedDate = c['reportedDate'].split('T')[0]
                PRGroupInCharge = c['groupIncharge']
                # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                PRClosedDate = c['lastMod'].split('T')[0]
                print 'before getPRCloseDate  PRID= %s!!!!!!!!!!!!!!!!!!!!!' % PRID
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                PRTitle = c['title']
                ReportedBy = c['reportedBy']
                FaultCoordinator = c['devFaultCoord']
                PRSeverity = c['severity']
                if PRSeverity == 'A - Critical':
                    HightestLevel = 1
                elif PRSeverity == 'B - Major':
                    HightestLevel = 2
                elif PRSeverity == 'C - Minor':
                    HightestLevel = 3
                PRRelease = c['softwareRelease']
                Feature = c['feature']
                PRAttached = c['problemReportIds']
                if not PRAttached:
                    PRAttached = ''
                else:
                    LabelItems = PRAttached
                    b = ''
                    for s in LabelItems:
                        b = b + s + ','
                        PRAttached = b.strip(',')
                PRProduct = c['product']
                JiraCustomerName = c['customerName']
                if 'customerName' in c.keys():
                    CustomerName = c['customerName']
                else:
                    CustomerName = 'Nokia'
        else:
            print 'Internal PR, ignore'
    BusinessLine, ProductLine = getBusinessLineAndProductLine(PRID)
    groupInChargeInfo = getGroupInChargeInfo(PRID, PRGroupInCharge)
    if groupInChargeInfo:
        JIRAProject = groupInChargeInfo.JIRAProject
        MatchFlag = False
        if HightestLevel == 1:
            MatchFlag = True
        elif HightestLevel == 2:
            customerDict = CustomerDict.query.filter_by(RCAFilterId='A_KEY_B').all()
            proik = APIPr(PRID)
            MatchFlag,cusname = FindCustomerNameAB(PRGetCustomerNameByDefault(proik), customerDict)
            if not MatchFlag:
                for ik in PRAttached.split(','):
                    proik = APIPr(ik)
                    MatchFlag,cusname = FindCustomerNameAB(PRGetCustomerNameByDefault(proik), customerDict)
                    if MatchFlag :
                        break
        elif HightestLevel == 3:
            MatchFlag = False
        if not MatchFlag:
            return
        BusinessLine = groupInChargeInfo.BusinessLine
        if groupInChargeInfo.ProductLine != 'All':
            ProductLine = groupInChargeInfo.ProductLine
        if PRAttached:
            attachedList = PRAttached.split(',')
            for item in attachedList:
                todo_item = MnTodo.query.get(item)
                if todo_item:
                    db.session.delete(todo_item)
                    db.session.commit()
        todomn = MnTodo.query.get(PRID)
        if todomn:
            todomn.ProductLine = ProductLine
            todomn.JiraCustomerName = JiraCustomerName
            todomn.RCAEDACategory = groupInChargeInfo.RCAEDACategory
            todomn.PRClosedDate = PRClosedDate
            todomn.PRAttached = PRAttached
            todomn.JIRAProject = JIRAProject
            todomn.PRSeverity = PRSeverity
            db.session.commit()
            todo_item = Todo.query.get(PRID)
            if todo_item:
                todo_item.PRSeverity = todomn.PRSeverity
                todo_item.PRAttached = todomn.PRAttached
                todo_item.PRGroupInCharge = todomn.PRGroupInCharge
                todo_item.PRProduct = todomn.PRProduct
                todo_item.ReportedBy = todomn.ReportedBy
                todo_item.FaultCoordinator = todomn.FaultCoordinator
                todo_item.CustomerName = todomn.CustomerName
                todo_item.Feature = todomn.Feature
                todo_item.JiraCustomerName = todomn.JiraCustomerName
                todo_item.JIRAProject = todomn.JIRAProject
                db.session.commit()
        else:
            PRRcaEdaAssessor = groupInChargeInfo.AssessorEmail
            EDA_AssessorEmail = groupInChargeInfo.EDA_AssessorEmail
            EDACreatingTime = groupInChargeInfo.EDACreatingTime
            EdaCaseType = groupInChargeInfo.EdaCaseType
            BusinessUnit = groupInChargeInfo.BusinessUnit
            # ProductLine = groupInChargeInfo.ProductLine
            RCAEDACategory = groupInChargeInfo.RCAEDACategory
            FaRcaEdaDecisionFlag = ''
            JiraTaskIsCreated = 'No'
            todo = MnTodo(PRID, PRTitle, PRReportedDate, PRClosedDate, PRRelease, PRSeverity, PRGroupInCharge,
                          PRAttached, PRRcaEdaAssessor, \
                          PRProduct, ReportedBy, FaultCoordinator, CustomerName, BusinessUnit, BusinessLine,
                          ProductLine, Feature, RCAEDACategory, \
                          JiraCustomerName, JIRAProject, FaRcaEdaDecisionFlag, EDA_AssessorEmail, EDACreatingTime,
                          EdaCaseType, JiraTaskIsCreated)
            # todo.jirauser = g.user
            db.session.add(todo)
            db.session.commit()
            todo_item = Todo.query.get(PRID)
            if todo_item:
                todo_item.PRSeverity = PRSeverity
                todo_item.PRAttached = PRAttached
                todo_item.PRGroupInCharge = PRGroupInCharge
                todo_item.PRProduct = PRProduct
                todo_item.ReportedBy = ReportedBy
                todo_item.FaultCoordinator = FaultCoordinator
                todo_item.CustomerName = CustomerName
                todo_item.Feature = Feature
                todo_item.JiraCustomerName = JiraCustomerName
                todo_item.JIRAProject = JIRAProject
            db.session.commit()

def debugProntoGenerator(PRID):
    probj = APIPr(PRID)
    try:
        # print '********************* i=%d @@@@@@@@@@@@@@@@@@@@@@@@@@@@@' % i
        PRID = probj['id']
        # if PRID != '02458610':
        #     continue
        BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(probj)
        state = probj['state']
        if state == 'Closed' or state == 'First Correction Complete' or \
                state == 'Final Response Sent':  # "Correction Plan Ready",
            author = probj['author']
            PRGroupInCharge = probj['groupIncharge']
        else:
            return
        todo_item = getGroupInChargeInfo(PRID, PRGroupInCharge)
        if todo_item:
            JIRAProject = todo_item.JIRAProject
            RCAFilter = todo_item.RCAFilter
            # if state == 'Closed' or state == 'First Correction Complete' or \
            #         state == 'Final Response Sent':  # "Correction Plan Ready",
            if RCAFilter == 'All Customer(A+B+C), All Internal':
                addToMnRcaAll(PRID)
                return
            if RCAFilter == 'Partial Customer(A+B) and Internal, both according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                addToMnRca(PRID, urlfa, probj)
                return
            if RCAFilter == 'No Internal, Partial Customer(A+B) according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                addToMnRca(PRID, urlfa, probj)
                return
            if RCAFilter == 'All Customer(A+B), Partial Internal according to FA rcaEdaDecision Flag':
                faultAnalysisId = probj['faultAnalysisId']
                urlfa = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/1/faultAnalysis/%s' % faultAnalysisId
                addToMnRca(PRID, urlfa, probj)
                # return
            if RCAFilter == 'Customer A + Key Customers B':
                addToMnRcaCustomerABnew(PRID, probj, author, JIRAProject)
                return
            # Below for Only Customer (A+B)
            if author != 'Electra':
                return
            # FNRed = False
            problemType = probj['problemType']  # Documentation
            FNRed = isFNRedNew(probj)
            if FNRed == True and JIRAProject != 'CUDOCIPBL':
                todo_item = MnTodo.query.get(PRID)
                if todo_item:
                    if todo_item.FaRcaEdaDecisionFlag == 'Yes':
                        return
                    else:
                        db.session.delete(todo_item)
                        db.session.commit()
                        return
            # PRID = probj['id']
            PRReportedDate = probj['reportedDate'].split('T')[0]
            PRClosedDate = probj['lastMod'].split('T')[0]
            # print PRID
            try:
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                # else:
                #     print 'No '
            except:
                pass

            PRTitle = probj['title']
            ReportedBy = probj['reportedBy']
            PRGroupInCharge = probj['groupIncharge']
            FaultCoordinator = probj['devFaultCoord']
            PRSeverity = probj['severity']
            if PRSeverity == 'A - Critical':
                rank = 1
            elif PRSeverity == 'B - Major':
                rank = 1
            elif PRSeverity == 'C - Minor':
                rank = 3
            PRRelease = probj['softwareRelease']
            Feature = probj['feature']
            PRProduct = probj['product']
            if 'customerName' in probj.keys():
                CustomerName = probj['customerName']
            else:
                CustomerName = 'Nokia'
            PRAttached = probj['problemReportIds']
            attachedlist = PRAttached
            if not PRAttached:
                PRAttached = ''
            else:
                LabelItems = PRAttached
                b = ''
                for s in LabelItems:
                    b = b + s + ','
                    PRAttached = b.strip(',')
            for item in attachedlist:
                c = APIPr(item)
                Attachedauthor = c['author']
                if Attachedauthor == 'Electra':
                    attachedSeverity = c['severity']
                    if attachedSeverity == 'A - Critical':
                        rank1 = 1
                    elif attachedSeverity == 'B - Major':
                        rank1 = 1
                    elif attachedSeverity == 'C - Minor':
                        rank1 = 3
                    if rank1 == rank:
                        AttachedPRReportedDate = c['reportedDate'].split('T')[0]
                        if AttachedPRReportedDate < PRReportedDate:
                            PRID = c['id']
                            BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                            PRReportedDate = c['reportedDate'].split('T')[0]
                            PRGroupInCharge = c['groupIncharge']
                            # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                            PRClosedDate = c['lastMod'].split('T')[0]
                            dateFlag, realDate = getPRCloseDate(PRID)
                            if dateFlag:
                                PRClosedDate = realDate
                                # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                            PRTitle = c['title']
                            ReportedBy = c['reportedBy']
                            FaultCoordinator = c['devFaultCoord']
                            PRSeverity = c['severity']
                            if PRSeverity == 'A - Critical':
                                rank = 1
                            elif PRSeverity == 'B - Major':
                                rank = 1
                            elif PRSeverity == 'C - Minor':
                                rank = 3
                            PRRelease = c['softwareRelease']
                            Feature = c['feature']
                            PRAttached = c['problemReportIds']
                            if not PRAttached:
                                PRAttached = ''
                            else:
                                LabelItems = PRAttached
                                b = ''
                                for s in LabelItems:
                                    b = b + s + ','
                                    PRAttached = b.strip(',')
                            PRProduct = c['product']
                            if 'customerName' in c.keys():
                                CustomerName = c['customerName']
                            else:
                                CustomerName = 'Nokia'
                    elif rank1 < rank:
                        PRID = c['id']
                        BusinessLine, ProductLine = getBusinessLineAndProductLineWithJson(c)
                        PRReportedDate = c['reportedDate'].split('T')[0]
                        PRGroupInCharge = c['groupIncharge']
                        # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                        PRClosedDate = c['lastMod'].split('T')[0]
                        # print 'before getPRCloseDate  PRID= %s!!!!!!!!!!!!!!!!!!!!!' % PRID
                        dateFlag, realDate = getPRCloseDate(PRID)
                        if dateFlag:
                            PRClosedDate = realDate
                            # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                        PRTitle = c['title']
                        ReportedBy = c['reportedBy']
                        FaultCoordinator = c['devFaultCoord']
                        PRSeverity = c['severity']
                        if PRSeverity == 'A - Critical':
                            rank = 1
                        elif PRSeverity == 'B - Major':
                            rank = 1
                        elif PRSeverity == 'C - Minor':
                            rank = 3
                        PRRelease = c['softwareRelease']
                        Feature = c['feature']
                        PRAttached = c['problemReportIds']
                        if not PRAttached:
                            PRAttached = ''
                        else:
                            LabelItems = PRAttached
                            b = ''
                            for s in LabelItems:
                                b = b + s + ','
                                PRAttached = b.strip(',')
                        PRProduct = c['product']
                        if 'customerName' in c.keys():
                            CustomerName = c['customerName']
                        else:
                            CustomerName = 'Nokia'
                # else:
                #     print 'Internal PR, ignore'
            # BusinessLine, ProductLine = getBusinessLineAndProductLine(PRID)
            groupInChargeInfo = getGroupInChargeInfo(PRID, PRGroupInCharge)
            if groupInChargeInfo:
                JIRAProject = groupInChargeInfo.JIRAProject
                JiraCustomerName = FindCustomerName(CustomerName.lower())
                # ProductLine = FindProductLine(PRProduct)
                BusinessLine = groupInChargeInfo.BusinessLine
                if groupInChargeInfo.ProductLine != 'All':
                    ProductLine = groupInChargeInfo.ProductLine

                if PRAttached:
                    attachedList = PRAttached.split(',')
                    for item in attachedList:
                        todo_item = MnTodo.query.get(item)
                        if todo_item:
                            db.session.delete(todo_item)
                            db.session.commit()
                todomn = MnTodo.query.get(PRID)
                if todomn:
                    todomn.ProductLine = ProductLine
                    todomn.JiraCustomerName = JiraCustomerName
                    todomn.RCAEDACategory = groupInChargeInfo.RCAEDACategory
                    todomn.PRClosedDate = PRClosedDate
                    todomn.PRAttached = PRAttached
                    todomn.JIRAProject = JIRAProject
                    todomn.PRSeverity = PRSeverity
                    todomn.PRGroupInCharge = PRGroupInCharge
                    db.session.commit()
                    todo_item = Todo.query.get(PRID)
                    if todo_item:
                        todo_item.PRSeverity = todomn.PRSeverity
                        todo_item.PRAttached = todomn.PRAttached
                        todo_item.PRGroupInCharge = todomn.PRGroupInCharge
                        todo_item.PRProduct = todomn.PRProduct
                        todo_item.ReportedBy = todomn.ReportedBy
                        todo_item.FaultCoordinator = todomn.FaultCoordinator
                        todo_item.CustomerName = todomn.CustomerName
                        todo_item.Feature = todomn.Feature
                        todo_item.JiraCustomerName = todomn.JiraCustomerName
                        todo_item.JIRAProject = todomn.JIRAProject
                        db.session.commit()
                else:
                    PRRcaEdaAssessor = groupInChargeInfo.AssessorEmail
                    EDA_AssessorEmail = groupInChargeInfo.EDA_AssessorEmail
                    EDACreatingTime = groupInChargeInfo.EDACreatingTime
                    EdaCaseType = groupInChargeInfo.EdaCaseType
                    BusinessUnit = groupInChargeInfo.BusinessUnit
                    # ProductLine = groupInChargeInfo.ProductLine
                    RCAEDACategory = groupInChargeInfo.RCAEDACategory
                    FaRcaEdaDecisionFlag = ''
                    JiraTaskIsCreated = 'No'
                    todo = MnTodo(PRID, PRTitle, PRReportedDate, PRClosedDate, PRRelease, PRSeverity,
                                  PRGroupInCharge, PRAttached, PRRcaEdaAssessor, \
                                  PRProduct, ReportedBy, FaultCoordinator, CustomerName, BusinessUnit,
                                  BusinessLine, ProductLine, Feature, RCAEDACategory, \
                                  JiraCustomerName, JIRAProject, FaRcaEdaDecisionFlag, EDA_AssessorEmail,
                                  EDACreatingTime, EdaCaseType, JiraTaskIsCreated)
                    # todo.jirauser = g.user
                    db.session.add(todo)
                    db.session.commit()
                    todo_item = Todo.query.get(PRID)
                    if todo_item:
                        todo_item.PRSeverity = PRSeverity
                        todo_item.PRAttached = PRAttached
                        todo_item.PRGroupInCharge = PRGroupInCharge
                        todo_item.PRProduct = PRProduct
                        todo_item.ReportedBy = ReportedBy
                        todo_item.FaultCoordinator = FaultCoordinator
                        todo_item.CustomerName = CustomerName
                        todo_item.Feature = Feature
                        todo_item.JiraCustomerName = JiraCustomerName
                        todo_item.JIRAProject = JIRAProject
                    db.session.commit()
            else:
                return
        else:
            return
    except BaseException as ec1:
        errmes = 'error: PRID=' + str(PRID) + ', errorMes=' + str(ec1)
        print(errmes)
        syslog_error('getCustomerPR_Thread', errormes=errmes)

def isFNRed(prid):
    return True

def addToMnRcaAll(PRID):
    # FaRcaEdaDecisionFlag=rcaEdaDecision
    flg = False
    try:
        c = APIPr(PRID)
    except:
        flg = True
        pass
    if flg:
        return
    faultAnalysisId = c['faultAnalysisId']
    try:
        b = APIFA(faultAnalysisId)
    except:
        flg = True
        pass
    if flg:
        return
    if 'rcaEdaDecision' in b.keys():
        rcaEdaDecision = b['rcaEdaDecision']
    else:
        rcaEdaDecision = ''
    FNRed = False
    FNRed = isFNRed(PRID)
    # correctionIds = c['correctionIds']
    # for correctionId in correctionIds:
    #     urlcf = "https://pronto.inside.nsn.com/prontoapi/rest/api/latest/correction/%s" % correctionId
    #     flg = False
    #     try:
    #         r = requests.get(urlcf, verify=False, auth=('krma', 'Nokia123'))
    #     except:
    #         flg = True
    #         pass
    #     if flg:
    #         continue
    #     b = r.json()
    #     targetBuild = b['targetBuild']
    #     state1 = b['state']
    #     if targetBuild:
    #         targetBuild = targetBuild.split('FNR')
    #         # if len(targetBuild) > 1 and state1 == 'Tested':
    #         if len(targetBuild) > 1:
    #             FNRed = True
    #         if len(targetBuild) < 1 and state1 == 'Tested':
    #             FNRed = False
    #             break
    #         targetBuild = b['targetBuild']
    #         targetBuild = targetBuild.split('FNE')
    #         if len(targetBuild) > 1 and state1 == 'Tested':
    #             FNRed = True
    #         if len(targetBuild) < 1 and state1 == 'Tested':
    #             FNRed = False
    #             break
    #         if FNRed == False and state1 == 'Tested':
    #             break
    if FNRed:
        todo_item = MnTodo.query.get(PRID)
        if todo_item:
            db.session.delete(todo_item)
            db.session.commit()
        return
    PRID = c['id']
    PRReportedDate = c['reportedDate'].split('T')[0]
    PRGroupInCharge = c['groupIncharge']
    PRClosedDate = c['lastMod'].split('T')[0]
    dateFlag, realDate = getPRCloseDate(PRID)
    if dateFlag:
        PRClosedDate = realDate
        # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
    PRTitle = c['title']
    ReportedBy = c['reportedBy']
    FaultCoordinator = c['devFaultCoord']
    PRSeverity = c['severity']
    if PRSeverity == 'A - Critical':
        rank = 1
    elif PRSeverity == 'B - Major':
        rank = 2
    elif PRSeverity == 'C - Minor':
        rank = 3
    PRRelease = c['softwareRelease']
    Feature = c['feature']
    PRProduct = c['product']
    if 'customerName' in c.keys():
        CustomerName = c['customerName']
    else:
        CustomerName = 'Nokia'
    PRAttached = c['problemReportIds']
    attachedlist = PRAttached
    if not PRAttached:
        PRAttached = ''
    else:
        LabelItems = PRAttached
        b = ''
        for s in LabelItems:
            b = b + s + ','
            PRAttached = b.strip(',')
    for item in attachedlist:
        c = APIPr(item)
        # if 'reportedBy' in c.keys():
        #     AttachedreportedBy = c['reportedBy']
        # else:
        #     AttachedreportedBy = ''
        # if AttachedreportedBy == 'Customer':
        attachedSeverity = c['severity']
        if attachedSeverity == 'A - Critical':
            rank1 = 1
        elif attachedSeverity == 'B - Major':
            rank1 = 2
        elif attachedSeverity == 'C - Minor':
            rank1 = 3
        if rank1 == rank:
            AttachedPRReportedDate = c['reportedDate'].split('T')[0]
            if AttachedPRReportedDate < PRReportedDate:
                PRID = c['id']
                PRReportedDate = c['reportedDate'].split('T')[0]
                PRGroupInCharge = c['groupIncharge']
                # todo_item = InChargeGroup.query.get(PRGroupInCharge)
                PRClosedDate = c['lastMod'].split('T')[0]
                dateFlag, realDate = getPRCloseDate(PRID)
                if dateFlag:
                    PRClosedDate = realDate
                    # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
                PRTitle = c['title']
                ReportedBy = c['reportedBy']
                FaultCoordinator = c['devFaultCoord']
                PRSeverity = c['severity']
                if PRSeverity == 'A - Critical':
                    rank = 1
                elif PRSeverity == 'B - Major':
                    rank = 2
                elif PRSeverity == 'C - Minor':
                    rank = 3
                PRRelease = c['softwareRelease']
                Feature = c['feature']
                PRAttached = c['problemReportIds']
                if not PRAttached:
                    PRAttached = ''
                else:
                    LabelItems = PRAttached
                    b = ''
                    for s in LabelItems:
                        b = b + s + ','
                        PRAttached = b.strip(',')
                PRProduct = c['product']
                if 'customerName' in c.keys():
                    CustomerName = c['customerName']
                else:
                    CustomerName = 'Nokia'
        elif rank1 < rank:
            PRID = c['id']
            PRReportedDate = c['reportedDate'].split('T')[0]
            PRGroupInCharge = c['groupIncharge']
            # todo_item = InChargeGroup.query.get(PRGroupInCharge)
            PRClosedDate = c['lastMod'].split('T')[0]
            dateFlag, realDate = getPRCloseDate(PRID)
            if dateFlag:
                PRClosedDate = realDate
                # print 'real close date find!!!!!!!!!!!!!!!!!!!!!'
            PRTitle = c['title']
            ReportedBy = c['reportedBy']
            FaultCoordinator = c['devFaultCoord']
            PRSeverity = c['severity']
            if PRSeverity == 'A - Critical':
                rank = 1
            elif PRSeverity == 'B - Major':
                rank = 2
            elif PRSeverity == 'C - Minor':
                rank = 3
            PRRelease = c['softwareRelease']
            Feature = c['feature']
            PRAttached = c['problemReportIds']
            if not PRAttached:
                PRAttached = ''
            else:
                LabelItems = PRAttached
                b = ''
                for s in LabelItems:
                    b = b + s + ','
                    PRAttached = b.strip(',')
            PRProduct = c['product']
            if 'customerName' in c.keys():
                CustomerName = c['customerName']
            else:
                CustomerName = 'Nokia'
    # PRProduct = c['product']
    # CustomerName = 'Nokia'
    # groupInChargeInfo = InChargeGroup.query.get(PRGroupInCharge)
    BusinessLine, ProductLine = getBusinessLineAndProductLine(PRID)
    # # groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
    # #                                                  BusinessLine=BusinessLine, ProductLine=ProductLine).first()
    # groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
    #                                                  BusinessLine=BusinessLine,
    #                                                  ProductLine='All').first()
    # if not groupInChargeInfo:
    #     groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
    #                                                      BusinessLine=BusinessLine,
    #                                                      ProductLine=ProductLine).first()
    # if not groupInChargeInfo:
    #     groupInChargeInfo = InChargeGroup.query.filter_by(InChargeGroupName=PRGroupInCharge, \
    #                                                      AddedBy='sufang.huang@nokia-sbell.com').first()
    groupInChargeInfo = getGroupInChargeInfo(PRID,PRGroupInCharge)
    JIRAProject = groupInChargeInfo.JIRAProject
    RCAFilter = groupInChargeInfo.RCAFilter
    JiraCustomerName = CustomerName
    # ProductLine = FindProductLine(PRProduct)
    if PRAttached:
        attachedList = PRAttached.split(',')
        for item in attachedList:
            todo_item = MnTodo.query.get(item)
            if todo_item:
                db.session.delete(todo_item)
                db.session.commit()
    # 'All Customer, No internal'
    # 'All Customer, Partial Internal according to FA rcaEdaDecision Flag'
    # 'Partial Customer and Internal, both according to FA rcaEdaDecision Flag'
    # 'No Internal, Partial Customer according to FA rcaEdaDecision Flag'
    if CustomerName != 'Nokia':
        JiraCustomerName = FindCustomerName(CustomerName.lower())
    # if RCAFilter == 'No Internal, Partial Customer according to FA rcaEdaDecision Flag':
    BusinessLine = groupInChargeInfo.BusinessLine
    if groupInChargeInfo.ProductLine != 'All':
        ProductLine = groupInChargeInfo.ProductLine
    # if groupInChargeInfo.BusinessLine == 'LTE' and PRRelease:
    #     if (len(PRRelease.split('TL')) != 1):
    #         ProductLine = 'TD LTE'
    #     elif (len(PRRelease.split('FL')) != 1):
    #         ProductLine = 'FDD LTE'
    #     elif (len(PRRelease.split('CBTS')) != 1):
    #         ProductLine = 'Cloud BTS'
    #     elif (len(PRRelease.split('SBTS')) != 1):
    #         ProductLine = 'SRAN/SBTS'
    #         BusinessLine = 'SRAN'
    # if groupInChargeInfo.BusinessLine == 'LTE' and PRProduct=='Flexi WCDMA Base Station':
    #     ProductLine = 'WBTS'
    #     BusinessLine = 'SRAN'
    todomn= MnTodo.query.get(PRID)
    if todomn:
        todomn.ProductLine = ProductLine
        todomn.JiraCustomerName = JiraCustomerName
        # todomn.RCAEDACategory = groupInChargeInfo.ProductLine
        todomn.RCAEDACategory = groupInChargeInfo.RCAEDACategory
        todomn.PRClosedDate = PRClosedDate
        todomn.PRAttached = PRAttached
        todomn.JIRAProject = JIRAProject
        todomn.FaRcaEdaDecisionFlag = rcaEdaDecision
        todomn.PRSeverity = PRSeverity
        db.session.commit()
    else:
        PRRcaEdaAssessor = groupInChargeInfo.AssessorEmail
        EDA_AssessorEmail = groupInChargeInfo.EDA_AssessorEmail
        EDACreatingTime = groupInChargeInfo.EDACreatingTime
        EdaCaseType = groupInChargeInfo.EdaCaseType
        BusinessUnit = groupInChargeInfo.BusinessUnit
        # BusinessLine = groupInChargeInfo.BusinessLine
        # ProductLine = groupInChargeInfo.ProductLine
        RCAEDACategory = groupInChargeInfo.RCAEDACategory
        FaRcaEdaDecisionFlag = rcaEdaDecision
        JiraTaskIsCreated = 'No'
        todo = MnTodo(PRID,PRTitle,PRReportedDate,PRClosedDate,PRRelease,PRSeverity,PRGroupInCharge,PRAttached,PRRcaEdaAssessor,\
             PRProduct,ReportedBy,FaultCoordinator,CustomerName,BusinessUnit,BusinessLine,ProductLine,Feature,RCAEDACategory,\
                      JiraCustomerName,JIRAProject,FaRcaEdaDecisionFlag,EDA_AssessorEmail,EDACreatingTime,EdaCaseType,JiraTaskIsCreated)
        db.session.add(todo)
        db.session.commit()

def jsonformat():
    json1 = '{"username": "mengxiang.li@nokia-sbell.com", "yxbz": "Y", "rcaformescapedefect": {"why_not_system_test": [{"ap_id": "ap3878730215", "completion_target_date": "", "uuid": 1438680374, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "7. Why didnt System Test                                catch                                this defect? ", "why_id": "why4633666088", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap6547839150", "completion_target_date": "", "uuid": 1390577763, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why1211200144", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_component_test": [{"ap_id": "ap2634003751", "completion_target_date": "", "uuid": 4374365699, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:333Answer:rrr", "why3": "", "why1": "Why didnt unit or component test catch this defect?", "why_id": "why2428193984", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap5145978402", "completion_target_date": "", "uuid": 6938275300, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why2343756332", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap2216372173", "completion_target_date": "", "uuid": 465398829, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:cccAnswer:ha", "why3": "", "why1": "", "why_id": "why2493598465", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_st_auto": [{"ap_id": "ap1392363480", "completion_target_date": "", "uuid": 5274551512, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:xAnswer:x", "why3": "", "why1": "Why didnt ST automated test                                cases                                catch this defect?", "why_id": "why1666385683", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap5600427014", "completion_target_date": "", "uuid": 8146981541, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why1263705508", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_requirements_review": [{"ap_id": "ap9739493088", "completion_target_date": "", "uuid": 7198345139, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt Requirements reviews catch this defect? ", "why_id": "why3604524808", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap3492706343", "completion_target_date": "", "uuid": 7930570700, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why2543062941", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap1727089601", "completion_target_date": "", "uuid": 5853649259, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why8953667536", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_et_auto": [{"ap_id": "ap5918505969", "completion_target_date": "", "uuid": 6208767226, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:y2xAnswer:y22x", "why3": "", "why1": "Why didnt ET automated test                                cases                                catch this defect?", "why_id": "why7086712883", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap609964132", "completion_target_date": "", "uuid": 1714815226, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why3476105220", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_inspections": [{"ap_id": "ap3714880029", "completion_target_date": "", "uuid": 6004754542, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt code inspections catch this defect?", "why_id": "why6719252535", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap4476929055", "completion_target_date": "", "uuid": 4460761551, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:_x000D_asdfx_x000D__x000D_Answer:_x000D_sadfx", "why3": "", "why1": "", "why_id": "why1495762353", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap5102286189", "completion_target_date": "", "uuid": 9579167508, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why6630532964", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_design_review": [{"ap_id": "ap4401005496", "completion_target_date": "", "uuid": 1158154711, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "Question:2jzxzzAnswer:2jyyzxz", "why3": "", "why1": "Why didnt Design reviews catch this defect? ", "why_id": "why5133293425", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap6469884461", "completion_target_date": "", "uuid": 6492657225, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why2372411049", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap5530641601", "completion_target_date": "", "uuid": 4738279296, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why2363599950", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_entity_test": [{"ap_id": "ap9009760602", "completion_target_date": "", "uuid": 1228051954, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt Entity Test test catch this defect? ", "why_id": "why5189143580", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap3143602873", "completion_target_date": "", "uuid": 9406705965, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why4732104011", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}], "why_not_analysis_tools": [{"ap_id": "ap6514294239", "completion_target_date": "", "uuid": 9336424099, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt code analysis tools such as Klocwork, Purify, etc catch this defect? ", "why_id": "why9745682740", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap6099663925", "completion_target_date": "", "uuid": 2568815758, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why6246740590", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}, {"ap_id": "ap9157843989", "completion_target_date": "", "uuid": 9196999397, "escape_cause_subcategory": "", "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why2798843549", "escape_cause": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "actionproposal": "", "escape_cause_category": ""}]}, "rcaformrootcause": {"why_was_the_fault_introduced": [{"why_id": "why2190952924", "completion_target_date": "2022-09-30", "uuid": 8711848714, "grading": "B", "rootcause": "import * as React from react;import Checkbox from @mui/material/Checkbox;import TextField from @mui/material/TextField;import Autocomplete from @mui/material/Autocomplete;import CheckBoxOutlineBlankIcon from @mui/icons-material/CheckBoxOutlineBlank;import CheckBoxIcon from @mui/icons-material/CheckBox;", "why4": "Question:3LSLSLSLSLSLSLLsAnswer:4", "why5": "s;s;s;s;s;s;8930348u9730940932409", "why2": "Question:asdfq;12Answer:sadfxl2", "why3": "Question:px123xNPS Tribe:why the counter was not know?12yzwrpAnswer:prx1No Req or hint leading us to think of a potential issue. After this CAS, A&S modified requirements in PR spec PR638766 (beamforming activated) and PR spec PR638223 (beamforming deactivated)1ywp", "why1": "Question:zzzzzzxxxxxx3Answer:zzzzzzzzz", "root_cause_category": {"text": "Architecture ", "dm": "02"}, "rca_action_type": {"text": "User Documentation (including Release Note) Improvement", "dm": "06"}, "ap_jiratask_id": "", "assigned_to": "enqing.lei@nokia-sbell.com", "ap_id": "ap7283957371", "actionproposal": "12345", "root_cause_subcategory": {"text": "Memory consumption error", "dm": "02"}}, {"why_id": "why1559683841", "completion_target_date": "", "uuid": 8959583869, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "root_cause_category": "", "rca_action_type": "", "ap_jiratask_id": "", "assigned_to": "", "ap_id": "ap6154707100", "actionproposal": "", "root_cause_subcategory": ""}, {"why_id": "why1382657208", "completion_target_date": "", "uuid": 7295336129, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "Question:2Answer:3", "why3": "Question:asdfAnswer:asdf", "why1": "Question:asdfAnswer:dsafzzz", "root_cause_category": "", "rca_action_type": "", "ap_jiratask_id": "", "assigned_to": "", "ap_id": "ap5686112505", "actionproposal": "", "root_cause_subcategory": ""}], "why_root_cause_was_not_found": [{"why_id": "why8826971642", "completion_target_date": "", "uuid": 5725248814, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "Question:sdAnswer:xxxccxccx", "why3": "Question:1Answer:zz", "why1": "Question:adsfAnswer:z12", "root_cause_category": "", "rca_action_type": "", "ap_jiratask_id": "", "assigned_to": "", "ap_id": "ap6233611630", "actionproposal": "", "root_cause_subcategory": ""}]}, "rcaformbase": {"product": "Flexi LTE Base Station", "issue_description": "[1. Impact on an operator:]Detailed description of what will be the most likely phenomena seen in the network once issue experienced (this is an extended description of the impact selected in the Customer impact Tester Analysis).: no user thp UL/DL[2. Impact on an end user:]Detailed description of problem on the End user, :No throughput, [3. Impacted HW System Module:]FSMF+FBBA*2[4. Impacted HW Radio Module:]FZNI", "assessors": "enqing.lei@nokia-sbell.com", "code_deficience": "*** DEFAULT TEMPLATE for FA IDENTIFICATION for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019)  PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Summary of the problem:] [2. Technical description of the fault:] Code Deficiency: What was wrong in the source code? Incorrect u64 to u32 data type transfer when calculate the AMBR limit. [3. Dependency on configuration:] Description if problem is based on certain configuration/feature(s)  What minimal environment (e.g. HW/SW) configuration should be used to reproduce the issue NN [4. Faulty component and version:] SW component and version where problem occurred first time. If problem occurred in different branches the all first broken versions for each branch shall be listed. TDDMACPS: SBTS19B_ENB_0000_001514_000000 *** END OF DEFAULT TEMPLATE ***", "pr_grade": {"text": "C \u2013 Clarifications required", "dm": "261616"}, "case_number": "01309318", "additional_facts": "dfg123456789", "triggering_scenario_category": {"text": "", "dm": ""}, "injection_type": {"text": "", "dm": ""}, "inheritance_recommendation": "xxx.", "how_many_times": "x12345", "quality_reviewer": "12xz1", "abstract_headline": "No user plane traffic in LTE TDD (old case ref. 00679635)", "triggering_scenario": "Problem 01314631: Transfer to PRONTO -- attention @Ma, Yanfang F. (NSB - CN/Hangzhou) yanfang.f.ma@nokia-sbell.comRefer to 00679635.[1. Detail Test Steps:] PRB01314631_CASE01309318: No user plane traffic in LTE TDD.", "correction_description": "*** DEFAULT TEMPLATE for FA Resolution for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019)  PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Workaround:] [2. Description of the correction:] What changes were done in code/HW architecture to fix the issue and how problem will be solved? Increase the max rate, and correct the u64 to u32 data transfer. [3. Test requirements:] How to test the correction and catch the problem in future in SCT/UT/MT level? UT*** END OF DEFAULT TEMPLATE ***", "injection_time": ""}, "pronto_id": "01309318"}'
    json2 = '{"username": "mengxiang.li@nokia-sbell.com", "yxbz": "Y", "rcaformescapedefect": {"why_not_system_test": [{"why_id": "why4633666088", "completion_target_date": "", "uuid": 7746735916, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "7. Why didnt System Test                                catch                                this defect? ", "ap_id": "ap3878730215", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why1211200144", "completion_target_date": "", "uuid": 3894624397, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap6547839150", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_component_test": [{"why_id": "why2428193984", "completion_target_date": "", "uuid": 7749292438, "grading": "", "why4": "", "why5": "", "why2": "Question:333Answer:rrr", "why3": "", "why1": "Why didnt unit or component test catch this defect?", "ap_id": "ap2634003751", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2343756332", "completion_target_date": "", "uuid": 5480422646, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap5145978402", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2493598465", "completion_target_date": "", "uuid": 9886151212, "grading": "", "why4": "", "why5": "", "why2": "Question:cccAnswer:ha", "why3": "", "why1": "", "ap_id": "ap2216372173", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_st_auto": [{"why_id": "why1666385683", "completion_target_date": "", "uuid": 1746125861, "grading": "", "why4": "", "why5": "", "why2": "Question:xAnswer:x", "why3": "", "why1": "Why didnt ST automated test                                cases                                catch this defect?", "ap_id": "ap1392363480", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why1263705508", "completion_target_date": "", "uuid": 275788643, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap5600427014", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_requirements_review": [{"why_id": "why3604524808", "completion_target_date": "", "uuid": 4721488644, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt Requirements reviews catch this defect? ", "ap_id": "ap9739493088", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2543062941", "completion_target_date": "", "uuid": 9931944892, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap3492706343", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why8953667536", "completion_target_date": "", "uuid": 2279307529, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap1727089601", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_design_review": [{"why_id": "why5133293425", "completion_target_date": "", "uuid": 9046676916, "grading": "", "why4": "", "why5": "", "why2": "Question:2jzxzzAnswer:2jyyzxz", "why3": "", "why1": "Why didnt Design reviews catch this defect? ", "ap_id": "ap4401005496", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2372411049", "completion_target_date": "", "uuid": 5942999593, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap6469884461", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2363599950", "completion_target_date": "", "uuid": 2490402760, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap5530641601", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_inspections": [{"why_id": "why6719252535", "completion_target_date": "", "uuid": 5629898145, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt code inspections catch this defect?", "ap_id": "ap3714880029", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why1495762353", "completion_target_date": "", "uuid": 5707912795, "grading": "", "why4": "", "why5": "", "why2": "Question:_x000D_asdfx_x000D__x000D_Answer:_x000D_sadfx", "why3": "", "why1": "", "ap_id": "ap4476929055", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why6630532964", "completion_target_date": "", "uuid": 9527838163, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap5102286189", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_et_auto": [{"why_id": "why7086712883", "completion_target_date": "", "uuid": 7880769129, "grading": "", "why4": "", "why5": "", "why2": "Question:y2xAnswer:y22x", "why3": "", "why1": "Why didnt ET automated test                                cases                                catch this defect?", "ap_id": "ap5918505969", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why3476105220", "completion_target_date": "", "uuid": 983567356, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap609964132", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_entity_test": [{"why_id": "why5189143580", "completion_target_date": "", "uuid": 2949602878, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt Entity Test test catch this defect? ", "ap_id": "ap9009760602", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why4732104011", "completion_target_date": "", "uuid": 7885535355, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap3143602873", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}], "why_not_analysis_tools": [{"why_id": "why9745682740", "completion_target_date": "", "uuid": 4381049684, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "Why didnt code analysis tools such as Klocwork, Purify, etc catch this defect? ", "ap_id": "ap6514294239", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why6246740590", "completion_target_date": "", "uuid": 232787477, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap6099663925", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}, {"why_id": "why2798843549", "completion_target_date": "", "uuid": 2867797896, "grading": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "ap_id": "ap9157843989", "escape_cause": "", "actionproposal": "", "ap_jiratask_id": "", "assigned_to": "", "eda_action_type": "", "where_could_defect_have_been_detected": "", "escape_cause_subcategory": "", "escape_cause_category": ""}]}, "rcaformrootcause": {"why_was_the_fault_introduced": [{"ap_jiratask_id": "xxx", "root_cause_category": {"text": "Architecture ", "dm": "02"}, "completion_target_date": "2022-09-30", "uuid": 4600236915, "grading": "B", "rootcause": "import * as React from react;import Checkbox from @mui/material/Checkbox;import TextField from @mui/material/TextField;import Autocomplete from @mui/material/Autocomplete;import CheckBoxOutlineBlankIcon from @mui/icons-material/CheckBoxOutlineBlank;import CheckBoxIcon from @mui/icons-material/CheckBox;", "why4": "Question:3LSLSLSLSLSLSLLsAnswer:4", "why5": "s;s;s;s;s;s;8930348u9730940932409", "why2": "Question:asdfq;12Answer:sadfxl2", "why3": "Question:px123xNPS Tribe:why the counter was not know?12yzwrpAnswer:prx1No Req or hint leading us to think of a potential issue. After this CAS, A&S modified requirements in PR spec PR638766 (beamforming activated) and PR spec PR638223 (beamforming deactivated)1ywp", "why1": "Question:zzzzzzxxxxxx3Answer:zzzzzzzzz", "why_id": "why2190952924", "root_cause_subcategory": {"text": "Memory consumption error", "dm": "02"}, "assigned_to": "enqing.lei@nokia-sbell.com", "ap_id": "ap7283957371", "actionproposal": "12345", "rca_action_type": {"text": "User Documentation (including Release Note) Improvement", "dm": "06"}}, {"ap_jiratask_id": "", "root_cause_category": "", "completion_target_date": "", "uuid": 9122585079, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "", "why3": "", "why1": "", "why_id": "why1559683841", "root_cause_subcategory": "", "assigned_to": "", "ap_id": "ap6154707100", "actionproposal": "", "rca_action_type": ""}, {"ap_jiratask_id": "", "root_cause_category": "", "completion_target_date": "", "uuid": 4977717265, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "Question:2Answer:3", "why3": "Question:asdfAnswer:asdf", "why1": "Question:asdfAnswer:dsafzzz", "why_id": "why1382657208", "root_cause_subcategory": "", "assigned_to": "", "ap_id": "ap5686112505", "actionproposal": "", "rca_action_type": ""}], "why_root_cause_was_not_found": [{"ap_jiratask_id": "", "root_cause_category": "", "completion_target_date": "", "uuid": 7282408683, "grading": "", "rootcause": "", "why4": "", "why5": "", "why2": "Question:sdAnswer:xxxccxccx", "why3": "Question:1Answer:zz", "why1": "Question:adsfAnswer:z12", "why_id": "why8826971642", "root_cause_subcategory": "", "assigned_to": "", "ap_id": "ap6233611630", "actionproposal": "", "rca_action_type": ""}]}, "rcaformbase": {"product": "Flexi LTE Base Station", "issue_description": "[1. Impact on an operator:]Detailed description of what will be the most likely phenomena seen in the network once issue experienced (this is an extended description of the impact selected in the Customer impact Tester Analysis).: no user thp UL/DL[2. Impact on an end user:]Detailed description of problem on the End user, :No throughput, [3. Impacted HW System Module:]FSMF+FBBA*2[4. Impacted HW Radio Module:]FZNI", "assessors": "enqing.lei@nokia-sbell.com", "code_deficience": "*** DEFAULT TEMPLATE for FA IDENTIFICATION for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019)  PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Summary of the problem:] [2. Technical description of the fault:] Code Deficiency: What was wrong in the source code? Incorrect u64 to u32 data type transfer when calculate the AMBR limit. [3. Dependency on configuration:] Description if problem is based on certain configuration/feature(s)  What minimal environment (e.g. HW/SW) configuration should be used to reproduce the issue NN [4. Faulty component and version:] SW component and version where problem occurred first time. If problem occurred in different branches the all first broken versions for each branch shall be listed. TDDMACPS: SBTS19B_ENB_0000_001514_000000 *** END OF DEFAULT TEMPLATE ***", "pr_grade": {"text": "C \u2013 Clarifications required", "dm": "261616"}, "quality_reviewer": "12xz1", "additional_facts": "dfg123456789", "triggering_scenario_category": {"text": "", "dm": ""}, "abstract_headline": "No user plane traffic in LTE TDD (old case ref. 00679635)", "inheritance_recommendation": "xxx.", "how_many_times": "x12345", "case_number": "01309318", "injection_type": {"text": "", "dm": ""}, "triggering_scenario": "Problem 01314631: Transfer to PRONTO -- attention @Ma, Yanfang F. (NSB - CN/Hangzhou) yanfang.f.ma@nokia-sbell.comRefer to 00679635.[1. Detail Test Steps:] PRB01314631_CASE01309318: No user plane traffic in LTE TDD.", "correction_description": "*** DEFAULT TEMPLATE for FA Resolution for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019)  PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Workaround:] [2. Description of the correction:] What changes were done in code/HW architecture to fix the issue and how problem will be solved? Increase the max rate, and correct the u64 to u32 data transfer. [3. Test requirements:] How to test the correction and catch the problem in future in SCT/UT/MT level? UT*** END OF DEFAULT TEMPLATE ***", "injection_time": ""}, "pronto_id": "01309318"}'
    json3 = '{"username":"mengxiang.li@nokia-sbell.com","yxbz":"Y","rcaformescapedefect":{"why_not_system_test":[{"ap_id":"ap3878730215","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"7. Why didnt System Test catch this defect? ","why_id":"why4633666088","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap6547839150","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why1211200144","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_component_test":[{"ap_id":"ap2634003751","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:333Answer:rrr","why3":"","why1":"Why didnt unit or component test catch this defect?","why_id":"why2428193984","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap5145978402","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why2343756332","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap2216372173","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:cccAnswer:ha","why3":"","why1":"","why_id":"why2493598465","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_st_auto":[{"ap_id":"ap1392363480","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:xAnswer:x","why3":"","why1":"Why didnt ST automated test cases catch this defect?","why_id":"why1666385683","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap5600427014","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why1263705508","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_requirements_review":[{"ap_id":"ap9739493088","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"Why didnt Requirements reviews catch this defect? ","why_id":"why3604524808","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap3492706343","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why2543062941","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap1727089601","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why8953667536","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_et_auto":[{"ap_id":"ap5918505969","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:y2xAnswer:y22x","why3":"","why1":"Why didnt ET automated test cases catch this defect?","why_id":"why7086712883","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap609964132","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why3476105220","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_inspections":[{"ap_id":"ap3714880029","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"Why didnt code inspections catch this defect?","why_id":"why6719252535","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap4476929055","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:_x000D_asdfx_x000D__x000D_Answer:_x000D_sadfx","why3":"","why1":"","why_id":"why1495762353","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap5102286189","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why6630532964","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_design_review":[{"ap_id":"ap4401005496","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"Question:2jzxzzAnswer:2jyyzxz","why3":"","why1":"Why didnt Design reviews catch this defect? ","why_id":"why5133293425","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap6469884461","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why2372411049","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap5530641601","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why2363599950","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_entity_test":[{"ap_id":"ap9009760602","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"Why didnt Entity Test test catch this defect? ","why_id":"why5189143580","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap3143602873","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why4732104011","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}],"why_not_analysis_tools":[{"ap_id":"ap6514294239","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"Why didnt code analysis tools such as Klocwork, Purify, etc catch this defect? ","why_id":"why9745682740","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap6099663925","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why6246740590","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""},{"ap_id":"ap9157843989","completion_target_date":"","escape_cause_category":"","grading":"","why4":"","why5":"","why2":"","why3":"","why1":"","why_id":"why2798843549","escape_cause":"","escape_cause_subcategory":"","assigned_to":"","eda_action_type":"","where_could_defect_have_been_detected":"","actionproposal":"","ap_jiratask_id":""}]},"rcaformrootcause":{"why_was_the_fault_introduced":[{"why_id":"why2190952924","completion_target_date":"2022-09-30","rca_action_type":{"text":"User Documentation (including Release Note) Improvement","dm":"06"},"grading":"C","rootcause":"import * as React from react;import Checkbox from @mui/material/Checkbox;import TextField from @mui/material/TextField;import Autocomplete from @mui/material/Autocomplete;import CheckBoxOutlineBlankIcon from @mui/icons-material/CheckBoxOutlineBlank;import CheckBoxIcon from @mui/icons-material/CheckBox;","why4":"Question:3LSLSLSLSLSLSLLsAnswer:4","why5":"s;s;s;s;s;s;8930348u9730940932409","why2":"Question:asdfq;12Answer:sadfxl2","why3":"Question:px123xNPS Tribe:why the counter was not know?12yzwrpAnswer:prx1No Req or hint leading us to think of a potential issue. After this CAS, A&S modified requirements in PR spec PR638766 (beamforming activated) and PR spec PR638223 (beamforming deactivated)1ywp","why1":"Question:zzzzzzxxxxxx3Answer:zzzzzzzzz","root_cause_category":{"text":"Architecture ","dm":"02"},"root_cause_subcategory":{"text":"Memory consumption error","dm":"02"},"assigned_to":"enqing.lei@nokia-sbell.com","ap_id":"ap7283957371","actionproposal":"12345","ap_jiratask_id":"xxx"},{"why_id":"why1559683841","completion_target_date":"","rca_action_type":"","grading":"","rootcause":"","why4":"","why5":"","why2":"","why3":"","why1":"","root_cause_category":"","root_cause_subcategory":"","assigned_to":"","ap_id":"ap6154707100","actionproposal":"","ap_jiratask_id":""},{"why_id":"why1382657208","completion_target_date":"","rca_action_type":"","grading":"","rootcause":"","why4":"","why5":"","why2":"Question:2Answer:3","why3":"Question:asdfAnswer:asdf","why1":"Question:asdfAnswer:dsafzzz","root_cause_category":"","root_cause_subcategory":"","assigned_to":"","ap_id":"ap5686112505","actionproposal":"","ap_jiratask_id":""}],"why_root_cause_was_not_found":[{"why_id":"why8826971642","completion_target_date":"","rca_action_type":"","grading":"","rootcause":"","why4":"","why5":"","why2":"Question:sdAnswer:xxxccxccx","why3":"Question:1Answer:zz","why1":"Question:adsfAnswer:z12","root_cause_category":"","root_cause_subcategory":"","assigned_to":"","ap_id":"ap6233611630","actionproposal":"","ap_jiratask_id":""}]},"rcaformbase":{"product":"Flexi LTE Base Station","issue_description":"[1. Impact on an operator:]Detailed description of what will be the most likely phenomena seen in the network once issue experienced (this is an extended description of the impact selected in the Customer impact Tester Analysis).: no user thp UL/DL[2. Impact on an end user:]Detailed description of problem on the End user, :No throughput, [3. Impacted HW System Module:]FSMF+FBBA*2[4. Impacted HW Radio Module:]FZNI","assessors":"enqing.lei@nokia-sbell.com","code_deficience":"*** DEFAULT TEMPLATE for FA IDENTIFICATION for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019) PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Summary of the problem:] [2. Technical description of the fault:] Code Deficiency: What was wrong in the source code? Incorrect u64 to u32 data type transfer when calculate the AMBR limit. [3. Dependency on configuration:] Description if problem is based on certain configuration/feature(s) What minimal environment (e.g. HW/SW) configuration should be used to reproduce the issue NN [4. Faulty component and version:] SW component and version where problem occurred first time. If problem occurred in different branches the all first broken versions for each branch shall be listed. TDDMACPS: SBTS19B_ENB_0000_001514_000000 *** END OF DEFAULT TEMPLATE ***","pr_grade":{"text":"C \u2013 Clarifications required","dm":"261616"},"case_number":"01309318","additional_facts":"dfg123456789","triggering_scenario_category":{"text":"","dm":""},"injection_type":{"text":"","dm":""},"inheritance_recommendation":"xxx.","how_many_times":"x12345","quality_reviewer":"12xz1","abstract_headline":"No user plane traffic in LTE TDD (old case ref. 00679635)","triggering_scenario":"Problem 01314631: Transfer to PRONTO -- attention @Ma, Yanfang F. (NSB - CN/Hangzhou) yanfang.f.ma@nokia-sbell.comRefer to 00679635.[1. Detail Test Steps:] PRB01314631_CASE01309318: No user plane traffic in LTE TDD.","correction_description":"*** DEFAULT TEMPLATE for FA Resolution for 2G-3G-4G-5G-SRAN-FDD-TDD-DCM-Micro-Controller common template v1.0.0 (03.06.2019) PLEASE FILL IN AND DO NOT CHANGE / REMOVE ANY SECTION OF THIS TEMPLATE ***[1. Workaround:] [2. Description of the correction:] What changes were done in code/HW architecture to fix the issue and how problem will be solved? Increase the max rate, and correct the u64 to u32 data transfer. [3. Test requirements:] How to test the correction and catch the problem in future in SCT/UT/MT level? UT*** END OF DEFAULT TEMPLATE ***","injection_time":""},"pronto_id":"01309318"}'
    jsonstr1 = json.loads(json1, object_pairs_hook=OrderedDict)
    jsonstr2 = json.loads(json2, object_pairs_hook=OrderedDict)
    jsonstr3 = json.loads(json3, object_pairs_hook=OrderedDict)
    print(json.dumps(jsonstr1, sort_keys=True))
    print(json.dumps(jsonstr2, sort_keys=True))
    print(json.dumps(jsonstr3, sort_keys=True))

def toolsGetDm(dmstr, key):
    if not dmstr:
        return ''
    json_data = ast.literal_eval(dmstr)
    return json_data[key] if key in json_data else ''

def toolsQAParser(qatext):
    if not qatext:
        return ['','']
    if 'Question:' in qatext:
        q = qatext[qatext.index('Question:') + len('Question:'): qatext.index('Answer:')]
        a = qatext[qatext.index('Answer:') + len('Answer:'): -1]
        return [q.strip(),a.strip()]
    return [qatext,'']

def toolsQAGenerate(q,a):
    if not q and not a:
        return ''
    return 'Question:\r\n' + q + '\r\nAnswer:\r\n' + a

TriggeringCategory = {  'Installation & Startup': u'219294',
                        'SW Upgrade':u'219295',
                        'SW Fallback':u'219296',
                        'Process of Configuration / Reconfiguration':u'219297',
                        'Not Supported Configuration':u'219298',
                        'OAM Operations':u'219299',
                        'Feature Interaction & interoperability':u'219300',
                        'OAM Robustness (high Load / stressful scenarios/long duration)':u'219301',
                        'Telecom Robustness (high Load / mobility /stressful scenarios/long duration)':u'219302',
                        'Debug (Counters/Alarms/Trace/Resource monitoring)':u'219303',
                        'Reset & Recovery':u'219304',
                        'HW Failure':u'219305',
                        'Required customer/vendor specific equipment':u'219306',
                        'Unknown triggering scenario':u'219307'
                        }
JIRA_STATUS = {'Closed': 2, 'Reopened' : 3, 'In Progress' : 4, 'Resolved' : 5}
def testTodo():
    RcaSubtaskJiraId=''
    jira = getjira()
    RcaAanlysistaskIssue = jira.issue('MNRCA-50048')
    todo_item = Todo.query.get('PR642255')
    if len(todo_item.TriggerScenarioCategory) != 0:  # TODO check if null
        TriggeringType = TriggeringCategory[todo_item.TriggerScenarioCategory]
        #Rca5WhyParentIssue.update(customfield_39290={'id': TriggeringType})
    shouldhavebeenfound=[]
    shouldhavebeenfound1=[]
    shouldhavebeenfound
    shouldhavebeenfound1
    if shouldhavebeenfound:  # Should have been found update
        dict1 = {'customfield_37199': shouldhavebeenfound}
        #Rca5WhyParentIssue.update(dict1)
        todo_item.ShouldHaveBeenDetected = shouldhavebeenfound1[0]
        shouldhavebeenfound = []
        shouldhavebeenfound1 = []
    status = str(RcaAanlysistaskIssue.fields.status)
    issuetype = str(RcaAanlysistaskIssue.fields.issuetype)
    print status + 'Status of JiraIssue!!!'
    if status in ['Open', 'Reopened', 'In Progress']:
        if issuetype == 'Analysis subtask':
            jira.transition_issue(RcaAanlysistaskIssue, JIRA_STATUS['Resolved'])
            mnrcametricsobj = MnRcaMetrics.query.get(RcaSubtaskJiraId)
            mnrcametricsobj.JiraIssueStatus = 'Resolved'
            db.session.commit()
        else:
            jira.transition_issue(RcaAanlysistaskIssue, 'RCA Done')
    # status = RcaAanlysistaskIssue.fields.status
    todo_item.PRRcaCompleteDate = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    todo_item.JiraIssueStatus = status
    todo_item.IsRcaCompleted = 'Yes'
    db.session.commit()
    # flash('AP item have been successfully imported RCA status also has been updated!')
    return True

class TodoAP(db.Model):
    __tablename__ = 'apstatus'
    APID = db.Column('APID', db.String(64), primary_key=True)
    PRID = db.Column(db.String(64))
    APDescription = db.Column(db.String(1024))
    APCreatedDate = db.Column(db.String(64))
    APDueDate = db.Column(db.String(64))
    APCompletedOn = db.Column(db.String(64))
    IsApCompleted = db.Column(db.String(32))
    APAssingnedTo = db.Column(db.String(128))
    QualityOwner = db.Column(db.String(128))

    # New added for JIRA
    InjectionRootCauseEdaCause = db.Column(db.String(1024))
    RcaEdaCauseType = db.Column(db.String(128))
    RcaEdaActionType = db.Column(db.String(128))
    TargetRelease = db.Column(db.String(128))
    CustomerAp = db.Column(db.String(32))
    ApCategory = db.Column(db.String(32))  # RCA/EDA
    ShouldHaveBeenDetected = db.Column(db.String(128))
    ApJiraId = db.Column(db.String(32))  # JIRA ID
    RcaEdaCauseCategory = db.Column(db.String(128))
    EvidenceOfCompleteness = db.Column(db.String(1024))
    rca_pronto_ap_id = db.Column(db.String(128))
    # End of added new field for JIRA


    def __init__(self, APID, PRID, APDescription, APCreatedDate, APDueDate, APCompletedOn, IsApCompleted, APAssingnedTo,
                 QualityOwner, \
                 InjectionRootCauseEdaCause, RcaEdaCauseType, RcaEdaActionType, TargetRelease, CustomerAp, ApCategory, \
                 ShouldHaveBeenDetected, ApJiraId, RcaEdaCauseCategory, rca_pronto_ap_id=''):
        self.APID = APID
        self.PRID = PRID
        self.APDescription = APDescription
        self.APCreatedDate = APCreatedDate
        self.APDueDate = APDueDate
        self.APCompletedOn = APCompletedOn
        self.IsApCompleted = IsApCompleted
        self.APAssingnedTo = APAssingnedTo
        self.QualityOwner = QualityOwner

        # New added field for JIRA
        self.InjectionRootCauseEdaCause = InjectionRootCauseEdaCause
        self.RcaEdaCauseType = RcaEdaCauseType   #Escape Cause Category
        self.RcaEdaActionType = RcaEdaActionType
        self.TargetRelease = TargetRelease
        self.CustomerAp = CustomerAp
        self.ApCategory = ApCategory  # RCA/EDA
        self.ShouldHaveBeenDetected = ShouldHaveBeenDetected
        self.ApJiraId = ApJiraId
        self.RcaEdaCauseCategory = RcaEdaCauseCategory   #Escape Cause Subcategory
        self.rca_pronto_ap_id = rca_pronto_ap_id
        # JIRA End

    # user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'))


def getLink(startAt,date):
    # RCALink = "https://jiradc.int.net.nokia.com/browse/{0}".format(RcaSubtaskJiraId)
    link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=%s'%(startAt,date)
    return link

def getNumberOfPRs():
    t = -45
    startAt = 0
    lastYearDate = (datetime.datetime.now() + datetime.timedelta(days=t)).strftime(
        "%Y-%m-%d")
    # link = 'https://pronto.int.net.nokia.com/prontoapi/rest/api/latest/problemReport?startAt=%d&maxResults=50&lastMod=2019-01-01'%startAt
    link = getLink(startAt, lastYearDate)
    url = link
    print('link is :'+str(link))
    r = requests.get(url, verify=False, auth=('krma', 'Nokia123'))
    r_status = r.status_code
    if r_status != 200:
        return 0,0
    a = json.loads(r.text)
    # a = r.json()
    total = a['total']
    prlist = a['values']

    return total,prlist


formposdict = {"datekey": ["completion_target_date"],
               "selectkey": ["injection_type", "triggering_scenario_category", "root_cause_category",
                             "root_cause_subcategory", "rca_action_type", "escape_cause_category",
                             "escape_cause_subcategory", "eda_action_type", "where_could_defect_have_been_detected"],
               "rcaformbase": {"product": {"3": "C"}, "issue_description": {"5": "D"}, "assessors": {"3": "H"},
                               "code_deficience": {"8": "D"}, "case_number": {"3": "B"},
                               "additional_facts": {"12": "D"},
                               "triggering_scenario_category": {"7": "D"}, "injection_type": {"10": "D"},
                               "inheritance_recommendation": {"13": "D"}, "how_many_times": {"14": "D"},
                               "quality_reviewer": {"3": "K"}, "abstract_headline": {"3": "D"},
                               "correction_description": {"9": "D"}, "triggering_scenario": {"6": "D"},
                               "injection_time": {"11": "D"}},
               "rcaformrootcause": {
                   "why_was_the_fault_introduced": {"row": 22, "span": 3,
                                                    "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9, "why5": 11,
                                                               "rootcause": 12, "actionproposal": 13,
                                                               "root_cause_category": 14,
                                                               "root_cause_subcategory": 15, "rca_action_type": 16,
                                                               "assigned_to": 17, "ap_jiratask_id": 18,
                                                               "completion_target_date": 19}},
                   "why_root_cause_was_not_found": {"row": 25, "span": 1,
                                                    "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9, "why5": 11,
                                                               "rootcause": 12, "actionproposal": 13,
                                                               "root_cause_category": 14,
                                                               "root_cause_subcategory": 15, "rca_action_type": 16,
                                                               "assigned_to": 17, "ap_jiratask_id": 18,
                                                               "completion_target_date": 19}},
                   "why_correction_took_longer_than_fct_target": {"row": 26, "span": 1,
                                                                  "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                             "why5": 11,
                                                                             "rootcause": 12, "actionproposal": 13,
                                                                             "root_cause_category": 14,
                                                                             "root_cause_subcategory": 15,
                                                                             "rca_action_type": 16,
                                                                             "assigned_to": 17, "ap_jiratask_id": 18,
                                                                             "completion_target_date": 19}}},
               "rcaformescapedefect": {"why_not_system_test": {"row": 59, "span": 4,
                                                               "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                          "why5": 11, "escape_cause": 12,
                                                                          "actionproposal": 13,
                                                                          "escape_cause_category": 14,
                                                                          "escape_cause_subcategory": 15,
                                                                          "eda_action_type": 16, "assigned_to": 17,
                                                                          "ap_jiratask_id": 18,
                                                                          "completion_target_date": 19,
                                                                          "where_could_defect_have_been_detected": 20}},
                                       "why_not_component_test": {"row": 50, "span": 3,
                                                                  "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                             "why5": 11, "escape_cause": 12,
                                                                             "actionproposal": 13,
                                                                             "escape_cause_category": 14,
                                                                             "escape_cause_subcategory": 15,
                                                                             "eda_action_type": 16, "assigned_to": 17,
                                                                             "ap_jiratask_id": 18,
                                                                             "completion_target_date": 19}},
                                       # "why_not_st_auto": {"row": 61, "span": 2,
                                       #                     "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                       #                                "why5": 11, "escape_cause": 12,
                                       #                                "actionproposal": 13, "escape_cause_category": 14,
                                       #                                "escape_cause_subcategory": 15,
                                       #                                "eda_action_type": 16, "assigned_to": 17,
                                       #                                "ap_jiratask_id": 18, "completion_target_date": 19,
                                       #                                "where_could_defect_have_been_detected": 20}},
                                       "why_not_requirements_review": {"row": 34, "span": 3,
                                                                       "keypos": {"why1": 3, "why2": 5, "why3": 7,
                                                                                  "why4": 9, "why5": 11,
                                                                                  "escape_cause": 12,
                                                                                  "actionproposal": 13,
                                                                                  "escape_cause_category": 14,
                                                                                  "escape_cause_subcategory": 15,
                                                                                  "eda_action_type": 16,
                                                                                  "assigned_to": 17,
                                                                                  "ap_jiratask_id": 18,
                                                                                  "completion_target_date": 19}},
                                       # "why_not_et_auto": {"row": 56, "span": 2,
                                       #                     "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                       #                                "why5": 11, "escape_cause": 12,
                                       #                                "actionproposal": 13, "escape_cause_category": 14,
                                       #                                "escape_cause_subcategory": 15,
                                       #                                "eda_action_type": 16, "assigned_to": 17,
                                       #                                "ap_jiratask_id": 18, "completion_target_date": 19,
                                       #                                "where_could_defect_have_been_detected": 20}},
                                       "why_not_inspections": {"row": 46, "span": 3,
                                                               "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                          "why5": 11, "escape_cause": 12,
                                                                          "actionproposal": 13,
                                                                          "escape_cause_category": 14,
                                                                          "escape_cause_subcategory": 15,
                                                                          "eda_action_type": 16, "assigned_to": 17,
                                                                          "ap_jiratask_id": 18,
                                                                          "completion_target_date": 19}},
                                       "why_not_design_review": {"row": 38, "span": 3,
                                                                 "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                            "why5": 11, "escape_cause": 12,
                                                                            "actionproposal": 13,
                                                                            "escape_cause_category": 14,
                                                                            "escape_cause_subcategory": 15,
                                                                            "eda_action_type": 16, "assigned_to": 17,
                                                                            "ap_jiratask_id": 18,
                                                                            "completion_target_date": 19}},
                                       "why_not_entity_test": {"row": 54, "span": 4,
                                                               "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                          "why5": 11, "escape_cause": 12,
                                                                          "actionproposal": 13,
                                                                          "escape_cause_category": 14,
                                                                          "escape_cause_subcategory": 15,
                                                                          "eda_action_type": 16, "assigned_to": 17,
                                                                          "ap_jiratask_id": 18,
                                                                          "completion_target_date": 19,
                                                                          "where_could_defect_have_been_detected": 20}},
                                       "why_not_analysis_tools": {"row": 42, "span": 3,
                                                                  "keypos": {"why1": 3, "why2": 5, "why3": 7, "why4": 9,
                                                                             "why5": 11, "escape_cause": 12,
                                                                             "actionproposal": 13,
                                                                             "escape_cause_category": 14,
                                                                             "escape_cause_subcategory": 15,
                                                                             "eda_action_type": 16, "assigned_to": 17,
                                                                             "ap_jiratask_id": 18,
                                                                             "completion_target_date": 19}},
                                       "why_customer_opened_a_ticket_on_a_known_defect": {"row": 64, "span": 2,
                                                                                          "keypos": {"why1": 3,
                                                                                                     "why2": 5,
                                                                                                     "why3": 7,
                                                                                                     "why4": 9,
                                                                                                     "why5": 11,
                                                                                                     "escape_cause": 12,
                                                                                                     "actionproposal": 13,
                                                                                                     "escape_cause_category": 14,
                                                                                                     "escape_cause_subcategory": 15,
                                                                                                     "eda_action_type": 16,
                                                                                                     "assigned_to": 17,
                                                                                                     "ap_jiratask_id": 18,
                                                                                                     "completion_target_date": 19}}},
               "lineorder": [("why_was_the_fault_introduced", "Why was the fault introduced?"),
                             ("why_root_cause_was_not_found",
                              "Why root cause was not found with first set of attached symptoms?"),
                             ("why_not_requirements_review", "1. Why didn't Requirements reviews catch this defect?"),
                             ("why_not_design_review", "2. Why didn't Design reviews catch this defect?"),
                             ("why_not_analysis_tools",
                              "3. Why didn't code analysis tools such as Klocwork, Purify, etc catch this defect?"),
                             ("why_not_inspections", "4. Why didn't code inspections catch this defect?"),
                             ("why_not_component_test", "5. Why didn't unit or component test catch this defect?"),
                             ("why_not_entity_test", "6. Why didn't Entity Test test catch this defect?"),
                             ("why_not_et_auto", "Why didn't ET automated test cases  catch this defect?"),
                             ("why_not_system_test", "7. Why didn't System Test catch this defect?"),
                             ("why_not_st_auto", "Why didn't ST automated test cases  catch this defect?")]}

def get_data_by_type_all():
    filter = []
    filter.append(RcaXtDm.yxbz == 'Y')
    # arr = RcaXtDm.query.order_by(RcaXtDm.dm.asc()).filter(*filter).all()
    # newarr = []
    # for i in arr:
    #     newarr.append(i.toJson())
    newarr = [{"type_id": '001', "type_name": '001',
                "dm": '001', "csz": '001', "p_type_id": '001',
                "p_dm": ''},
              {"type_id": '002', "type_name": '002',
               "dm": '002', "csz": '002', "p_type_id": '002',
               "p_dm": ''}]
    return newarr

def tranExcelPos(pos):
    newpos = []
    for k, v in pos.items():
        newpos.append(int(k))
        newpos.append(ord(v) - 64)
        break
    return newpos
# upload sharepoint file
def transactCharact(fddrcajson):
    if not fddrcajson:
        return ''
    elif type(fddrcajson) == 'str':
        return strip_non_ascii(fddrcajson)  # fddrcajson.replace('　',' ')
    elif type(fddrcajson) == 'date':
        return fddrcajson
    else:
        return strip_non_ascii(str(fddrcajson))  # fddrcajson.replace('　',' ')

def getXtDmByText(xtdmlist, text):
    if text == '':
        return ''
    for i in xtdmlist:
        if i["csz"].strip() == text.strip():
            return i["dm"].strip()

def adjustformdict(formposdict, sheet):
    # if all element in right position? adjust if not.
    # lineorder
    fddrcaobj = {"rcaformbase": {}}
    xtdmlist = get_data_by_type_all()
    try:
        for table in [("rcaformrootcause", formposdict["rcaformrootcause"]),
                      ("rcaformescapedefect", formposdict["rcaformescapedefect"])]:
            fddrcaobj[table[0]] = {}
            for i, j in table[1].items():
                fddrcaobj[table[0]][i] = []
                row = j["row"]
                for ix in range(1, 100):
                    pos = tranExcelPos({ix: 'B'})
                    cellvalue = sheet.cell(pos[0], pos[1]).value
                    dictkey = i[i.index(i)]
                    if cellvalue == i:
                        row = ix
                        print(str(cellvalue) + '   , ' + str(row) )
                span = j["span"]
                for r in range(0, span):
                    newline = {}
                    fddrcaobj[table[0]][i].append(newline)
                    rowvalid = True
                    for k, v in j["keypos"].items():
                        try:
                            cellvaluewhy1 = sheet.cell(row + r, j["keypos"]["why1"]).value
                            if cellvaluewhy1 and 'Why 1' in cellvaluewhy1:
                                rowvalid = False
                            if not rowvalid:
                                continue
                            cellvalue = sheet.cell(row + r, v).value
                            if k in formposdict["selectkey"]:
                                newline[k] = {"dm": getXtDmByText(xtdmlist,
                                                                  cellvalue if cellvalue else ''),
                                              "text": cellvalue if cellvalue else ''}
                            elif k in formposdict["datekey"]:
                                newline[k] = str(cellvalue).replace(' 00:00:00', '') if cellvalue else ''
                            else:
                                newline[k] = cellvalue.replace('_x000D_','') if cellvalue else ''
                        except:
                            print("%s, %s, %s" %(k,v,traceback.format_exc()))
    except Exception as e:
        print(str(e))
    return formposdict

def parxlstopojo(fpath):
    # parse file
    workbook = load_workbook(fpath)
    sheet = workbook['RcaEda']
    # format the formposdict
    localformposdict = formposdict
    localformposdict = adjustformdict(formposdict,sheet)
    return
    # generate pronto_json
    fddrcaobj = {"rcaformbase": {}}
    # dm list
    xtdmlist = get_data_by_type_all()
    case_number_pos = tranExcelPos(localformposdict["rcaformbase"]["case_number"])
    fddrcaobj["pronto_id"] = sheet.cell(case_number_pos[0], case_number_pos[1]).value
    for k, v in localformposdict["rcaformbase"].items():
        posxy = tranExcelPos(v)
        posx = posxy[0]
        posy = posxy[1]
        value = sheet.cell(posx, posy).value
        try:
            value = transactCharact(value)
        except Exception as e:
            value = value
            print(k + '  : ' + str(e))
        try:
            if k in localformposdict["selectkey"]:
                fddrcaobj["rcaformbase"][k] = {"dm": getXtDmByText(xtdmlist, value), "text": value}
            else:
                fddrcaobj["rcaformbase"][k] = value
        except Exception as e:
            print(e)

    try:
        for table in [("rcaformrootcause", localformposdict["rcaformrootcause"]),
                      ("rcaformescapedefect", localformposdict["rcaformescapedefect"])]:
            fddrcaobj[table[0]] = {}
            for i, j in table[1].items():
                fddrcaobj[table[0]][i] = []
                row = j["row"]
                span = j["span"]
                for r in range(0, span):
                    newline = {}
                    fddrcaobj[table[0]][i].append(newline)
                    rowvalid = True
                    for k, v in j["keypos"].items():
                        try:
                            cellvaluewhy1 = sheet.cell(row + r, j["keypos"]["why1"]).value
                            if cellvaluewhy1 and 'Why 1' in cellvaluewhy1:
                                rowvalid = False
                            if not rowvalid:
                                continue
                            cellvalue = sheet.cell(row + r, v).value
                            if k in localformposdict["selectkey"]:
                                newline[k] = {"dm": getXtDmByText(xtdmlist,
                                                                  cellvalue if cellvalue else ''),
                                              "text": cellvalue if cellvalue else ''}
                            elif k in localformposdict["datekey"]:
                                newline[k] = str(cellvalue).replace(' 00:00:00', '') if cellvalue else ''
                            else:
                                newline[k] = cellvalue.replace('_x000D_','') if cellvalue else ''
                        except:
                            print("%s, %s, %s" %(k,v,traceback.format_exc()))
    except Exception as e:
        print(str(e))
    return fddrcaobj


if __name__ == '__main__':
    # createParent5WhyRcaTaskFirst(getjira(), None, None, {'problemReportIds':['02359277','02357066']})
    #pr = APIPr('PR626937')
    #pass
    # for i in range(10):
    #     try:
    #         i / (i % 2)
    #     except BaseException as e1:
    #         errmes = 'error: PRID=' + str(i) + ', errorMes=' + str(e1) + 'RcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilledRcaFactsGetFilled'
    #         print(errmes)
    #         syslog_error('getCustomerPR_Thread', errormes=errmes)
    # # fa = APIFA(' FA718951')
    # testxls()
    # pass
    #initjira()
    # print(getjirains())
    # print(getjirains())
    # for i in range(10):
    #     j = threading.Thread(target=myjob, args=(str(i)))
    #     j.start()
    # print('end')

    # user = getDisplayNameByAccountId('mengxiang')
    # print(user)

    #batchPullEmail()
    # print(1)
    # print(2)
    # print(3)
    # print(4)
    # print(5)
    # try:
    #     raiseExe()
    # except Exception as e1:
    #     print(e1.message + "," + str(traceback.format_exc()))

    # conn = APIGetconn()
    # filter = '(mail=*%s*)' % 'men'
    # attrs = ['mail', 'displayName', 'uid']
    # base_dn = 'o=NSN'
    # flg = False
    # req_ctrl = SimplePagedResultsControl(size=30, cookie='')
    # result = conn.search_ext_s(base_dn, ldap.SCOPE_SUBTREE, filter, attrs, serverctrls=[req_ctrl])
    # print(len(result))
    # # RcaAanlysistaskIssue = getjirains().issue('MNRCA-48743')
    # # print(RcaAanlysistaskIssue)
    #
    # syslog_error('getCustomerPR_Thread', errormes='getCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_ThreadgetCustomerPR_Thread')

    # checkPR()
    # checkAP()
    # checkPRnotgenerate()
    # checkPRnotgeneratelist(['02408288','02399531','02357691'])


    # syslog_error('prontoweb_modify_fromexcelbyprid', errormes='123')
    # syslog_error('prontoweb_modify_fromexcelbyprid', errormes='1234', user='lmx',username='mengxiang.li@nokia-sbell.com')

    # checkJiraissue()
    # pullEmailByName('Wang, Fuqiang-Frank (NSB - CN/Hangzhou)')

    # LineManagerEmail, squadGroup, lineDisplayName, tribeLeadDisplayName, tribeName = getSquadTribeInfoWithConn(
    #     'fuqiang-frank.wang@nokia-sbell.com', get_ldap_connection())
    # print(LineManagerEmail)

    # LineManagerEmail, squadGroup, lineDisplayName, tribeLeadDisplayName2, tribeName1 = getSquadTribeInfo(
    #     'fuqiang-frank.wang@nokia-sbell.com')
    # print(LineManagerEmail)

    # syncjira('PR635549')
    # syncjira('PR636274')

    # testjiracomment()
    # testConnSingle()

    # testJiraGrade()

    # logging.debug('This message should appear on the console')
    # print('ok')

    #APIjiraissue('MNRCA-51002')
    # APIjiracreate()

    # json_data = "{'property': 'text'}"
    # json_data = ast.literal_eval(json_data)
    # print(json.dumps(json_data))

    # debugProntoGenerator('PR643292')

    # jsonformat()

    # json_data = '{"text": u"User Documentation (including Release Note) Improvement", "dm": u"06"}'
    # json_data2 = "{'text': u'User Documentation (including Release Note) Improvement', 'dm': u'06'}"
    # print(toolsGetDm(json_data,'text'))
    # print(toolsGetDm(json_data2,'text'))

    # print(toolsQAParser('Question: asdf Answer: dsafzzz'))
    # print(toolsQAGenerate('asdf','dsafzzz'))
    # print(toolsQAGenerate('asdf',''))
    # print(toolsQAGenerate('',''))
    # print(toolsQAGenerate('','dsafzzz'))
    # print(toolsQAParser(toolsQAGenerate('asdf','dsafzzz')))

    # testTodo()

    # todos = TodoAP.query.filter_by(APID='AP2022051817532585', IsApCompleted='No').first()
    # print(todos)

    # jira = getjira()
    # coms = jira.issue('MNRCA-46699')
    # if coms and coms.fields:
    #     if coms.fields.subtasks:
    #         for st in coms.fields.subtasks:
    #             print(st.key + ' , '+ str(st.fields.status))
    # print(coms)

    # print(getAccountIdByEmailName('yan-jack.chen@nokia-sbell.com'))
    # print(getAccountIdByEmailName('yaohong.h.chen@nokia-sbell.com'))
    # print(getAccountIdByEmailName('chunna.han@nokia-sbell.com'))

    # batchPullTribe()
    # getNumberOfPRs()

    # str1 = '[{"model": "mnrcahome.mnrcametricstable", "pk": "MNRCA-19246", "fields": {"JiraIssueParentTaskId": "MNRCA-19245", "JiraIssueSummary": "CAS-182423-Z9Q3:[Bharti mMIMO NPI] Sector beam ratio increased from 10% to 75% during high load (Kerala)", "JiraIssueType": "Analysis subtask", "JiraIssueCaseType": "RCA", "JiraIssueStatus": "Resolved", "JiraIssueParentStatus": "Closed", "JiraIssueAssignee": "Jiang, Zhilong (NSB - CN/Hangzhou)", "JiraIssueReporter": "Xia, Aaron (NSB - CN/Hangzhou)", "PRID": "CAS-182423-Z9Q3", "FAID": "FA449611", "PRTitle": "[Bharti mMIMO NPI] Sector beam ratio increased from 10% to 75% during high load (Kerala)", "PRRcaEdaAssessor": "zhilong.jiang@nokia-sbell.com", "PRAttached": "", "PRSeverity": "B - Major", "PRGroupInCharge": "NIHZSMAC", "JiraIssueCreatedDate": "2019-01-25", "JiraIssueParentCreatedDate": "2019-01-25", "JiraIssueResolutionDate": "2019-03-01", "JiraIssueParentResolutionDate": "2019-06-03", "JiraIssueDueDate": "2019-02-20", "JiraIssueOpenDays": 35, "JiraIssueParentOpenDays": 129, "JiraIssueAssigneeSquadGroup": "NSB MN RAN L2 SW 3 CN 1 SG", "JiraIssueAssigneeSquadGroupLead": "Zhang, Yijie (NSB - CN/Hangzhou)", "JiraIssueAssigneeTribe": "NSB MN RAN L2 SW 3 CN", "JiraIssueAssigneeTribeLead": "Chen, Zhuofei (NSB - CN/Hangzhou)", "JiraIssueOverDue": "Yes", "JiraIssueParentOverDue": "Yes", "JiraIssueOverDueReason": "", "JiraIssueParentOverDueReason": "", "ReportedBy": "Customer"}}, {"model": "mnrcahome.mnrcametricstable", "pk": "MNRCA-19248", "fields": {"JiraIssueParentTaskId": "MNRCA-19247", "JiraIssueSummary": "CAS-189164-N5K3:[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "JiraIssueType": "Analysis subtask", "JiraIssueCaseType": "RCA and EDA", "JiraIssueStatus": "Resolved", "JiraIssueParentStatus": "Resolved", "JiraIssueAssignee": "Jiang, Zhilong (NSB - CN/Hangzhou)", "JiraIssueReporter": "Xia, Aaron (NSB - CN/Hangzhou)", "PRID": "CAS-189164-N5K3", "FAID": "FA463847", "PRTitle": "[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "PRRcaEdaAssessor": "zhilong.jiang@nokia-sbell.com", "PRAttached": "", "PRSeverity": "B - Major", "PRGroupInCharge": "NIHZSMAC", "JiraIssueCreatedDate": "2019-01-25", "JiraIssueParentCreatedDate": "2019-01-25", "JiraIssueResolutionDate": "2019-02-22", "JiraIssueParentResolutionDate": "2019-03-06", "JiraIssueDueDate": "2019-02-22", "JiraIssueOpenDays": 28, "JiraIssueParentOpenDays": 40, "JiraIssueAssigneeSquadGroup": "NSB MN RAN L2 SW 3 CN 1 SG", "JiraIssueAssigneeSquadGroupLead": "Zhang, Yijie (NSB - CN/Hangzhou)", "JiraIssueAssigneeTribe": "NSB MN RAN L2 SW 3 CN", "JiraIssueAssigneeTribeLead": "Chen, Zhuofei (NSB - CN/Hangzhou)", "JiraIssueOverDue": "No", "JiraIssueParentOverDue": "Yes", "JiraIssueOverDueReason": "", "JiraIssueParentOverDueReason": "", "ReportedBy": "Customer"}}, {"model": "mnrcahome.mnrcametricstable", "pk": "MNRCA-19685", "fields": {"JiraIssueParentTaskId": "MNRCA-19247", "JiraIssueSummary": "CAS-189164-N5K3:[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "JiraIssueType": "Action for RCA", "JiraIssueCaseType": "", "JiraIssueStatus": "Resolved", "JiraIssueParentStatus": "Resolved", "JiraIssueAssignee": "Jiang, Zhilong (NSB - CN/Hangzhou)", "JiraIssueReporter": "Jiang, Zhilong (NSB - CN/Hangzhou)", "PRID": "CAS-189164-N5K3", "FAID": "FA463847", "PRTitle": "[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "PRRcaEdaAssessor": "zhilong.jiang@nokia-sbell.com", "PRAttached": "", "PRSeverity": "B - Major", "PRGroupInCharge": "NIHZSMAC", "JiraIssueCreatedDate": "2019-02-22", "JiraIssueParentCreatedDate": "2019-01-25", "JiraIssueResolutionDate": "2019-03-21", "JiraIssueParentResolutionDate": "2019-03-06", "JiraIssueDueDate": "2019-03-31", "JiraIssueOpenDays": 27, "JiraIssueParentOpenDays": 40, "JiraIssueAssigneeSquadGroup": "NSB MN RAN L2 SW 3 CN 1 SG", "JiraIssueAssigneeSquadGroupLead": "Zhang, Yijie (NSB - CN/Hangzhou)", "JiraIssueAssigneeTribe": "NSB MN RAN L2 SW 3 CN", "JiraIssueAssigneeTribeLead": "Chen, Zhuofei (NSB - CN/Hangzhou)", "JiraIssueOverDue": "No", "JiraIssueParentOverDue": "Yes", "JiraIssueOverDueReason": "", "JiraIssueParentOverDueReason": "", "ReportedBy": "Customer"}}, {"model": "mnrcahome.mnrcametricstable", "pk": "MNRCA-19686", "fields": {"JiraIssueParentTaskId": "MNRCA-19247", "JiraIssueSummary": "CAS-189164-N5K3:[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "JiraIssueType": "Action for RCA", "JiraIssueCaseType": "", "JiraIssueStatus": "Resolved", "JiraIssueParentStatus": "Resolved", "JiraIssueAssignee": "Jiang, Zhilong (NSB - CN/Hangzhou)", "JiraIssueReporter": "Jiang, Zhilong (NSB - CN/Hangzhou)", "PRID": "CAS-189164-N5K3", "FAID": "FA463847", "PRTitle": "[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "PRRcaEdaAssessor": "zhilong.jiang@nokia-sbell.com", "PRAttached": "", "PRSeverity": "B - Major", "PRGroupInCharge": "NIHZSMAC", "JiraIssueCreatedDate": "2019-02-22", "JiraIssueParentCreatedDate": "2019-01-25", "JiraIssueResolutionDate": "2019-04-24", "JiraIssueParentResolutionDate": "2019-03-06", "JiraIssueDueDate": "2019-04-30", "JiraIssueOpenDays": 61, "JiraIssueParentOpenDays": 40, "JiraIssueAssigneeSquadGroup": "NSB MN RAN L2 SW 3 CN 1 SG", "JiraIssueAssigneeSquadGroupLead": "Zhang, Yijie (NSB - CN/Hangzhou)", "JiraIssueAssigneeTribe": "NSB MN RAN L2 SW 3 CN", "JiraIssueAssigneeTribeLead": "Chen, Zhuofei (NSB - CN/Hangzhou)", "JiraIssueOverDue": "No", "JiraIssueParentOverDue": "Yes", "JiraIssueOverDueReason": "", "JiraIssueParentOverDueReason": "", "ReportedBy": "Customer"}}, {"model": "mnrcahome.mnrcametricstable", "pk": "MNRCA-19687", "fields": {"JiraIssueParentTaskId": "MNRCA-19247", "JiraIssueSummary": "CAS-189164-N5K3:[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "JiraIssueType": "Action for RCA", "JiraIssueCaseType": "", "JiraIssueStatus": "Resolved", "JiraIssueParentStatus": "Resolved", "JiraIssueAssignee": "Jiang, Zhilong (NSB - CN/Hangzhou)", "JiraIssueReporter": "Jiang, Zhilong (NSB - CN/Hangzhou)", "PRID": "CAS-189164-N5K3", "FAID": "FA463847", "PRTitle": "[VF mMIMO TL18SP NPI] [KPI] mMIMO sites are throughput dip randomly", "PRRcaEdaAssessor": "zhilong.jiang@nokia-sbell.com", "PRAttached": "", "PRSeverity": "B - Major", "PRGroupInCharge": "NIHZSMAC", "JiraIssueCreatedDate": "2019-02-22", "JiraIssueParentCreatedDate": "2019-01-25", "JiraIssueResolutionDate": "2019-03-28", "JiraIssueParentResolutionDate": "2019-03-06", "JiraIssueDueDate": "2019-03-31", "JiraIssueOpenDays": 34, "JiraIssueParentOpenDays": 40, "JiraIssueAssigneeSquadGroup": "NSB MN RAN L2 SW 3 CN 1 SG", "JiraIssueAssigneeSquadGroupLead": "Zhang, Yijie (NSB - CN/Hangzhou)", "JiraIssueAssigneeTribe": "NSB MN RAN L2 SW 3 CN", "JiraIssueAssigneeTribeLead": "Chen, Zhuofei (NSB - CN/Hangzhou)", "JiraIssueOverDue": "No", "JiraIssueParentOverDue": "Yes", "JiraIssueOverDueReason": "", "JiraIssueParentOverDueReason": "", "ReportedBy": "Customer"}}]'
    # print(json.loads(str1))

    parxlstopojo('C:\\Work\\templates\\RCA_EDA_Analysis_Template_LTE_BL-mode.xlsx')

    # stra = '123_asdf_dfsk'
    # print('rfind:'+str(stra.rfind('g')))
    # print('find:'+str(stra.find('g')))
    # print('index:'+str(stra.rindex('g',-5)))
    # print('index:'+str(stra.index('g',-5)))





